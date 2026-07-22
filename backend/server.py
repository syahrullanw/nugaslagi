from __future__ import annotations

import io
import asyncio
import base64
import gzip
import hashlib
import html
import json
import logging
import os
import re
import secrets
import time
import uuid
import zipfile
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any, Dict, List, Optional
from zoneinfo import ZoneInfo
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit
from xml.etree import ElementTree

import bcrypt
from cryptography.fernet import Fernet, InvalidToken
from cryptography.exceptions import InvalidSignature
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import padding, rsa
from dotenv import load_dotenv
import httpx
from fastapi import APIRouter, BackgroundTasks, Depends, FastAPI, File, Form, Header, HTTPException, Request, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from google.auth.transport.requests import Request as GoogleAuthRequest
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseUpload
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, EmailStr, Field
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.gzip import GZipMiddleware

try:
    from .postgres_database import PostgresDatabase
    from .app_version import version_payload
except ImportError:  # Supports `uvicorn server:app` from the backend directory.
    from postgres_database import PostgresDatabase
    from app_version import version_payload


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

db = PostgresDatabase(os.environ["DATABASE_URL"])

app = FastAPI(title="E-Learning Dosen API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)
_whatsapp_send_lock = asyncio.Lock()
_material_creation_lock = asyncio.Lock()

_settings_cache: Dict[str, Any] = {}
_settings_cache_times: Dict[str, float] = {}
_SETTINGS_CACHE_TTL = 30.0
_oidc_discovery_cache: Dict[str, Any] = {}
_oidc_discovery_cached_at = 0.0
_OIDC_DISCOVERY_CACHE_TTL = 300.0
_oidc_runtime_settings: Dict[str, Any] = {}


def env_enabled(name: str, default: bool = False) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def environment_oidc_settings() -> Dict[str, Any]:
    discovery_url = os.environ.get("OIDC_DISCOVERY_URL", "").strip()
    client_id = os.environ.get("OIDC_CLIENT_ID", "").strip()
    redirect_uri = os.environ.get("OIDC_REDIRECT_URI", "").strip()
    frontend_url = os.environ.get("OIDC_FRONTEND_URL", os.environ.get("APP_URL", "http://127.0.0.1:3000")).strip().rstrip("/")
    return {
        "enabled": env_enabled("OIDC_ENABLED", bool(discovery_url and client_id and redirect_uri)),
        "discovery_url": discovery_url,
        "issuer": os.environ.get("OIDC_ISSUER", "").strip().rstrip("/"),
        "client_id": client_id,
        "client_secret": os.environ.get("OIDC_CLIENT_SECRET", "").strip(),
        "redirect_uri": redirect_uri,
        "frontend_url": frontend_url,
        "scopes": os.environ.get("OIDC_SCOPES", "openid profile email roles").strip(),
        "local_login_enabled": env_enabled("OIDC_LOCAL_LOGIN_ENABLED", True),
    }


def oidc_settings() -> Dict[str, Any]:
    return {**environment_oidc_settings(), **_oidc_runtime_settings}


def clear_oidc_discovery_cache() -> None:
    global _oidc_discovery_cache, _oidc_discovery_cached_at
    _oidc_discovery_cache = {}
    _oidc_discovery_cached_at = 0.0


def oidc_require_settings() -> Dict[str, Any]:
    settings = oidc_settings()
    missing = [key for key in ["discovery_url", "client_id", "redirect_uri", "frontend_url"] if not settings.get(key)]
    if not settings["enabled"] or missing:
        raise HTTPException(status_code=503, detail="Login SCI-ID belum dikonfigurasi")
    return settings


def _get_cached_settings(key: str) -> Optional[Dict[str, Any]]:
    now = time.time()
    if key in _settings_cache and (now - _settings_cache_times.get(key, 0)) < _SETTINGS_CACHE_TTL:
        return _settings_cache[key]
    return None


def _set_cached_settings(key: str, value: Dict[str, Any]) -> None:
    _settings_cache[key] = value
    _settings_cache_times[key] = time.time()


def _invalidate_settings_cache(key: str = "") -> None:
    if key:
        _settings_cache.pop(key, None)
        _settings_cache_times.pop(key, None)
    else:
        _settings_cache.clear()
        _settings_cache_times.clear()


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def base64url_decode(value: str) -> bytes:
    return base64.urlsafe_b64decode(value + ("=" * (-len(value) % 4)))


def base64url_encode(value: bytes) -> str:
    return base64.urlsafe_b64encode(value).decode("ascii").rstrip("=")


async def oidc_discovery(force: bool = False) -> Dict[str, Any]:
    global _oidc_discovery_cache, _oidc_discovery_cached_at
    settings = oidc_require_settings()
    if not force and _oidc_discovery_cache and time.time() - _oidc_discovery_cached_at < _OIDC_DISCOVERY_CACHE_TTL:
        return _oidc_discovery_cache
    try:
        async with httpx.AsyncClient(timeout=10) as http:
            response = await http.get(settings["discovery_url"], headers={"Accept": "application/json"})
            response.raise_for_status()
            metadata = response.json()
    except (httpx.HTTPError, ValueError) as exc:
        logger.warning("OIDC discovery SCI-ID gagal: %s", exc)
        raise HTTPException(status_code=503, detail="SCI-ID sedang tidak dapat dihubungi") from exc
    required = ["issuer", "authorization_endpoint", "token_endpoint", "jwks_uri", "end_session_endpoint"]
    if not isinstance(metadata, dict) or any(not metadata.get(key) for key in required):
        raise HTTPException(status_code=503, detail="Metadata OIDC SCI-ID tidak lengkap")
    expected_issuer = settings["issuer"] or metadata["issuer"]
    if metadata["issuer"].rstrip("/") != expected_issuer.rstrip("/"):
        raise HTTPException(status_code=503, detail="Issuer OIDC SCI-ID tidak sesuai konfigurasi")
    _oidc_discovery_cache = metadata
    _oidc_discovery_cached_at = time.time()
    return metadata


def oidc_roles(claims: Dict[str, Any], client_id: str) -> List[str]:
    roles: set[str] = set()
    realm_access = claims.get("realm_access")
    if isinstance(realm_access, dict) and isinstance(realm_access.get("roles"), list):
        roles.update(str(role) for role in realm_access["roles"])
    resource_access = claims.get("resource_access")
    client_access = resource_access.get(client_id) if isinstance(resource_access, dict) else None
    if isinstance(client_access, dict) and isinstance(client_access.get("roles"), list):
        roles.update(str(role) for role in client_access["roles"])
    return sorted(roles)


def oidc_application_role(claims: Dict[str, Any], client_id: str) -> str:
    roles = set(oidc_roles(claims, client_id))
    if "super_admin" in roles:
        return "admin"
    if "dosen" in roles:
        return "lecturer"
    if "mahasiswa" in roles:
        return "student"
    raise HTTPException(status_code=403, detail="Akun SCI-ID belum memiliki role dosen, mahasiswa, atau super_admin")


async def validate_oidc_id_token(id_token: str, metadata: Dict[str, Any], expected_nonce: str) -> Dict[str, Any]:
    settings = oidc_require_settings()
    parts = id_token.split(".")
    if len(parts) != 3:
        raise HTTPException(status_code=401, detail="Format ID token SCI-ID tidak valid")
    try:
        header = json.loads(base64url_decode(parts[0]))
        claims = json.loads(base64url_decode(parts[1]))
        signature = base64url_decode(parts[2])
    except (ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=401, detail="ID token SCI-ID tidak dapat dibaca") from exc
    if header.get("alg") != "RS256" or not header.get("kid"):
        raise HTTPException(status_code=401, detail="Algoritma ID token SCI-ID tidak diizinkan")
    try:
        async with httpx.AsyncClient(timeout=10) as http:
            jwks_response = await http.get(metadata["jwks_uri"], headers={"Accept": "application/json"})
            jwks_response.raise_for_status()
            jwks = jwks_response.json()
    except (httpx.HTTPError, ValueError) as exc:
        raise HTTPException(status_code=503, detail="Kunci verifikasi SCI-ID tidak dapat dimuat") from exc
    key = next((item for item in jwks.get("keys", []) if item.get("kid") == header["kid"] and item.get("kty") == "RSA"), None)
    if not key or not key.get("n") or not key.get("e"):
        raise HTTPException(status_code=401, detail="Kunci ID token SCI-ID tidak ditemukan")
    try:
        public_key = rsa.RSAPublicNumbers(
            int.from_bytes(base64url_decode(key["e"]), "big"),
            int.from_bytes(base64url_decode(key["n"]), "big"),
        ).public_key()
        public_key.verify(signature, f"{parts[0]}.{parts[1]}".encode("ascii"), padding.PKCS1v15(), hashes.SHA256())
    except (InvalidSignature, ValueError) as exc:
        raise HTTPException(status_code=401, detail="Signature ID token SCI-ID tidak valid") from exc
    now = int(time.time())
    issuer = settings["issuer"] or metadata["issuer"]
    audiences = claims.get("aud") if isinstance(claims.get("aud"), list) else [claims.get("aud")]
    if claims.get("iss", "").rstrip("/") != issuer.rstrip("/"):
        raise HTTPException(status_code=401, detail="Issuer ID token SCI-ID tidak sesuai")
    if settings["client_id"] not in audiences:
        raise HTTPException(status_code=401, detail="Audience ID token SCI-ID tidak sesuai")
    if claims.get("azp") and claims["azp"] != settings["client_id"]:
        raise HTTPException(status_code=401, detail="Authorized party ID token SCI-ID tidak sesuai")
    if not isinstance(claims.get("exp"), (int, float)) or now - 30 >= int(claims["exp"]):
        raise HTTPException(status_code=401, detail="ID token SCI-ID sudah kedaluwarsa")
    if isinstance(claims.get("nbf"), (int, float)) and now + 30 < int(claims["nbf"]):
        raise HTTPException(status_code=401, detail="ID token SCI-ID belum berlaku")
    if isinstance(claims.get("iat"), (int, float)) and int(claims["iat"]) > now + 60:
        raise HTTPException(status_code=401, detail="Waktu penerbitan ID token SCI-ID tidak valid")
    if not claims.get("nonce") or not secrets.compare_digest(str(claims["nonce"]), expected_nonce):
        raise HTTPException(status_code=401, detail="Nonce login SCI-ID tidak sesuai")
    if not str(claims.get("sub", "")).strip():
        raise HTTPException(status_code=401, detail="Subject ID token SCI-ID tidak valid")
    return claims


def oidc_frontend_redirect(**params: str) -> str:
    base = oidc_require_settings()["frontend_url"]
    clean = {key: value for key, value in params.items() if value}
    return f"{base}/?{urlencode(clean)}" if clean else base


async def provision_oidc_user(claims: Dict[str, Any]) -> Dict[str, Any]:
    settings = oidc_require_settings()
    issuer = settings["issuer"] or str(claims["iss"])
    subject = str(claims["sub"])
    role = oidc_application_role(claims, settings["client_id"])
    username = str(claims.get("preferred_username") or f"sso-{subject[:12]}").strip().lower()
    email = str(claims.get("email") or f"sso-{subject}@local.invalid").strip().lower()
    name = str(claims.get("name") or username).strip()
    linked = await db.users.find_one({"sso_issuer": issuer, "sso_subject": subject}, {"_id": 0})
    if not linked:
        matches = await db.users.find({"$or": [{"email": email}, {"username": username}]}, {"_id": 0}).to_list(3)
        unique_matches = {item["id"]: item for item in matches}
        if len(unique_matches) > 1:
            raise HTTPException(status_code=409, detail="Email dan username SCI-ID terhubung ke akun lokal yang berbeda")
        linked = next(iter(unique_matches.values()), None)
    role_claims = oidc_roles(claims, settings["client_id"])
    synced = {
        "role": role,
        "name": name,
        "sso_issuer": issuer,
        "sso_subject": subject,
        "sso_roles": role_claims,
        "auth_source": "sso",
        "status": "active",
        "last_login_at": now_iso(),
        "sso_synced_at": now_iso(),
    }
    employee_id = str(claims.get("employee_id") or claims.get("nidn") or claims.get("nip") or "").strip()
    nim = str(claims.get("nim") or "").strip()
    if employee_id:
        synced["employee_id"] = employee_id
    if nim:
        synced["nim"] = nim
    if linked:
        await db.users.update_one({"id": linked["id"]}, {"$set": synced})
        user = await db.users.find_one({"id": linked["id"]}, {"_id": 0})
    else:
        candidate = username
        if await db.users.find_one({"username": candidate}, {"_id": 0, "id": 1}):
            candidate = f"{candidate}-{subject[:6]}"
        doc = {
            "id": new_id(),
            "username": candidate,
            "email": email,
            "whatsapp": str(claims.get("phone_number") or ""),
            "password_hash": hash_password(secrets.token_urlsafe(32)),
            "class_ids": [],
            "created_at": now_iso(),
            **synced,
        }
        await db.users.insert_one(doc)
        user = doc
    return public_doc(user)


def parse_iso_datetime(value: str) -> Optional[datetime]:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)
    except Exception:
        return None


def normalize_optional_datetime(value: str, field_label: str) -> str:
    if not value:
        return ""
    parsed = parse_iso_datetime(value)
    if not parsed:
        raise HTTPException(status_code=400, detail=f"{field_label} tidak valid")
    return parsed.isoformat()


def assignment_is_published(assignment: Dict[str, Any], reference: Optional[datetime] = None) -> bool:
    published_at = parse_iso_datetime(assignment.get("published_at", ""))
    if not published_at:
        return True
    return published_at <= (reference or datetime.now(timezone.utc))


def assignment_publish_status(assignment: Dict[str, Any]) -> str:
    return "published" if assignment_is_published(assignment) else "scheduled"


def new_id() -> str:
    return str(uuid.uuid4())


def clean_code(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", value.upper())[:10]


def public_doc(doc: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not doc:
        return None
    doc.pop("_id", None)
    doc.pop("password_hash", None)
    doc.pop("local_path", None)
    doc.pop("drive_error", None)
    doc.pop("drive_hierarchy", None)
    return doc


STORAGE_ROOT = ROOT_DIR / "storage" / "E-Learning Dosen"
DEFAULT_SUBMISSION_MAX_FILE_MB = 5


def safe_path_segment(value: str) -> str:
    cleaned = re.sub(r"[\\/]+", "-", str(value or "").strip())
    cleaned = re.sub(r"[^A-Za-z0-9._ \-()]+", "-", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .-")
    return cleaned[:100] or "Tanpa Nama"


def build_local_file_path(
    hierarchy: List[str], student_nim: str, student_name: str, file_token: str, original_filename: str
) -> tuple[Path, str, str]:
    safe_hierarchy = [safe_path_segment(item) for item in hierarchy]
    student_folder = safe_path_segment(f"{student_nim or 'NO-NIM'} - {student_name}")
    safe_filename = safe_path_segment(original_filename or "submission.bin")
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    final_filename = f"{timestamp}_{file_token[:8]}_{safe_filename}"
    folder = STORAGE_ROOT.joinpath(*safe_hierarchy, student_folder)
    absolute_path = folder / final_filename
    relative_folder = str(Path("E-Learning Dosen").joinpath(*safe_hierarchy, student_folder))
    relative_file = str(Path(relative_folder) / final_filename)
    return absolute_path, relative_folder, relative_file


def local_file_urls(file_id: str) -> Dict[str, str]:
    return {
        "file_url": f"/api/files/{file_id}/download",
        "preview_url": f"/api/files/{file_id}/preview",
        "inline_url": f"/api/files/{file_id}/inline",
    }


async def multipart_uploads(request: Request, fields: List[str]) -> tuple[Any, List[UploadFile]]:
    form = await request.form()
    uploads: List[UploadFile] = []
    for field in fields:
        for item in form.getlist(field):
            if getattr(item, "filename", "") and hasattr(item, "read"):
                uploads.append(item)
    return form, uploads


def assignment_max_file_size_mb(assignment: Dict[str, Any]) -> float:
    raw_value = assignment.get("max_file_size_mb", assignment.get("max_submission_size_mb", DEFAULT_SUBMISSION_MAX_FILE_MB))
    try:
        value = float(raw_value)
    except (TypeError, ValueError):
        value = DEFAULT_SUBMISSION_MAX_FILE_MB
    return value if value > 0 else DEFAULT_SUBMISSION_MAX_FILE_MB


async def validate_upload_file_sizes(uploads: List[UploadFile], max_mb: float, label: str) -> None:
    max_bytes = int(max_mb * 1024 * 1024)
    for upload in uploads:
        file_size = 0
        while True:
            chunk = await upload.read(1024 * 1024)
            if not chunk:
                break
            file_size += len(chunk)
            if file_size > max_bytes:
                raise HTTPException(status_code=400, detail=f"Ukuran file {upload.filename or label} maksimal {max_mb:g} MB")
        await upload.seek(0)


def enrich_file_urls(file_doc: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not file_doc:
        return file_doc
    file_id = file_doc.get("file_id") or file_doc.get("id")
    if file_id:
        file_doc.update(local_file_urls(file_id))
    return file_doc


def enrich_submission_file_urls(submission: Dict[str, Any]) -> Dict[str, Any]:
    if isinstance(submission.get("file"), dict):
        submission["file"] = enrich_file_urls(submission["file"])
    if isinstance(submission.get("files"), list):
        submission["files"] = [enrich_file_urls(item) if isinstance(item, dict) else item for item in submission["files"]]
    return submission


def mask_secret(value: str) -> str:
    if not value:
        return ""
    if len(value) <= 6:
        return "******"
    return f"{value[:3]}***{value[-3:]}"


def normalize_phone(value: str) -> str:
    cleaned = re.sub(r"\D+", "", value or "")
    if cleaned.startswith("0"):
        cleaned = "62" + cleaned[1:]
    return cleaned


def normalize_http_base_url(value: str) -> str:
    base_url = (value or "").strip()
    if not base_url:
        return ""
    if not re.match(r"^https?://", base_url, flags=re.IGNORECASE):
        base_url = f"http://{base_url}"
    base_url = base_url.rstrip("/")
    if base_url.lower().endswith("/api"):
        base_url = base_url[:-4].rstrip("/")
    return base_url


def normalize_app_url(value: str) -> str:
    app_url = (value or "").strip()
    if not app_url or "domain-aplikasi" in app_url.lower():
        app_url = os.environ.get("APP_URL", "").strip()
    if not app_url:
        return ""
    if not re.match(r"^https?://", app_url, flags=re.IGNORECASE):
        if app_url.startswith(("localhost", "127.", "10.", "192.168.")) or re.match(r"^\d{1,3}(\.\d{1,3}){3}", app_url):
            app_url = f"http://{app_url}"
        else:
            app_url = f"https://{app_url}"
    return app_url.rstrip("/")


def build_password_reset_link(app_url: str, identifier: str) -> str:
    base_url = normalize_app_url(app_url)
    if not base_url:
        return "Buka halaman aplikasi dan pilih Lupa Password"
    parts = urlsplit(base_url)
    query = dict(parse_qsl(parts.query, keep_blank_values=True))
    query.update({"reset": "password", "identifier": identifier})
    path = parts.path or "/"
    return urlunsplit((parts.scheme, parts.netloc, path, urlencode(query), parts.fragment))


def build_app_fragment_link(app_url: str, fragment: str = "") -> str:
    base_url = normalize_app_url(app_url)
    if not base_url:
        return "Buka halaman aplikasi"
    parts = urlsplit(base_url)
    return urlunsplit((parts.scheme, parts.netloc, parts.path or "/", parts.query, fragment.lstrip("#")))


def waha_chat_id(phone_or_chat_id: str) -> str:
    raw = (phone_or_chat_id or "").strip()
    if raw.endswith(("@c.us", "@g.us", "@newsletter")):
        return raw
    phone = normalize_phone(raw)
    if not phone:
        raise RuntimeError("Nomor tujuan WhatsApp kosong")
    return f"{phone}@c.us"


def waha_headers(api_key: str = "") -> Dict[str, str]:
    headers = {"Accept": "application/json", "Content-Type": "application/json"}
    api_key = (api_key or "").strip()
    if api_key:
        headers["X-Api-Key"] = api_key
    return headers


def response_excerpt(response: httpx.Response) -> str:
    try:
        data = response.json()
        return json.dumps(data, ensure_ascii=False)[:600]
    except Exception:
        return response.text[:600]


def file_extension(file_doc: Dict[str, Any], path: Optional[Path] = None) -> str:
    source = file_doc.get("file_name") or file_doc.get("storage_path") or ""
    if path:
        source = source or path.name
    return source.rsplit(".", 1)[-1].lower() if "." in source else ""


def preview_kind(file_doc: Dict[str, Any], path: Path) -> str:
    mime = (file_doc.get("mime_type") or "").lower()
    ext = file_extension(file_doc, path)
    if mime == "application/pdf" or ext == "pdf":
        return "pdf"
    if mime.startswith("image/") or ext in {"png", "jpg", "jpeg", "webp", "gif"}:
        return "image"
    if ext in {"docx"}:
        return "docx"
    if ext in {"xlsx", "xlsm"}:
        return "xlsx"
    if ext in {"txt", "md", "csv", "json"} or mime.startswith("text/"):
        return "text"
    return "unsupported"


def html_panel(title: str, body: str) -> str:
    return f"<section class=\"doc-preview-section\"><h3>{html.escape(title)}</h3>{body}</section>"


def preview_docx_html(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        xml_data = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml_data)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    blocks: List[str] = []
    for paragraph in root.findall(".//w:p", namespace):
        texts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        line = "".join(texts).strip()
        if line:
            blocks.append(f"<p>{html.escape(line)}</p>")
    if not blocks:
        blocks.append("<p class=\"doc-preview-muted\">Dokumen tidak memiliki teks yang bisa diekstrak.</p>")
    return html_panel(path.name, "".join(blocks[:1200]))


def preview_xlsx_html(path: Path) -> str:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sections: List[str] = []
    try:
        for sheet in workbook.worksheets[:3]:
            rows_html: List[str] = []
            for row_index, row in enumerate(sheet.iter_rows(max_row=80, max_col=20, values_only=True), start=1):
                cells = ["" if value is None else str(value) for value in row]
                if not any(cells):
                    continue
                cell_tag = "th" if row_index == 1 else "td"
                row_html = "".join(f"<{cell_tag}>{html.escape(value)}</{cell_tag}>" for value in cells)
                rows_html.append(f"<tr>{row_html}</tr>")
            table = "<table>" + "".join(rows_html) + "</table>" if rows_html else "<p class=\"doc-preview-muted\">Sheet kosong.</p>"
            sections.append(f"<h3>{html.escape(sheet.title)}</h3>{table}")
    finally:
        workbook.close()
    return "<section class=\"doc-preview-section\">" + "".join(sections) + "</section>"


def preview_text_html(path: Path) -> str:
    content = path.read_text(encoding="utf-8", errors="replace")[:120000]
    return html_panel(path.name, f"<pre>{html.escape(content)}</pre>")


async def fetch_waha_session_status(client: httpx.AsyncClient, base_url: str, session: str, headers: Dict[str, str]) -> Optional[Dict[str, Any]]:
    response = await client.get(f"{base_url}/api/sessions/{session}", headers=headers)
    if response.status_code == 404:
        return None
    if response.status_code >= 400:
        raise RuntimeError(f"WAHA session check gagal ({response.status_code}): {response_excerpt(response)}")
    return response.json()


async def ensure_waha_session_working(client: httpx.AsyncClient, base_url: str, session: str, headers: Dict[str, str]) -> None:
    session_doc = await fetch_waha_session_status(client, base_url, session, headers)
    if session_doc is None:
        raise RuntimeError(f"Sesi WAHA '{session}' tidak ditemukan. Buat/start session di WAHA lalu scan QR.")
    status = str(session_doc.get("status") or "").upper()
    if status and status != "WORKING":
        raise RuntimeError(f"Sesi WAHA '{session}' belum WORKING (status: {status}). Start session dan scan QR di dashboard WAHA.")


def generate_otp() -> str:
    return "".join(str(secrets.randbelow(10)) for _ in range(6))


def local_reset_otp_enabled() -> bool:
    return os.environ.get("ALLOW_LOCAL_RESET_OTP", "").lower() in {"1", "true", "yes", "on"}


def identity_query(identifier: str) -> Dict[str, Any]:
    raw = (identifier or "").strip()
    lowered = raw.lower()
    normalized_phone = normalize_phone(raw)
    candidates: List[Dict[str, str]] = [
        {"email": lowered},
        {"username": lowered},
        {"employee_id": raw},
        {"nim": raw.upper()},
        {"nim": raw},
        {"whatsapp": raw},
    ]
    if normalized_phone and normalized_phone != raw:
        candidates.append({"whatsapp": normalized_phone})
    return {"$or": candidates}


DEFAULT_GRADE_PREDICATES = [
    {"label": "A", "min_score": 85, "max_score": 100},
    {"label": "B", "min_score": 70, "max_score": 84.99},
    {"label": "C", "min_score": 60, "max_score": 69.99},
    {"label": "D", "min_score": 50, "max_score": 59.99},
    {"label": "E", "min_score": 0, "max_score": 49.99},
]


def default_whatsapp_settings() -> Dict[str, Any]:
    return {
        "id": "main",
        "provider": "disabled",
        "app_url": os.environ.get("APP_URL", ""),
        "fonnte_token": "",
        "fonnte_url": "https://api.fonnte.com/send",
        "waha_base_url": "",
        "waha_api_key": "",
        "waha_session": "default",
        "send_delay_seconds": 3,
        "typing_simulation_seconds": 30,
        "otp_template": "Kode OTP reset password Anda: {code}. Berlaku {minutes} menit. Link: {link}",
        "assignment_template": "Halo {name}, ada tugas baru: {title}. Kelas: {class_name}. Deadline: {deadline}. Link: {link}",
        "grade_template": "Halo {name}, tugas {title} sudah dinilai. Nilai: {grade} ({predicate}). Feedback: {feedback}. Link: {link}",
        "revision_template": "Halo {name}, tugas {title} perlu revisi. Catatan: {revision_note}. Link: {link}",
    }


def default_email_settings() -> Dict[str, Any]:
    return {
        "id": "main",
        "enabled": False,
        "smtp_host": os.environ.get("SMTP_HOST", ""),
        "smtp_port": int(os.environ.get("SMTP_PORT", "587")),
        "smtp_user": os.environ.get("SMTP_USER", ""),
        "smtp_password": os.environ.get("SMTP_PASSWORD", ""),
        "smtp_use_tls": os.environ.get("SMTP_USE_TLS", "true").lower() in {"1", "true", "yes", "on"},
        "from_name": os.environ.get("SMTP_FROM_NAME", "E-Learning Dosen"),
        "from_email": os.environ.get("SMTP_FROM_EMAIL", ""),
    }


async def get_email_settings(mask: bool = False) -> Dict[str, Any]:
    cached = _get_cached_settings("email_settings")
    if cached is None:
        settings = await db.email_settings.find_one({"id": "main"}, {"_id": 0})
        cached = {**default_email_settings(), **(settings or {})}
        _set_cached_settings("email_settings", cached)
    if mask:
        settings = cached.copy()
        settings["smtp_password_masked"] = mask_secret(settings.get("smtp_password", ""))
        settings.pop("smtp_password", None)
        return settings
    return cached


async def send_email_message(to_email: str, subject: str, html_body: str) -> Dict[str, Any]:
    settings = await get_email_settings(mask=False)
    if not settings.get("enabled"):
        return {"ok": False, "error": "Email belum diaktifkan"}
    host = settings.get("smtp_host", "")
    port = int(settings.get("smtp_port", 587))
    user = settings.get("smtp_user", "")
    password = settings.get("smtp_password", "")
    use_tls = settings.get("smtp_use_tls", True)
    from_name = settings.get("from_name", "E-Learning Dosen")
    from_email = settings.get("from_email", "") or user
    if not host or not from_email:
        return {"ok": False, "error": "SMTP host atau email pengirim belum dikonfigurasi"}
    try:
        msg = MIMEMultipart("alternative")
        msg["From"] = f"{from_name} <{from_email}>"
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(html_body, "html", "utf-8"))
        if port == 465:
            server = smtplib.SMTP_SSL(host, port, timeout=15)
        else:
            server = smtplib.SMTP(host, port, timeout=15)
            server.ehlo()
            if use_tls:
                try:
                    server.starttls()
                    server.ehlo()
                except smtplib.SMTPNotSupportedError:
                    pass
        auth_capable = server.has_extn("AUTH")
        if user and password and auth_capable:
            server.login(user, password)
        server.sendmail(from_email, [to_email], msg.as_string())
        server.quit()
        return {"ok": True}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


def default_app_settings() -> Dict[str, Any]:
    return {
        "id": "main",
        "app_name": "E-Learning Dosen",
        "campus_name": "Kampus Demo",
        "campus_address": "Alamat kampus",
        "program_name": "Program Studi",
        "lecturer_name": "Dosen Admin",
        "lecturer_email": "dosen@demo.id",
        "campus_logo_url": "",
        "active_academic_year": "2025/2026",
        "active_semester": "Ganjil",
    }


async def get_app_settings_cached() -> Dict[str, Any]:
    cached = _get_cached_settings("app_settings")
    if cached is None:
        cached = await db.app_settings.find_one({"id": "main"}, {"_id": 0}) or default_app_settings()
        _set_cached_settings("app_settings", cached)
    return cached


def validate_grade_predicates(predicates: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    cleaned = []
    for item in predicates:
        label = str(item.get("label", "")).strip().upper()
        min_score = float(item.get("min_score", 0))
        max_score = float(item.get("max_score", 0))
        if not label:
            raise HTTPException(status_code=400, detail="Label predikat wajib diisi")
        if min_score < 0 or max_score > 100 or min_score > max_score:
            raise HTTPException(status_code=400, detail="Range predikat harus 0-100 dan minimum <= maksimum")
        cleaned.append({"label": label, "min_score": min_score, "max_score": max_score})
    ordered = sorted(cleaned, key=lambda item: item["min_score"])
    for idx in range(len(ordered) - 1):
        if ordered[idx + 1]["min_score"] <= ordered[idx]["max_score"]:
            raise HTTPException(status_code=400, detail="Range predikat tidak boleh tumpang tindih")
    return cleaned


async def get_grade_predicates_for_class(class_id: str = "") -> List[Dict[str, Any]]:
    if class_id:
        class_doc = await db.grade_predicates.find_one({"class_id": class_id}, {"_id": 0})
        if class_doc:
            return class_doc.get("predicates", DEFAULT_GRADE_PREDICATES)
    global_doc = await db.grade_predicates.find_one({"class_id": ""}, {"_id": 0})
    return global_doc.get("predicates", DEFAULT_GRADE_PREDICATES) if global_doc else DEFAULT_GRADE_PREDICATES


async def calculate_grade_predicate(score: float, class_id: str = "") -> str:
    predicates = await get_grade_predicates_for_class(class_id)
    for item in predicates:
        if float(item["min_score"]) <= float(score) <= float(item["max_score"]):
            return item["label"]
    return "-"


async def get_whatsapp_settings(mask: bool = False) -> Dict[str, Any]:
    cached = _get_cached_settings("whatsapp_settings")
    if cached is None:
        settings = await db.whatsapp_settings.find_one({"id": "main"}, {"_id": 0})
        cached = {**default_whatsapp_settings(), **(settings or {})}
        _set_cached_settings("whatsapp_settings", cached)
    if mask:
        settings = cached.copy()
        settings["fonnte_token_masked"] = mask_secret(settings.get("fonnte_token", ""))
        settings["waha_api_key_masked"] = mask_secret(settings.get("waha_api_key", ""))
        settings.pop("fonnte_token", None)
        settings.pop("waha_api_key", None)
        return settings
    return cached


class SafeTemplateContext(dict):
    def __missing__(self, key: str) -> str:
        return "{" + key + "}"


def format_message_template(template: str, context: Dict[str, Any]) -> str:
    cleaned = {key: "" if value is None else str(value) for key, value in context.items()}
    try:
        return template.format_map(SafeTemplateContext(cleaned))
    except Exception:
        return template


def format_message_datetime(value: str) -> str:
    if not value:
        return "-"
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        return parsed.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return value


def whatsapp_send_delay_seconds(settings: Dict[str, Any]) -> int:
    try:
        delay = int(settings.get("send_delay_seconds", 3))
    except Exception:
        delay = 3
    return max(0, min(delay, 300))


async def wait_for_whatsapp_send_window(delay_seconds: int) -> None:
    if delay_seconds <= 0:
        return
    gate = await db.whatsapp_runtime.find_one({"id": "send_gate"}, {"_id": 0})
    last_attempt = (gate or {}).get("last_attempt_at", "")
    if last_attempt:
        try:
            last_at = datetime.fromisoformat(last_attempt.replace("Z", "+00:00"))
            wait_seconds = delay_seconds - (datetime.now(timezone.utc) - last_at).total_seconds()
            if wait_seconds > 0:
                await asyncio.sleep(wait_seconds)
        except Exception:
            pass
    await db.whatsapp_runtime.update_one({"id": "send_gate"}, {"$set": {"last_attempt_at": now_iso()}}, upsert=True)


async def enqueue_whatsapp_message(to: str, message: str, message_type: str, ref_id: str = "") -> Dict[str, Any]:
    settings = await get_whatsapp_settings(mask=False)
    provider = settings.get("provider", "disabled")
    status = "pending_config" if provider == "disabled" else "pending"
    doc = {
        "id": new_id(),
        "to": normalize_phone(to),
        "message": message,
        "message_type": message_type,
        "ref_id": ref_id,
        "provider": provider,
        "status": status,
        "response": "",
        "error": "Gateway belum dikonfigurasi" if status == "pending_config" else "",
        "created_at": now_iso(),
        "sent_at": "",
    }
    await db.whatsapp_messages.insert_one(doc)
    return public_doc(doc.copy())


def public_whatsapp_delivery_status(message: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not message:
        return {"message_id": "", "status": "", "provider": "", "created_at": "", "sent_at": "", "error": ""}
    status = message.get("status", "")
    return {
        "message_id": message.get("id", ""),
        "status": status,
        "provider": message.get("provider", ""),
        "created_at": message.get("created_at", ""),
        "sent_at": message.get("sent_at", ""),
        "error": message.get("error", "") if status in {"failed", "pending_config"} else "",
    }


def forgot_password_response_message(delivery: Dict[str, Any]) -> str:
    status = delivery.get("status", "")
    provider = delivery.get("provider", "")
    if status == "pending":
        provider_label = f" via {provider}" if provider and provider != "disabled" else ""
        return f"Permintaan reset password diproses. OTP masuk antrian WhatsApp{provider_label}. Tunggu sampai status berubah menjadi terkirim."
    if status == "sent":
        return "Permintaan reset password diproses. OTP sudah dikirim via WhatsApp."
    if status == "pending_config":
        return "Permintaan reset password diproses. OTP dibuat, tetapi gateway WhatsApp belum aktif. Pesan masuk antrian konfigurasi."
    if status == "no_whatsapp":
        return "Permintaan reset password diproses, tetapi nomor WhatsApp belum terdaftar."
    return "Permintaan reset password diproses."


async def simulate_whatsapp_typing(
    client: httpx.AsyncClient,
    provider: str,
    to: str,
    settings: Dict[str, Any],
    typing_seconds: int,
) -> None:
    if typing_seconds <= 0:
        return
    try:
        if provider == "fonnte":
            token = settings.get("fonnte_token", "")
            if token:
                await client.post(
                    settings.get("fonnte_url") or "https://api.fonnte.com/send",
                    data={"target": to, "message": "...", "delay": str(typing_seconds), "typing": "true"},
                    headers={"Authorization": token},
                    timeout=10,
                )
        elif provider == "waha":
            base_url = normalize_http_base_url(settings.get("waha_base_url", ""))
            api_key = settings.get("waha_api_key", "")
            session = (settings.get("waha_session") or "default").strip() or "default"
            if base_url:
                headers = waha_headers(api_key)
                chat_id = waha_chat_id(to)
                for _ in range(typing_seconds // 5):
                    await client.post(
                        f"{base_url}/api/sendPresence",
                        json={"session": session, "chatId": chat_id, "presence": "typing"},
                        headers=headers,
                        timeout=10,
                    )
                    await asyncio.sleep(5)
                remaining = typing_seconds % 5
                if remaining > 0:
                    await asyncio.sleep(remaining)
                return
        await asyncio.sleep(typing_seconds)
    except Exception:
        logger.debug("Typing simulation gagal, melanjutkan kirim pesan", exc_info=True)
        await asyncio.sleep(typing_seconds)


async def send_whatsapp_message(message_id: str) -> None:
    msg = await db.whatsapp_messages.find_one({"id": message_id}, {"_id": 0})
    if not msg:
        return
    provider = "disabled"
    async with _whatsapp_send_lock:
        settings = await get_whatsapp_settings(mask=False)
        provider = settings.get("provider", "disabled")
        if provider == "disabled":
            await db.whatsapp_messages.update_one(
                {"id": message_id}, {"$set": {"status": "pending_config", "error": "Gateway belum dikonfigurasi"}}
            )
            return
        try:
            await wait_for_whatsapp_send_window(whatsapp_send_delay_seconds(settings))
            typing_seconds = max(0, min(int(settings.get("typing_simulation_seconds", 30)), 120))
            async with httpx.AsyncClient(timeout=20) as client:
                await simulate_whatsapp_typing(client, provider, msg["to"], settings, typing_seconds)
                if provider == "fonnte":
                    token = settings.get("fonnte_token", "")
                    if not token:
                        raise RuntimeError("Token Fonnte belum diisi")
                    response = await client.post(
                        settings.get("fonnte_url") or "https://api.fonnte.com/send",
                        data={"target": msg["to"], "message": msg["message"]},
                        headers={"Authorization": token},
                    )
                elif provider == "waha":
                    base_url = normalize_http_base_url(settings.get("waha_base_url", ""))
                    api_key = settings.get("waha_api_key", "")
                    session = (settings.get("waha_session") or "default").strip() or "default"
                    if not base_url:
                        raise RuntimeError("WAHA Base URL belum diisi")
                    headers = waha_headers(api_key)
                    await ensure_waha_session_working(client, base_url, session, headers)
                    response = await client.post(
                        f"{base_url}/api/sendText",
                        json={"session": session, "chatId": waha_chat_id(msg["to"]), "text": msg["message"]},
                        headers=headers,
                    )
                else:
                    raise RuntimeError("Provider WhatsApp tidak dikenal")
            if response.status_code >= 400:
                raise RuntimeError(f"Gateway {provider} gagal ({response.status_code}): {response_excerpt(response)}")
            await db.whatsapp_messages.update_one(
                {"id": message_id},
                {"$set": {"status": "sent", "response": response.text[:2000], "sent_at": now_iso(), "provider": provider, "error": ""}},
            )
        except Exception as exc:
            await db.whatsapp_messages.update_one(
                {"id": message_id},
                {"$set": {"status": "failed", "error": str(exc), "provider": provider}},
            )


async def queue_student_whatsapp_message(
    student: Dict[str, Any],
    message: str,
    message_type: str,
    ref_id: str,
    background_tasks: BackgroundTasks,
) -> Optional[Dict[str, Any]]:
    if not normalize_phone(student.get("whatsapp", "")):
        return None
    queued = await enqueue_whatsapp_message(student.get("whatsapp", ""), message, message_type, ref_id)
    if queued.get("status") == "pending":
        background_tasks.add_task(send_whatsapp_message, queued["id"])
    return queued


def unique_ids(values: List[str]) -> List[str]:
    seen = set()
    result = []
    for value in values or []:
        item = str(value or "").strip()
        if item and item not in seen:
            seen.add(item)
            result.append(item)
    return result


async def get_manageable_class(class_id: str, active_only: bool = False, user: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    query: Dict[str, Any] = {"id": class_id}
    if active_only:
        query["status"] = "active"
    else:
        query["status"] = {"$ne": "deleted"}
    if user and user.get("role") == "lecturer":
        query["lecturer_id"] = user["id"]
    class_doc = await db.classes.find_one(query, {"_id": 0})
    if not class_doc:
        raise HTTPException(status_code=404, detail="Kelas aktif tidak ditemukan" if active_only else "Kelas tidak ditemukan")
    return class_doc


async def get_active_student(student_id: str) -> Dict[str, Any]:
    student = await db.users.find_one({"id": student_id, "role": "student"}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Mahasiswa tidak ditemukan")
    if student.get("status", "active") != "active":
        raise HTTPException(status_code=400, detail="Hanya mahasiswa aktif yang bisa diproses")
    return student


async def add_student_to_class_record(class_doc: Dict[str, Any], student: Dict[str, Any], actor_id: str) -> Dict[str, Any]:
    class_id = class_doc["id"]
    student_id = student["id"]
    already_joined = class_id in student.get("class_ids", [])
    await db.users.update_one({"id": student_id}, {"$addToSet": {"class_ids": class_id}})
    await db.classes.update_one({"id": class_id}, {"$addToSet": {"student_ids": student_id}})
    await db.enrollment_requests.update_many(
        {"class_id": class_id, "student_id": student_id, "status": {"$in": ["pending", "invited"]}},
        {"$set": {"status": "approved", "approved_at": now_iso(), "approved_by": actor_id}},
    )
    return {
        "student_id": student_id,
        "student_name": student.get("name", ""),
        "status": "already_joined" if already_joined else "approved",
    }


async def invite_student_to_class_record(
    class_doc: Dict[str, Any],
    student: Dict[str, Any],
    background_tasks: BackgroundTasks,
    actor_id: str,
) -> Dict[str, Any]:
    class_id = class_doc["id"]
    if class_id in student.get("class_ids", []):
        return {
            "student_id": student["id"],
            "student_name": student.get("name", ""),
            "status": "already_joined",
            "delivery_status": "",
            "message_id": "",
        }

    now = now_iso()
    existing = await db.enrollment_requests.find_one(
        {"class_id": class_id, "student_id": student["id"], "status": {"$in": ["pending", "invited"]}}, {"_id": 0}
    )
    request_doc = {
        "id": existing.get("id") if existing else new_id(),
        "class_id": class_doc["id"],
        "class_name": class_doc.get("name", ""),
        "class_code": class_doc.get("class_code", ""),
        "lecturer_id": class_doc.get("lecturer_id", ""),
        "lecturer_name": class_doc.get("lecturer_name", ""),
        "student_id": student["id"],
        "student_name": student.get("name", ""),
        "student_nim": student.get("nim", ""),
        "student_email": student.get("email", ""),
        "status": existing.get("status", "invited") if existing else "invited",
        "requested_at": existing.get("requested_at", now) if existing else now,
        "invited_at": now,
        "invited_by": actor_id,
    }

    settings = await get_whatsapp_settings(mask=False)
    link = build_app_fragment_link(settings.get("app_url", ""))
    message = format_message_template(
        "Halo {name}, Anda diundang bergabung ke kelas {class_name}. Kode kelas: {class_code}. Login ke aplikasi lalu masukkan kode kelas tersebut: {link}",
        {
            "name": student.get("name", ""),
            "nim": student.get("nim", ""),
            "class_name": class_doc.get("name", ""),
            "course_name": class_doc.get("course_name", ""),
            "class_code": class_doc.get("class_code", ""),
            "link": link,
        },
    )
    queued = await queue_student_whatsapp_message(student, message, "invite_kelas", request_doc["id"], background_tasks)
    request_doc["delivery_status"] = queued.get("status", "") if queued else "no_whatsapp"
    request_doc["message_id"] = queued.get("id", "") if queued else ""
    if existing:
        await db.enrollment_requests.update_one({"id": existing["id"]}, {"$set": request_doc})
    else:
        await db.enrollment_requests.insert_one(request_doc)
    return public_doc(request_doc.copy())


async def active_students_for_class(class_id: str, student_ids: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    conditions: List[Dict[str, Any]] = [{"class_ids": class_id}]
    if student_ids:
        conditions.append({"id": {"$in": student_ids}})
    students = await db.users.find(
        {
            "role": "student",
            "status": "active",
            "whatsapp": {"$exists": True, "$ne": ""},
            "$or": conditions,
        },
        {"_id": 0},
    ).to_list(2000)
    deduped: Dict[str, Dict[str, Any]] = {}
    for student in students:
        if normalize_phone(student.get("whatsapp", "")):
            deduped[student["id"]] = student
    return list(deduped.values())


async def notify_new_assignment_whatsapp(doc: Dict[str, Any], class_doc: Dict[str, Any], background_tasks: BackgroundTasks) -> None:
    settings = await get_whatsapp_settings(mask=False)
    link = build_app_fragment_link(settings.get("app_url", ""), f"assignment-{doc['id']}")
    students = await active_students_for_class(doc["class_id"], class_doc.get("student_ids", []))
    template = settings.get("assignment_template") or default_whatsapp_settings()["assignment_template"]
    for student in students:
        message = format_message_template(
            template,
            {
                "name": student.get("name", ""),
                "nim": student.get("nim", ""),
                "title": doc.get("title", ""),
                "class_name": doc.get("class_name", ""),
                "course_name": doc.get("course_name", ""),
                "deadline": format_message_datetime(doc.get("deadline", "")),
                "description": doc.get("description", ""),
                "link": link,
            },
        )
        await queue_student_whatsapp_message(student, message, "tugas_baru", doc["id"], background_tasks)


async def create_assignment_publication_reminders(doc: Dict[str, Any], class_doc: Dict[str, Any]) -> None:
    for student_id in class_doc.get("student_ids", []):
        await db.reminder_logs.insert_one(
            {
                "id": new_id(),
                "assignment_id": doc["id"],
                "student_id": student_id,
                "reminder_type": "tugas_baru",
                "sent_at": now_iso(),
                "status": "in_app",
                "response": "Reminder tampil di aplikasi",
            }
        )


async def send_assignment_publication_notifications(
    doc: Dict[str, Any], class_doc: Dict[str, Any], background_tasks: BackgroundTasks
) -> None:
    await create_assignment_publication_reminders(doc, class_doc)
    await notify_new_assignment_whatsapp(doc, class_doc, background_tasks)


async def dispatch_due_assignment_notifications(background_tasks: BackgroundTasks) -> None:
    query = {
        "is_active": True,
        "published_at": {"$nin": ["", None]},
        "$or": [{"published_notification_sent_at": {"$exists": False}}, {"published_notification_sent_at": ""}],
    }
    scheduled = await db.assignments.find(query, {"_id": 0}).to_list(1000)
    for assignment in scheduled:
        if not assignment_is_published(assignment):
            continue
        update = await db.assignments.update_one(
            {
                "id": assignment["id"],
                "$or": [{"published_notification_sent_at": {"$exists": False}}, {"published_notification_sent_at": ""}],
            },
            {"$set": {"published_notification_sent_at": now_iso()}},
        )
        if update.modified_count:
            class_doc = await db.classes.find_one({"id": assignment["class_id"]}, {"_id": 0}) or {}
            await send_assignment_publication_notifications(assignment, class_doc, background_tasks)


async def notify_submission_status_whatsapp(
    submission: Dict[str, Any],
    assignment: Dict[str, Any],
    message_type: str,
    background_tasks: BackgroundTasks,
    extra: Optional[Dict[str, Any]] = None,
) -> None:
    student = await db.users.find_one({"id": submission.get("student_id")}, {"_id": 0})
    if not student:
        return
    settings = await get_whatsapp_settings(mask=False)
    link = build_app_fragment_link(settings.get("app_url", ""), f"assignment-{submission.get('assignment_id', '')}")
    defaults = default_whatsapp_settings()
    template_key = "revision_template" if message_type == "revisi_tugas" else "grade_template"
    context = {
        "name": student.get("name", submission.get("student_name", "")),
        "nim": student.get("nim", submission.get("student_nim", "")),
        "title": assignment.get("title") or submission.get("assignment_title", ""),
        "class_name": assignment.get("class_name", ""),
        "course_name": assignment.get("course_name", ""),
        "grade": submission.get("grade", ""),
        "predicate": submission.get("grade_predicate", ""),
        "feedback": submission.get("feedback", ""),
        "revision_note": submission.get("revision_note", ""),
        "link": link,
    }
    context.update(extra or {})
    message = format_message_template(settings.get(template_key) or defaults[template_key], context)
    await queue_student_whatsapp_message(student, message, message_type, submission.get("id", ""), background_tasks)


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(password: str, password_hash: str) -> bool:
    return bcrypt.checkpw(password.encode("utf-8"), password_hash.encode("utf-8"))


async def find_user(user_id: str) -> Dict[str, Any]:
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Sesi tidak valid")
    if user.get("status", "active") != "active":
        raise HTTPException(status_code=403, detail="Akun tidak aktif")
    return user


async def get_current_user(authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Token diperlukan")
    token = authorization.replace("Bearer ", "", 1).strip()
    session = await db.sessions.find_one({"token": token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Sesi tidak ditemukan")
    return await find_user(session["user_id"])


async def require_admin(user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
    if user.get("role") not in {"admin", "lecturer"}:
        raise HTTPException(status_code=403, detail="Hanya dosen atau admin kampus")
    return user


async def require_campus_admin(user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Hanya admin kampus")
    return user


def is_campus_admin(user: Dict[str, Any]) -> bool:
    return user.get("role") == "admin"


async def lecturer_class_ids(user: Dict[str, Any], include_deleted: bool = False) -> List[str]:
    if user.get("role") == "student":
        return list(user.get("class_ids", []))
    query: Dict[str, Any] = {}
    if not include_deleted:
        query["status"] = {"$ne": "deleted"}
    if not is_campus_admin(user):
        query["lecturer_id"] = user["id"]
    docs = await db.classes.find(query, {"_id": 0, "id": 1}).to_list(5000)
    return [item["id"] for item in docs]


async def require_class_access(class_id: str, user: Dict[str, Any], active_only: bool = False) -> Dict[str, Any]:
    query: Dict[str, Any] = {"id": class_id}
    query["status"] = "active" if active_only else {"$ne": "deleted"}
    if user.get("role") == "lecturer":
        query["lecturer_id"] = user["id"]
    elif user.get("role") == "student":
        query["id"] = {"$in": user.get("class_ids", [])}
    class_doc = await db.classes.find_one(query, {"_id": 0})
    if not class_doc:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan atau bukan kelas yang Anda kelola")
    return class_doc


async def require_submission_access(submission_id: str, user: Dict[str, Any]) -> Dict[str, Any]:
    submission = await db.submissions.find_one({"id": submission_id}, {"_id": 0})
    if not submission:
        raise HTTPException(status_code=404, detail="Submission tidak ditemukan")
    await require_class_access(submission.get("class_id", ""), user)
    return submission


def chat_conversation_id(first_user_id: str, second_user_id: str) -> str:
    participants = sorted([first_user_id, second_user_id])
    return hashlib.sha256(":".join(participants).encode("utf-8")).hexdigest()


def chat_user_payload(user: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": user.get("id", ""),
        "role": user.get("role", ""),
        "username": user.get("username", ""),
        "nim": user.get("nim", ""),
        "name": user.get("name", ""),
        "email": user.get("email", ""),
    }


async def chat_contact_payload(user: Dict[str, Any]) -> Dict[str, Any]:
    payload = chat_user_payload(user)
    if user.get("role") != "admin":
        return payload
    settings = await db.app_settings.find_one({"id": "main"}, {"_id": 0, "lecturer_name": 1, "lecturer_email": 1}) or {}
    configured_name = str(settings.get("lecturer_name", "")).strip()
    configured_email = str(settings.get("lecturer_email", "")).strip().lower()
    applies_to_user = not configured_email or configured_email == str(user.get("email", "")).strip().lower()
    if applies_to_user and configured_name and configured_name.lower() != "dosen admin":
        payload["name"] = configured_name
    elif str(payload.get("name", "")).strip().lower() == "dosen admin":
        payload["name"] = "Dosen Pengampu"
    return payload


class ChatConnectionManager:
    def __init__(self) -> None:
        self.connections: Dict[str, set[WebSocket]] = {}
        self.viewing: Dict[WebSocket, str] = {}

    def is_online(self, user_id: str) -> bool:
        return bool(self.connections.get(user_id))

    def is_viewing(self, viewer_id: str, target_user_id: str) -> bool:
        return any(self.viewing.get(socket) == target_user_id for socket in self.connections.get(viewer_id, set()))

    async def connect(self, user_id: str, websocket: WebSocket) -> bool:
        await websocket.accept()
        self.connections.setdefault(user_id, set()).add(websocket)
        try:
            await websocket.send_json({"type": "presence_snapshot", "online_user_ids": list(self.connections.keys())})
            await self.broadcast({"type": "presence", "user_id": user_id, "online": True})
            return True
        except Exception:
            sockets = self.connections.get(user_id, set())
            sockets.discard(websocket)
            if not sockets:
                self.connections.pop(user_id, None)
            return False

    async def disconnect(self, user_id: str, websocket: WebSocket) -> None:
        viewed_user_id = self.viewing.pop(websocket, "")
        sockets = self.connections.get(user_id, set())
        sockets.discard(websocket)
        if not sockets:
            self.connections.pop(user_id, None)
        if viewed_user_id:
            await self.send_to_users(
                {viewed_user_id},
                {"type": "chat_focus", "user_id": user_id, "viewing": self.is_viewing(user_id, viewed_user_id)},
            )
        await self.broadcast({"type": "presence", "user_id": user_id, "online": self.is_online(user_id)})

    async def set_viewing(self, user_id: str, websocket: WebSocket, target_user_id: str) -> None:
        previous = self.viewing.get(websocket, "")
        self.viewing[websocket] = target_user_id
        if previous and previous != target_user_id:
            await self.send_to_users(
                {previous},
                {"type": "chat_focus", "user_id": user_id, "viewing": self.is_viewing(user_id, previous)},
            )
        if target_user_id:
            await self.send_to_users(
                {target_user_id},
                {"type": "chat_focus", "user_id": user_id, "viewing": True},
            )

    async def send_to_users(self, user_ids: set[str], payload: Dict[str, Any]) -> None:
        for user_id in user_ids:
            for socket in list(self.connections.get(user_id, set())):
                try:
                    await socket.send_json(payload)
                except Exception:
                    await self.disconnect(user_id, socket)

    async def broadcast(self, payload: Dict[str, Any]) -> None:
        await self.send_to_users(set(self.connections.keys()), payload)


chat_connections = ChatConnectionManager()


class LoginInput(BaseModel):
    identifier: str = ""
    email: str = ""
    password: str = Field(min_length=3)


class SsoExchangeInput(BaseModel):
    ticket: str = Field(min_length=20)


class RegisterStudentInput(BaseModel):
    username: str = ""
    nim: str
    name: str
    email: EmailStr
    whatsapp: str = ""
    password: str = Field(min_length=3)


class ForgotPasswordInput(BaseModel):
    identifier: str


class ChangePasswordInput(BaseModel):
    current_password: str
    new_password: str = Field(min_length=3)


class ProfileInput(BaseModel):
    name: str = Field(min_length=1)
    username: str = Field(min_length=1)
    email: EmailStr
    whatsapp: str = ""


class ResetPasswordOtpInput(BaseModel):
    identifier: str
    otp: str
    new_password: str = Field(min_length=3)


class JoinClassInput(BaseModel):
    class_code: str
    nim: str
    name: str
    email: EmailStr
    whatsapp: str = ""
    password: str = Field(min_length=3)


class ProgramInput(BaseModel):
    code: str
    name: str
    description: str = ""


class CourseInput(BaseModel):
    program_id: str = ""
    code: str
    name: str
    credits: int = 3
    description: str = ""


class ClassInput(BaseModel):
    academic_year: str
    semester: str
    course_id: str
    name: str
    schedule: str = ""


class StudentInput(BaseModel):
    nim: str
    name: str
    email: EmailStr
    whatsapp: str = ""
    class_id: str
    status: str = "active"
    password: str = "Mahasiswa123!"


class LecturerInput(BaseModel):
    employee_id: str = ""
    username: str = Field(min_length=3)
    name: str = Field(min_length=1)
    email: EmailStr
    whatsapp: str = ""
    password: str = Field(default="Dosen123!", min_length=6)
    status: str = "active"


class LecturerUpdateInput(BaseModel):
    employee_id: str = ""
    username: str = Field(min_length=3)
    name: str = Field(min_length=1)
    email: EmailStr
    whatsapp: str = ""
    status: str = "active"


class MaterialInput(BaseModel):
    class_id: str
    title: str
    description: str = ""
    meeting: str = "Pertemuan 1"
    file_url: str = ""
    video_url: str = ""
    meeting_type: str = "offline"
    meeting_url: str = ""
    is_active: bool = True
    locked_until: str = ""


class GoogleMeetInput(BaseModel):
    class_id: str
    title: str = ""


class CommentInput(BaseModel):
    material_id: str
    content: str
    parent_id: str = ""


class JoinRequestInput(BaseModel):
    class_code: str


class StudentIdsInput(BaseModel):
    student_ids: List[str] = Field(default_factory=list)


class RubricItem(BaseModel):
    criterion: str
    weight: float


class AssignmentInput(BaseModel):
    class_id: str
    title: str
    description: str
    attachment_link: str = ""
    deadline: str
    published_at: str = ""
    tolerance_hours: int = 0
    allowed_formats: List[str] = Field(default_factory=lambda: ["pdf", "docx", "xlsx", "zip", "png", "jpg"])
    max_file_size_mb: float = Field(default=DEFAULT_SUBMISSION_MAX_FILE_MB, gt=0)
    rubric: List[RubricItem] = Field(default_factory=list)
    assignment_type: str = "individu"
    allow_revision: bool = True
    is_active: bool = True
    is_practicum: bool = False
    practicum_goal: str = ""
    practicum_tools: str = ""
    practicum_steps: List[str] = Field(default_factory=list)
    required_screenshot: bool = False
    late_penalty_per_day: float = 0
    close_after_deadline: bool = False
    material_id: str = ""


class GradeItem(BaseModel):
    criterion: str
    weight: float
    score: float = Field(..., ge=0, le=100)


class GradeInput(BaseModel):
    rubric_scores: List[GradeItem]
    feedback: str = ""
    revision_note: str = ""
    status: str = "Dinilai"


class GradePredicateItem(BaseModel):
    label: str
    min_score: float
    max_score: float


class GradePredicateInput(BaseModel):
    class_id: str = ""
    predicates: List[GradePredicateItem]


class CleanDataInput(BaseModel):
    confirmation: str = ""


class ReminderInput(BaseModel):
    assignment_id: str
    student_id: str = ""
    reminder_type: str = "manual"
    message: str = ""


class BulkGradeItem(BaseModel):
    submission_id: str
    score: float = Field(..., ge=0, le=100)
    feedback: str = ""
    revision_note: str = ""


class BulkGradeInput(BaseModel):
    grades: List[BulkGradeItem]


class ResetPasswordInput(BaseModel):
    password: str = ""


class AppSettingsInput(BaseModel):
    app_name: str = "E-Learning Dosen"
    campus_name: str = ""
    campus_address: str = ""
    program_name: str = ""
    lecturer_name: str = ""
    lecturer_email: str = ""
    campus_logo_url: str = ""
    active_academic_year: str = ""
    active_semester: str = ""


class GoogleDriveSettingsInput(BaseModel):
    enabled: bool = True
    root_folder_id: str = ""
    root_folder_name: str = "E-Learning Dosen"
    require_upload: bool = False
    lecturer_folder_sharing_enabled: bool = False
    lecturer_folder_role: str = "reader"
    google_meet_enabled: bool = False
    google_workspace_delegated_user: str = ""
    service_account_json: str = ""
    clear_service_account: bool = False


class DatabaseBackupSettingsInput(BaseModel):
    enabled: bool = False
    frequency: str = "daily"
    run_time: str = "02:00"
    weekly_day: int = Field(default=0, ge=0, le=6)
    retention_count: int = Field(default=14, ge=1, le=90)
    upload_to_drive: bool = True
    keep_local: bool = True


class WhatsAppSettingsInput(BaseModel):
    provider: str = "disabled"
    app_url: str = ""
    fonnte_token: str = ""
    fonnte_url: str = "https://api.fonnte.com/send"
    waha_base_url: str = ""
    waha_api_key: str = ""
    waha_session: str = "default"
    send_delay_seconds: int = Field(default=3, ge=0, le=300)
    typing_simulation_seconds: int = Field(default=30, ge=0, le=120)
    otp_template: str = "Kode OTP reset password Anda: {code}. Berlaku {minutes} menit. Link: {link}"
    assignment_template: str = default_whatsapp_settings()["assignment_template"]
    grade_template: str = default_whatsapp_settings()["grade_template"]
    revision_template: str = default_whatsapp_settings()["revision_template"]


class EmailSettingsInput(BaseModel):
    enabled: bool = False
    smtp_host: str = ""
    smtp_port: int = Field(default=587, ge=1, le=65535)
    smtp_user: str = ""
    smtp_password: str = ""
    smtp_use_tls: bool = True
    from_name: str = "E-Learning Dosen"
    from_email: str = ""


class OidcSettingsInput(BaseModel):
    enabled: bool = True
    discovery_url: str = ""
    issuer: str = ""
    client_id: str = ""
    client_secret: str = ""
    redirect_uri: str = ""
    frontend_url: str = ""
    scopes: str = "openid profile email roles"
    local_login_enabled: bool = True
    clear_client_secret: bool = False


async def seed_data() -> None:
    existing_admin = await db.users.find_one({"role": "admin"}, {"_id": 0})
    if existing_admin:
        return

    admin_id = new_id()
    await db.users.insert_one(
        {
            "id": admin_id,
            "role": "admin",
            "username": "dosenadmin",
            "name": "Dosen Admin",
            "email": "dosen@demo.id",
            "whatsapp": "628000000001",
            "password_hash": hash_password("Dosen123!"),
            "status": "active",
            "created_at": now_iso(),
            "last_login_at": "",
        }
    )

    program_id = new_id()
    course_id = new_id()
    class_id = new_id()
    student_id = new_id()
    assignment_id = new_id()
    material_id = new_id()

    await db.programs.insert_one(
        {
            "id": program_id,
            "code": "IF",
            "name": "Teknik Informatika",
            "description": "Program studi demo untuk kelas e-learning.",
            "status": "active",
            "created_at": now_iso(),
        }
    )
    await db.courses.insert_one(
        {
            "id": course_id,
            "program_id": program_id,
            "program_name": "Teknik Informatika",
            "code": "IF401",
            "name": "Pemrograman Web Lanjut",
            "credits": 3,
            "description": "Mata kuliah praktis untuk membangun aplikasi web modern.",
            "status": "active",
            "created_at": now_iso(),
        }
    )
    await db.classes.insert_one(
        {
            "id": class_id,
            "academic_year": "2025/2026",
            "semester": "Ganjil",
            "course_id": course_id,
            "course_name": "Pemrograman Web Lanjut",
            "name": "IF-4A",
            "schedule": "Selasa 09.00",
            "class_code": "WEB4A1",
            "lecturer_id": admin_id,
            "lecturer_name": "Dosen Admin",
            "status": "active",
            "student_ids": [student_id],
            "created_at": now_iso(),
        }
    )
    await db.users.insert_one(
        {
            "id": student_id,
            "role": "student",
            "username": "alya",
            "nim": "230001001",
            "name": "Alya Pratama",
            "email": "alya@demo.id",
            "whatsapp": "628123456789",
            "password_hash": hash_password("Mahasiswa123!"),
            "status": "active",
            "class_ids": [class_id],
            "created_at": now_iso(),
            "last_login_at": "",
        }
    )
    deadline = (datetime.now(timezone.utc) + timedelta(days=3)).isoformat()
    await db.assignments.insert_one(
        {
            "id": assignment_id,
            "class_id": class_id,
            "course_id": course_id,
            "course_name": "Pemrograman Web Lanjut",
            "class_name": "IF-4A",
            "title": "Praktikum CRUD API",
            "description": "Buat endpoint CRUD sederhana dan unggah laporan singkat.",
            "deadline": deadline,
            "published_at": "",
            "tolerance_hours": 6,
            "allowed_formats": ["pdf", "docx", "zip"],
            "rubric": [
                {"criterion": "Ketepatan jawaban", "weight": 40},
                {"criterion": "Kerapian laporan", "weight": 20},
                {"criterion": "Kreativitas", "weight": 20},
                {"criterion": "Ketepatan waktu", "weight": 20},
            ],
            "assignment_type": "individu",
            "allow_revision": True,
            "is_active": True,
            "is_practicum": True,
            "practicum_goal": "Mahasiswa memahami struktur API CRUD.",
            "practicum_tools": "Python, FastAPI, PostgreSQL",
            "practicum_steps": ["Rancang endpoint", "Uji dengan curl", "Tulis laporan"],
            "required_screenshot": True,
            "late_penalty_per_day": 5,
            "close_after_deadline": False,
            "material_id": material_id,
            "created_at": now_iso(),
            "created_by": admin_id,
            "lecturer_id": admin_id,
        }
    )
    await db.materials.insert_one(
        {
            "id": material_id,
            "class_id": class_id,
            "title": "Arsitektur Aplikasi Full-stack",
            "description": "Materi pembuka tentang frontend, backend, dan database.",
            "meeting": "Pertemuan 1",
            "file_url": "https://drive.google.com/",
            "video_url": "",
            "is_active": True,
            "locked_until": "",
            "created_at": now_iso(),
            "created_by": admin_id,
            "lecturer_id": admin_id,
        }
    )
    logger.info("Seed data e-learning dibuat")


async def ensure_program_course_links() -> None:
    program = await db.programs.find_one({}, {"_id": 0})
    if not program:
        settings = await db.app_settings.find_one({"id": "main"}, {"_id": 0}) or {}
        program = {
            "id": new_id(),
            "code": "PRODI",
            "name": settings.get("program_name") or "Program Studi",
            "description": "",
            "status": "active",
            "created_at": now_iso(),
        }
        await db.programs.insert_one(program)
    await db.courses.update_many(
        {"$or": [{"program_id": {"$exists": False}}, {"program_id": ""}, {"program_id": None}]},
        {"$set": {"program_id": program["id"], "program_name": program["name"]}},
    )
    async for course in db.courses.find({"program_id": {"$exists": True}}, {"_id": 0, "id": 1, "program_id": 1}):
        linked = await db.programs.find_one({"id": course.get("program_id")}, {"_id": 0})
        if linked:
            await db.courses.update_one({"id": course["id"]}, {"$set": {"program_name": linked.get("name", "")}})


async def ensure_multi_lecturer_schema() -> None:
    """Backfill ownership so existing single-lecturer installations remain usable."""
    fallback = await db.users.find_one({"role": "admin", "status": {"$ne": "deleted"}}, {"_id": 0})
    if not fallback:
        return
    fallback_id = fallback["id"]
    fallback_name = fallback.get("name", "Admin Kampus")
    async for class_doc in db.classes.find(
        {"$or": [{"lecturer_id": {"$exists": False}}, {"lecturer_id": ""}, {"lecturer_id": None}]},
        {"_id": 0, "id": 1, "created_by": 1},
    ):
        owner_id = class_doc.get("created_by") or fallback_id
        owner = await db.users.find_one({"id": owner_id, "role": {"$in": ["admin", "lecturer"]}}, {"_id": 0}) or fallback
        await db.classes.update_one(
            {"id": class_doc["id"]},
            {"$set": {"lecturer_id": owner["id"], "lecturer_name": owner.get("name", fallback_name)}},
        )
    async for class_doc in db.classes.find({}, {"_id": 0, "id": 1, "lecturer_id": 1, "lecturer_name": 1}):
        owner_id = class_doc.get("lecturer_id") or fallback_id
        owner_name = class_doc.get("lecturer_name") or fallback_name
        class_id = class_doc["id"]
        for collection in [db.materials, db.assignments, db.submissions, db.enrollment_requests, db.reminder_logs]:
            await collection.update_many(
                {"class_id": class_id, "$or": [{"lecturer_id": {"$exists": False}}, {"lecturer_id": ""}, {"lecturer_id": None}]},
                {"$set": {"lecturer_id": owner_id, "lecturer_name": owner_name}},
            )


def env_flag(name: str, default: bool = False) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


def env_enabled(name: str, default: bool = True) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() not in {"0", "false", "no", "n", "off", "disabled"}


def env_service_account_configured() -> bool:
    return bool(
        os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
        or os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON_B64")
        or os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE")
    )


def drive_key_file_path() -> Path:
    configured = os.environ.get("GOOGLE_DRIVE_CONFIG_KEY_FILE", "").strip()
    return Path(configured).expanduser() if configured else ROOT_DIR / ".drive_config.key"


def normalize_fernet_key(value: str) -> bytes:
    raw = value.strip().encode()
    try:
        Fernet(raw)
        return raw
    except Exception:
        return base64.urlsafe_b64encode(hashlib.sha256(raw).digest())


def get_drive_config_key() -> bytes:
    env_key = os.environ.get("GOOGLE_DRIVE_CONFIG_KEY", "").strip()
    if env_key:
        return normalize_fernet_key(env_key)
    key_path = drive_key_file_path()
    if key_path.exists():
        return key_path.read_bytes().strip()
    key = Fernet.generate_key()
    key_path.write_bytes(key)
    os.chmod(key_path, 0o600)
    return key


def encrypt_secret(value: str) -> str:
    return Fernet(get_drive_config_key()).encrypt(value.encode("utf-8")).decode("utf-8")


def decrypt_secret(value: str) -> str:
    if not value:
        return ""
    try:
        return Fernet(get_drive_config_key()).decrypt(value.encode("utf-8")).decode("utf-8")
    except InvalidToken:
        logger.error("Kunci enkripsi aplikasi tidak cocok dengan credential tersimpan")
        return ""


def validate_oidc_url(value: str, label: str, required: bool = True) -> str:
    normalized = (value or "").strip().rstrip("/")
    if not normalized and not required:
        return ""
    parsed = urlsplit(normalized)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise HTTPException(status_code=400, detail=f"{label} harus berupa URL HTTP/HTTPS yang valid")
    host = (parsed.hostname or "").lower()
    local_http = host in {"localhost", "127.0.0.1", "::1"} or host.startswith(("10.", "192.168."))
    if parsed.scheme != "https" and not local_http:
        raise HTTPException(status_code=400, detail=f"{label} non-local wajib menggunakan HTTPS")
    return normalized


def oidc_admin_view(settings: Dict[str, Any], source: str = "environment", metadata: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    secret_configured = bool(settings.get("client_secret"))
    return {
        "enabled": bool(settings.get("enabled")),
        "provider": "SCI-ID",
        "discovery_url": settings.get("discovery_url", ""),
        "issuer": settings.get("issuer", ""),
        "client_id": settings.get("client_id", ""),
        "client_secret_configured": secret_configured,
        "client_secret_masked": "••••••••••••" if secret_configured else "",
        "redirect_uri": settings.get("redirect_uri", ""),
        "frontend_url": settings.get("frontend_url", ""),
        "scopes": settings.get("scopes", "openid profile email roles"),
        "local_login_enabled": bool(settings.get("local_login_enabled", True)),
        "source": source,
        "updated_at": (metadata or {}).get("updated_at", ""),
        "updated_by": (metadata or {}).get("updated_by", ""),
    }


async def load_oidc_runtime_settings() -> Dict[str, Any]:
    doc = await db.oidc_settings.find_one({"id": "main"}, {"_id": 0}) or {}
    _oidc_runtime_settings.clear()
    if doc:
        _oidc_runtime_settings.update(
            {
                "enabled": bool(doc.get("enabled")),
                "discovery_url": str(doc.get("discovery_url", "")).strip(),
                "issuer": str(doc.get("issuer", "")).strip().rstrip("/"),
                "client_id": str(doc.get("client_id", "")).strip(),
                "client_secret": decrypt_secret(str(doc.get("client_secret_encrypted", ""))),
                "redirect_uri": str(doc.get("redirect_uri", "")).strip(),
                "frontend_url": str(doc.get("frontend_url", "")).strip().rstrip("/"),
                "scopes": str(doc.get("scopes", "openid profile email roles")).strip(),
                "local_login_enabled": bool(doc.get("local_login_enabled", True)),
            }
        )
    clear_oidc_discovery_cache()
    return oidc_admin_view(oidc_settings(), "admin_ui" if doc else "environment", doc)


def normalize_service_account_payload(value: str) -> Dict[str, Any]:
    raw = (value or "").strip()
    if not raw:
        return {}
    if not raw.startswith("{"):
        raw = base64.b64decode(raw).decode("utf-8")
    info = json.loads(raw)
    if not isinstance(info, dict) or info.get("type") != "service_account" or not info.get("client_email") or not info.get("private_key"):
        raise ValueError("JSON service account tidak valid")
    return info


async def get_google_drive_settings(mask: bool = True) -> Dict[str, Any]:
    cached = _get_cached_settings("google_drive_settings")
    if cached is not None:
        if mask:
            return cached.copy()
        return cached
    doc = await db.google_drive_settings.find_one({"id": "main"}, {"_id": 0}) or {}
    credential = ""
    credential_source = ""
    service_account_email = doc.get("service_account_email", "")
    if doc.get("service_account_json_encrypted"):
        credential = decrypt_secret(doc.get("service_account_json_encrypted", ""))
        credential_source = "admin_ui" if credential else ""
    env_has_credential = env_service_account_configured()
    if not credential and env_has_credential:
        credential_source = "environment"
        service_account_email = service_account_email or service_account_email_from_env()
    settings = {
        "id": "main",
        "enabled": bool(doc.get("enabled", env_enabled("GOOGLE_DRIVE_ENABLED", True))),
        "root_folder_id": doc.get("root_folder_id", os.environ.get("GOOGLE_DRIVE_ROOT_FOLDER_ID", "")),
        "root_folder_name": safe_path_segment(doc.get("root_folder_name", os.environ.get("GOOGLE_DRIVE_ROOT_FOLDER_NAME", "E-Learning Dosen"))),
        "require_upload": bool(doc.get("require_upload", env_flag("GOOGLE_DRIVE_REQUIRE_UPLOAD", False))),
        "lecturer_folder_sharing_enabled": bool(doc.get("lecturer_folder_sharing_enabled", False)),
        "lecturer_folder_role": doc.get("lecturer_folder_role", "reader") if doc.get("lecturer_folder_role") in {"reader", "writer"} else "reader",
        "google_meet_enabled": bool(doc.get("google_meet_enabled", env_enabled("GOOGLE_MEET_ENABLED", False))),
        "google_workspace_delegated_user": str(doc.get("google_workspace_delegated_user") or os.environ.get("GOOGLE_WORKSPACE_DELEGATED_USER", "")).strip().lower(),
        "service_account_configured": bool(credential or env_has_credential),
        "service_account_source": credential_source,
        "service_account_email": service_account_email,
        "updated_at": doc.get("updated_at", ""),
        "updated_by": doc.get("updated_by", ""),
    }
    settings["google_meet_ready"] = bool(
        settings["google_meet_enabled"]
        and settings["service_account_configured"]
        and settings["google_workspace_delegated_user"]
    )
    settings["google_meet_scope"] = "https://www.googleapis.com/auth/meetings.space.created"
    if not mask:
        settings["service_account_json"] = credential
    _set_cached_settings("google_drive_settings", settings)
    if mask:
        return settings.copy()
    return settings


def service_account_email_from_env() -> str:
    try:
        json_payload = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
        json_payload_b64 = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON_B64")
        file_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE")
        if json_payload:
            return json.loads(json_payload).get("client_email", "")
        if json_payload_b64:
            return json.loads(base64.b64decode(json_payload_b64).decode("utf-8")).get("client_email", "")
        if file_path and Path(file_path).exists():
            return json.loads(Path(file_path).read_text()).get("client_email", "")
    except Exception:
        return ""
    return ""


def google_drive_upload_enabled(settings: Dict[str, Any]) -> bool:
    return bool(settings.get("enabled") and settings.get("service_account_configured"))


def google_drive_upload_required(settings: Dict[str, Any]) -> bool:
    return bool(settings.get("require_upload"))


def google_drive_error_message(exc: Exception) -> str:
    text = str(exc)
    if "storageQuotaExceeded" in text or "Service Accounts do not have storage quota" in text:
        return (
            "Google Drive menolak upload file karena service account tidak punya kuota penyimpanan. "
            "Gunakan Shared Drive atau folder Drive yang mendukung upload service account, lalu jalankan retry."
        )
    if "insufficientFilePermissions" in text or "The user does not have sufficient permissions" in text:
        return "Service account belum punya izin upload ke folder Google Drive. Bagikan folder dengan akses Editor."
    if "File not found" in text or "notFound" in text:
        return "Folder Google Drive tidak ditemukan atau belum dibagikan ke service account."
    return text[:500]


async def storage_status_summary() -> Dict[str, Any]:
    settings = await get_google_drive_settings(mask=True)
    drive_enabled = google_drive_upload_enabled(settings)
    drive_sync_counts = {
        "pending": await db.stored_files.count_documents({"drive_sync_status": "pending"}),
        "synced": await db.stored_files.count_documents({"drive_sync_status": "synced"}),
        "failed": await db.stored_files.count_documents({"drive_sync_status": "failed"}),
        "not_configured": await db.stored_files.count_documents({"drive_sync_status": "not_configured"}),
    }
    return {
        "storage_mode": "google_drive" if drive_enabled else "server_local",
        "drive_configured": bool(settings.get("service_account_configured")),
        "drive_enabled": drive_enabled,
        "drive_required": google_drive_upload_required(settings),
        "drive_root_folder_id_configured": bool(settings.get("root_folder_id")),
        "drive_root_folder_name": settings.get("root_folder_name", "E-Learning Dosen"),
        "drive_service_account_email": settings.get("service_account_email", ""),
        "drive_service_account_source": settings.get("service_account_source", ""),
        "drive_sync": drive_sync_counts,
    }


def get_drive_service(settings: Optional[Dict[str, Any]] = None):
    scopes = ["https://www.googleapis.com/auth/drive"]
    settings = settings or {}
    if settings.get("service_account_json"):
        credentials = service_account.Credentials.from_service_account_info(json.loads(settings["service_account_json"]), scopes=scopes)
        return build("drive", "v3", credentials=credentials, cache_discovery=False)
    json_payload = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    json_payload_b64 = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON_B64")
    file_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE")
    if json_payload:
        credentials = service_account.Credentials.from_service_account_info(json.loads(json_payload), scopes=scopes)
    elif json_payload_b64:
        credentials = service_account.Credentials.from_service_account_info(
            json.loads(base64.b64decode(json_payload_b64).decode("utf-8")),
            scopes=scopes,
        )
    elif file_path:
        credentials = service_account.Credentials.from_service_account_file(file_path, scopes=scopes)
    else:
        raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_JSON, GOOGLE_SERVICE_ACCOUNT_JSON_B64, atau GOOGLE_SERVICE_ACCOUNT_FILE belum dikonfigurasi")
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


BACKUP_ROOT = STORAGE_ROOT / "Database Backups"
BACKUP_TIMEZONE = ZoneInfo(os.environ.get("BACKUP_TIMEZONE", "Asia/Jakarta"))
_database_backup_scheduler_task: Optional[asyncio.Task] = None


def default_database_backup_settings() -> Dict[str, Any]:
    return {
        "id": "main",
        "enabled": False,
        "frequency": "daily",
        "run_time": "02:00",
        "weekly_day": 0,
        "retention_count": 14,
        "upload_to_drive": True,
        "keep_local": True,
        "timezone": str(BACKUP_TIMEZONE),
    }


async def get_database_backup_settings() -> Dict[str, Any]:
    doc = await db.database_backup_settings.find_one({"id": "main"}, {"_id": 0}) or {}
    return {**default_database_backup_settings(), **doc, "running": bool(doc.get("running"))}


def next_database_backup_at(settings: Dict[str, Any], after: Optional[datetime] = None) -> str:
    current = (after or datetime.now(timezone.utc)).astimezone(BACKUP_TIMEZONE)
    try:
        hour, minute = [int(item) for item in str(settings.get("run_time", "02:00")).split(":", 1)]
    except (TypeError, ValueError):
        hour, minute = 2, 0
    candidate = current.replace(hour=hour, minute=minute, second=0, microsecond=0)
    if settings.get("frequency") == "weekly":
        target_day = int(settings.get("weekly_day", 0))
        candidate += timedelta(days=(target_day - candidate.weekday()) % 7)
        if candidate <= current:
            candidate += timedelta(days=7)
    elif candidate <= current:
        candidate += timedelta(days=1)
    return candidate.astimezone(timezone.utc).isoformat()


async def database_backup_payload() -> bytes:
    collection_names = sorted(
        name for name in await db.list_collection_names() if not name.startswith("system.")
    )
    collections: Dict[str, List[Dict[str, Any]]] = {}
    for name in collection_names:
        documents: List[Dict[str, Any]] = []
        async for document in db[name].find({}):
            documents.append(document)
        collections[name] = documents
    payload = {
        "format": "nugaslagi-postgresql-jsonb",
        "version": 2,
        "database": db.name,
        "created_at": now_iso(),
        "collections": collections,
    }
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    return gzip.compress(raw, compresslevel=6)


def upload_database_backup_to_drive_sync(
    content: bytes, file_name: str, settings: Dict[str, Any]
) -> Dict[str, str]:
    service = get_drive_service(settings)
    root_folder_id = settings.get("root_folder_id", "")
    if not root_folder_id:
        root_folder_id = drive_find_or_create_folder(
            service,
            safe_path_segment(settings.get("root_folder_name") or "E-Learning Dosen"),
            None,
        )
    backup_folder_id = drive_find_or_create_folder(
        service, "Database Backups", root_folder_id
    )
    media = MediaIoBaseUpload(
        io.BytesIO(content), mimetype="application/gzip", resumable=True
    )
    uploaded = (
        service.files()
        .create(
            body={"name": file_name, "parents": [backup_folder_id]},
            media_body=media,
            fields="id,name,webViewLink",
            supportsAllDrives=True,
        )
        .execute()
    )
    return {
        "drive_file_id": uploaded.get("id", ""),
        "drive_file_url": uploaded.get("webViewLink", ""),
    }


def delete_database_backup_from_drive_sync(file_id: str, settings: Dict[str, Any]) -> None:
    if file_id:
        get_drive_service(settings).files().delete(
            fileId=file_id, supportsAllDrives=True
        ).execute()


async def enforce_database_backup_retention(settings: Dict[str, Any]) -> None:
    retention = max(1, int(settings.get("retention_count", 14)))
    expired = await db.database_backups.find(
        {"status": {"$in": ["completed", "completed_with_warning"]}}, {"_id": 0}
    ).sort("created_at", -1).skip(retention).to_list(500)
    drive_settings = await get_google_drive_settings(mask=False)
    for item in expired:
        local_path = item.get("local_path", "")
        if local_path:
            path = Path(local_path)
            if path.exists() and BACKUP_ROOT.resolve() in path.resolve().parents:
                path.unlink(missing_ok=True)
        if item.get("drive_file_id") and google_drive_upload_enabled(drive_settings):
            try:
                await asyncio.to_thread(
                    delete_database_backup_from_drive_sync,
                    item["drive_file_id"],
                    drive_settings,
                )
            except Exception as exc:
                logger.warning("Backup lama gagal dihapus dari Drive: %s", exc)
        await db.database_backups.delete_one({"id": item["id"]})


async def create_database_backup(trigger: str, user_id: str = "system") -> Dict[str, Any]:
    backup_id = str(uuid.uuid4())
    created_at = now_iso()
    file_name = f"nugaslagi-db-{datetime.now(BACKUP_TIMEZONE).strftime('%Y%m%d-%H%M%S')}.json.gz"
    record = {
        "id": backup_id,
        "file_name": file_name,
        "trigger": trigger,
        "status": "running",
        "created_at": created_at,
        "created_by": user_id,
        "size": 0,
        "local_available": False,
        "drive_file_id": "",
        "drive_file_url": "",
        "error": "",
    }
    await db.database_backups.insert_one(record.copy())
    settings = await get_database_backup_settings()
    try:
        content = await database_backup_payload()
        BACKUP_ROOT.mkdir(parents=True, exist_ok=True)
        local_path = BACKUP_ROOT / file_name
        local_path.write_bytes(content)
        update: Dict[str, Any] = {
            "status": "completed",
            "completed_at": now_iso(),
            "size": len(content),
            "local_path": str(local_path),
            "local_available": True,
        }
        if settings.get("upload_to_drive"):
            drive_settings = await get_google_drive_settings(mask=False)
            if not google_drive_upload_enabled(drive_settings):
                update.update(
                    status="completed_with_warning",
                    error="Backup lokal berhasil, tetapi Google Drive belum aktif atau belum dikonfigurasi.",
                )
            else:
                try:
                    drive_result = await asyncio.to_thread(
                        upload_database_backup_to_drive_sync,
                        content,
                        file_name,
                        drive_settings,
                    )
                    update.update(drive_result)
                except Exception as exc:
                    update.update(
                        status="completed_with_warning",
                        error=f"Backup lokal berhasil, upload Drive gagal: {google_drive_error_message(exc)}",
                    )
        if not settings.get("keep_local") and update.get("drive_file_id"):
            local_path.unlink(missing_ok=True)
            update.update(local_path="", local_available=False)
        await db.database_backups.update_one({"id": backup_id}, {"$set": update})
        await enforce_database_backup_retention(settings)
    except Exception as exc:
        logger.exception("Backup database gagal: %s", exc)
        await db.database_backups.update_one(
            {"id": backup_id},
            {"$set": {"status": "failed", "completed_at": now_iso(), "error": str(exc)[:500]}},
        )
    return await db.database_backups.find_one({"id": backup_id}, {"_id": 0})


async def database_backup_scheduler() -> None:
    while True:
        try:
            settings = await get_database_backup_settings()
            if settings.get("enabled"):
                now = now_iso()
                next_run = settings.get("next_run_at") or next_database_backup_at(settings)
                if not settings.get("next_run_at"):
                    await db.database_backup_settings.update_one(
                        {"id": "main"}, {"$set": {"next_run_at": next_run}}, upsert=True
                    )
                if next_run <= now:
                    lock_until = (datetime.now(timezone.utc) + timedelta(minutes=30)).isoformat()
                    claimed = await db.database_backup_settings.update_one(
                        {
                            "id": "main",
                            "enabled": True,
                            "next_run_at": {"$lte": now},
                            "$or": [
                                {"auto_lock_until": {"$exists": False}},
                                {"auto_lock_until": {"$lt": now}},
                            ],
                        },
                        {"$set": {"running": True, "auto_lock_until": lock_until}},
                    )
                    if claimed.modified_count:
                        await create_database_backup("automatic", "system")
                        latest = await get_database_backup_settings()
                        await db.database_backup_settings.update_one(
                            {"id": "main"},
                            {"$set": {
                                "running": False,
                                "last_run_at": now_iso(),
                                "next_run_at": next_database_backup_at(latest),
                            }, "$unset": {"auto_lock_until": ""}},
                        )
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            logger.exception("Scheduler backup database bermasalah: %s", exc)
        await asyncio.sleep(60)


def get_meet_access_token(settings: Dict[str, Any], delegated_user: str = "") -> str:
    """Get a user-authorized token for Meet using Workspace domain-wide delegation."""
    credential_payload = settings.get("service_account_json", "")
    scopes = ["https://www.googleapis.com/auth/meetings.space.created"]
    if credential_payload:
        credentials = service_account.Credentials.from_service_account_info(
            json.loads(credential_payload), scopes=scopes
        )
    else:
        json_payload = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
        json_payload_b64 = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON_B64")
        file_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE")
        if json_payload:
            credentials = service_account.Credentials.from_service_account_info(json.loads(json_payload), scopes=scopes)
        elif json_payload_b64:
            credentials = service_account.Credentials.from_service_account_info(
                json.loads(base64.b64decode(json_payload_b64).decode("utf-8")), scopes=scopes
            )
        elif file_path:
            credentials = service_account.Credentials.from_service_account_file(file_path, scopes=scopes)
        else:
            raise RuntimeError("Service account Google belum dikonfigurasi")
    effective_user = str(delegated_user or settings.get("google_workspace_delegated_user", "")).strip()
    if effective_user:
        credentials = credentials.with_subject(effective_user)
    credentials.refresh(GoogleAuthRequest())
    if not credentials.token:
        raise RuntimeError("Token Google Meet tidak berhasil dibuat")
    return credentials.token


def create_google_meet_space_sync(settings: Dict[str, Any], delegated_user: str = "") -> Dict[str, str]:
    if not settings.get("service_account_configured"):
        raise RuntimeError("Service account Google belum dikonfigurasi")
    effective_user = str(delegated_user or settings.get("google_workspace_delegated_user", "")).strip()
    if not effective_user:
        raise RuntimeError(
            "Isi Google Workspace user untuk Meet dan aktifkan Domain-wide Delegation pada service account"
        )
    token = get_meet_access_token(settings, effective_user)
    response = httpx.post(
        "https://meet.googleapis.com/v2/spaces",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json={},
        timeout=20,
    )
    if response.status_code >= 400:
        detail = ""
        try:
            detail = response.json().get("error", {}).get("message", "")
        except ValueError:
            detail = response.text
        raise RuntimeError(f"Google Meet menolak pembuatan ruang: {detail[:400]}")
    payload = response.json()
    meeting_uri = str(payload.get("meetingUri", "")).strip()
    meeting_code = str(payload.get("meetingCode", "")).strip()
    if not meeting_uri:
        raise RuntimeError("Google Meet tidak mengembalikan link pertemuan")
    return {"meeting_url": meeting_uri, "meeting_code": meeting_code, "meeting_space_name": str(payload.get("name", ""))}


def drive_query_literal(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'")


def drive_find_or_create_folder(service, name: str, parent_id: Optional[str]) -> str:
    folder_name = safe_path_segment(name or "Tanpa Nama")
    safe_name = drive_query_literal(folder_name)
    parent_clause = f" and '{parent_id}' in parents" if parent_id else " and 'root' in parents"
    query = (
        "mimeType='application/vnd.google-apps.folder' "
        f"and name='{safe_name}' and trashed=false{parent_clause}"
    )
    result = (
        service.files()
        .list(
            q=query,
            spaces="drive",
            corpora="allDrives",
            fields="files(id,name)",
            pageSize=1,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        )
        .execute()
    )
    files = result.get("files", [])
    if files:
        return files[0]["id"]
    metadata = {"name": folder_name, "mimeType": "application/vnd.google-apps.folder"}
    if parent_id:
        metadata["parents"] = [parent_id]
    folder = service.files().create(body=metadata, fields="id", supportsAllDrives=True).execute()
    return folder["id"]


def lecturer_drive_folder_name(lecturer: Dict[str, Any]) -> str:
    existing_name = str(lecturer.get("drive_folder_name", "")).strip()
    if existing_name:
        return safe_path_segment(existing_name)
    lecturer_code = str(lecturer.get("employee_id") or lecturer.get("id") or "DOSEN").strip()
    lecturer_name = str(lecturer.get("name") or "Dosen").strip()
    return safe_path_segment(f"{lecturer_code} - {lecturer_name}")


def drive_ensure_user_permission(service, folder_id: str, email: str, role: str) -> Dict[str, str]:
    clean_email = (email or "").strip().lower()
    clean_role = role if role in {"reader", "writer"} else "reader"
    if not clean_email:
        return {"drive_access_status": "missing_email", "drive_permission_id": ""}
    permissions = (
        service.permissions()
        .list(
            fileId=folder_id,
            fields="permissions(id,type,emailAddress,role)",
            supportsAllDrives=True,
        )
        .execute()
        .get("permissions", [])
    )
    existing = next(
        (
            item
            for item in permissions
            if item.get("type") == "user" and str(item.get("emailAddress", "")).lower() == clean_email
        ),
        None,
    )
    if existing:
        permission_id = existing.get("id", "")
        if existing.get("role") != clean_role:
            service.permissions().update(
                fileId=folder_id,
                permissionId=permission_id,
                body={"role": clean_role},
                supportsAllDrives=True,
            ).execute()
    else:
        created = (
            service.permissions()
            .create(
                fileId=folder_id,
                body={"type": "user", "role": clean_role, "emailAddress": clean_email},
                fields="id",
                sendNotificationEmail=False,
                supportsAllDrives=True,
            )
            .execute()
        )
        permission_id = created.get("id", "")
    return {
        "drive_access_status": "shared",
        "drive_permission_id": permission_id,
        "drive_permission_email": clean_email,
        "drive_permission_role": clean_role,
    }


def drive_revoke_user_permission(service, folder_id: str, permission_id: str) -> None:
    if folder_id and permission_id:
        service.permissions().delete(
            fileId=folder_id,
            permissionId=permission_id,
            supportsAllDrives=True,
        ).execute()


def drive_prepare_lecturer_folder(
    settings: Dict[str, Any], lecturer: Dict[str, Any]
) -> Dict[str, str]:
    service = get_drive_service(settings)
    root_folder_name = safe_path_segment(settings.get("root_folder_name") or "E-Learning Dosen")
    root_folder_id = settings.get("root_folder_id") or os.environ.get("GOOGLE_DRIVE_ROOT_FOLDER_ID")
    if not root_folder_id:
        root_folder_id = drive_find_or_create_folder(service, root_folder_name, None)
    lecturers_folder_id = drive_find_or_create_folder(service, "Dosen", root_folder_id)
    folder_name = lecturer_drive_folder_name(lecturer)
    folder_id = drive_find_or_create_folder(service, folder_name, lecturers_folder_id)
    result = {
        "drive_folder_id": folder_id,
        "drive_folder_name": folder_name,
        "drive_access_status": "app_only",
        "drive_permission_id": "",
        "drive_permission_email": "",
        "drive_permission_role": "",
        "drive_access_error": "",
    }
    if settings.get("lecturer_folder_sharing_enabled") and lecturer.get("status", "active") == "active":
        try:
            result.update(
                drive_ensure_user_permission(
                    service,
                    folder_id,
                    lecturer.get("email", ""),
                    settings.get("lecturer_folder_role", "reader"),
                )
            )
        except Exception as exc:
            result.update(
                {
                    "drive_access_status": "share_failed",
                    "drive_access_error": google_drive_error_message(exc),
                }
            )
    return result


async def reconcile_lecturer_drive_access(lecturer_id: str) -> None:
    lecturer = await db.users.find_one({"id": lecturer_id, "role": "lecturer"}, {"_id": 0})
    if not lecturer:
        return
    settings = await get_google_drive_settings(mask=False)
    if not google_drive_upload_enabled(settings):
        return
    old_folder_id = str(lecturer.get("drive_folder_id", ""))
    old_permission_id = str(lecturer.get("drive_permission_id", ""))
    should_share = bool(
        settings.get("lecturer_folder_sharing_enabled") and lecturer.get("status", "active") == "active"
    )
    if old_folder_id and old_permission_id and not should_share:
        try:
            await asyncio.to_thread(
                drive_revoke_user_permission,
                get_drive_service(settings),
                old_folder_id,
                old_permission_id,
            )
        except Exception as exc:
            logger.warning("Gagal mencabut akses Drive dosen %s: %s", lecturer_id, exc)
    try:
        drive_doc = await asyncio.to_thread(drive_prepare_lecturer_folder, settings, lecturer)
    except Exception as exc:
        logger.exception("Gagal menyiapkan folder Drive dosen %s: %s", lecturer_id, exc)
        await db.users.update_one(
            {"id": lecturer_id},
            {"$set": {"drive_access_status": "provision_failed", "drive_access_error": google_drive_error_message(exc)}},
        )
        return
    await db.users.update_one(
        {"id": lecturer_id},
        {"$set": {**drive_doc, "drive_access_updated_at": now_iso()}},
    )


async def reconcile_all_lecturer_drive_access() -> None:
    lecturer_ids = await db.users.distinct("id", {"role": "lecturer", "status": {"$ne": "deleted"}})
    for lecturer_id in lecturer_ids:
        await reconcile_lecturer_drive_access(str(lecturer_id))


def upload_to_drive(
    temp_path: str,
    filename: str,
    mime_type: str,
    hierarchy: List[str],
    settings: Dict[str, Any],
    lecturer_email: str = "",
) -> Dict[str, Any]:
    service = get_drive_service(settings)
    root_folder_name = safe_path_segment(settings.get("root_folder_name") or "E-Learning Dosen")
    parent_id = settings.get("root_folder_id") or os.environ.get("GOOGLE_DRIVE_ROOT_FOLDER_ID")
    path_parts = [root_folder_name]
    if not parent_id:
        parent_id = drive_find_or_create_folder(service, root_folder_name, None)
    lecturer_folder_id = ""
    for folder_index, folder_name in enumerate(hierarchy):
        clean_folder_name = safe_path_segment(folder_name or "Tanpa Nama")
        parent_id = drive_find_or_create_folder(service, clean_folder_name, parent_id)
        path_parts.append(clean_folder_name)
        if folder_index == 1 and hierarchy[0] == "Dosen":
            lecturer_folder_id = parent_id
    media = MediaFileUpload(temp_path, mimetype=mime_type or "application/octet-stream", resumable=True)
    metadata = {"name": safe_path_segment(filename), "parents": [parent_id]}
    uploaded = (
        service.files()
        .create(
            body=metadata,
            media_body=media,
            fields="id,name,webViewLink,mimeType,size,parents",
            supportsAllDrives=True,
        )
        .execute()
    )
    result = {
        "drive_file_id": uploaded.get("id", ""),
        "drive_file_name": uploaded.get("name", filename),
        "drive_file_url": uploaded.get("webViewLink", ""),
        "drive_mime_type": uploaded.get("mimeType", mime_type),
        "drive_folder_id": parent_id,
        "drive_folder_path": str(Path(*path_parts)),
        "drive_uploaded_at": now_iso(),
    }
    if lecturer_folder_id:
        result.update(
            {
                "drive_lecturer_folder_id": lecturer_folder_id,
                "drive_access_status": "app_only",
                "drive_permission_id": "",
                "drive_permission_email": "",
                "drive_permission_role": "",
                "drive_access_error": "",
            }
        )
        if settings.get("lecturer_folder_sharing_enabled"):
            try:
                result.update(
                    drive_ensure_user_permission(
                        service,
                        lecturer_folder_id,
                        lecturer_email,
                        settings.get("lecturer_folder_role", "reader"),
                    )
                )
            except Exception as exc:
                result.update(
                    {
                        "drive_access_status": "share_failed",
                        "drive_access_error": google_drive_error_message(exc),
                    }
                )
    return result


async def refresh_embedded_file_references(file_id: str) -> None:
    updated = await db.stored_files.find_one({"id": file_id}, {"_id": 0})
    if not updated:
        return
    public_file = enrich_file_urls(public_doc(updated.copy()))
    await db.submissions.update_many(
        {"file.file_id": file_id},
        {"$set": {"file": public_file}},
    )
    await db.submissions.update_many(
        {"files.file_id": file_id},
        {"$set": {"files.$[item]": public_file}},
        array_filters=[{"item.file_id": file_id}],
    )
    await db.assignments.update_many(
        {"attachments.file_id": file_id},
        {"$set": {"attachments.$[item]": public_file}},
        array_filters=[{"item.file_id": file_id}],
    )


async def sync_stored_file_to_drive(file_id: str) -> None:
    file_doc = await db.stored_files.find_one({"id": file_id}, {"_id": 0})
    if not file_doc:
        return
    settings = await get_google_drive_settings(mask=False)
    if not google_drive_upload_enabled(settings):
        await db.stored_files.update_one(
            {"id": file_id},
            {"$set": {"drive_sync_status": "not_configured", "updated_at": now_iso()}},
        )
        await refresh_embedded_file_references(file_id)
        return
    local_path = Path(file_doc.get("local_path", ""))
    if not file_doc.get("local_path") or not local_path.exists():
        await db.stored_files.update_one(
            {"id": file_id},
            {
                "$set": {
                    "upload_status": "drive_upload_failed",
                    "drive_sync_status": "failed",
                    "drive_error": "File lokal tidak ditemukan untuk sinkron Google Drive.",
                    "updated_at": now_iso(),
                }
            },
        )
        await refresh_embedded_file_references(file_id)
        return
    try:
        drive_doc = await asyncio.to_thread(
            upload_to_drive,
            str(local_path),
            local_path.name,
            file_doc.get("mime_type") or "application/octet-stream",
            file_doc.get("drive_hierarchy") or [],
            settings,
            file_doc.get("lecturer_email", ""),
        )
        await db.stored_files.update_one(
            {"id": file_id},
            {
                "$set": {
                    **drive_doc,
                    "storage_provider": "google_drive",
                    "upload_status": "uploaded_to_drive",
                    "drive_sync_status": "synced",
                    "updated_at": now_iso(),
                },
                "$unset": {"drive_error": ""},
            },
        )
        if file_doc.get("lecturer_id") and drive_doc.get("drive_lecturer_folder_id"):
            await db.users.update_one(
                {"id": file_doc["lecturer_id"], "role": "lecturer"},
                {
                    "$set": {
                        "drive_folder_id": drive_doc.get("drive_lecturer_folder_id", ""),
                        "drive_folder_name": file_doc.get("lecturer_folder_name", ""),
                        "drive_access_status": drive_doc.get("drive_access_status", "app_only"),
                        "drive_permission_id": drive_doc.get("drive_permission_id", ""),
                        "drive_permission_email": drive_doc.get("drive_permission_email", ""),
                        "drive_permission_role": drive_doc.get("drive_permission_role", ""),
                        "drive_access_error": drive_doc.get("drive_access_error", ""),
                        "drive_access_updated_at": now_iso(),
                    }
                },
            )
        await refresh_embedded_file_references(file_id)
    except Exception as exc:
        logger.exception("Sinkron Google Drive background gagal untuk %s: %s", file_doc.get("file_name", file_id), exc)
        drive_error = google_drive_error_message(exc)
        await db.stored_files.update_one(
            {"id": file_id},
            {
                "$set": {
                    "upload_status": "drive_upload_failed",
                    "drive_sync_status": "failed",
                    "drive_error": drive_error,
                    "updated_at": now_iso(),
                }
            },
        )
        await refresh_embedded_file_references(file_id)


async def save_uploaded_file_record(
    upload: UploadFile,
    hierarchy: List[str],
    owner_code: str,
    owner_name: str,
    uploaded_by: str,
    submission_id: str = "",
    assignment_id: str = "",
    record_type: str = "submission",
    sync_drive: bool = True,
    background_tasks: Optional[BackgroundTasks] = None,
    async_drive: bool = False,
    lecturer_id: str = "",
) -> Dict[str, Any]:
    drive_settings = await get_google_drive_settings(mask=False) if sync_drive else {}
    if sync_drive and google_drive_upload_required(drive_settings) and not google_drive_upload_enabled(drive_settings):
        raise HTTPException(
            status_code=503,
            detail="Google Drive belum dikonfigurasi. Isi konfigurasi Google Drive di menu admin terlebih dahulu.",
        )

    lecturer = {}
    if lecturer_id:
        lecturer = await db.users.find_one(
            {"id": lecturer_id, "role": "lecturer", "status": {"$ne": "deleted"}},
            {"_id": 0, "id": 1, "employee_id": 1, "name": 1, "email": 1, "drive_folder_name": 1},
        ) or {}
    lecturer_folder = lecturer_drive_folder_name(lecturer) if lecturer else ""
    scoped_hierarchy = ["Dosen", lecturer_folder, *hierarchy] if lecturer_folder else list(hierarchy)

    file_id = new_id()
    original_filename = upload.filename or "upload.bin"
    mime_type = upload.content_type or "application/octet-stream"
    local_path, folder_path, storage_path = build_local_file_path(
        scoped_hierarchy,
        owner_code,
        owner_name,
        file_id,
        original_filename,
    )
    local_path.parent.mkdir(parents=True, exist_ok=True)
    size = 0
    with local_path.open("wb") as output:
        while True:
            chunk = await upload.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            output.write(chunk)
    await upload.close()
    owner_folder = safe_path_segment(f"{owner_code or 'NO-ID'} - {owner_name}")
    drive_hierarchy = [*scoped_hierarchy, owner_folder]
    file_doc = {
        "id": file_id,
        "file_name": original_filename,
        "file_id": file_id,
        **local_file_urls(file_id),
        "mime_type": mime_type,
        "size": size,
        "folder_path": folder_path,
        "storage_path": storage_path,
        "storage_provider": "server_local",
        "local_path": str(local_path),
        "uploaded_by": uploaded_by,
        "uploaded_at": now_iso(),
        "submission_id": submission_id,
        "assignment_id": assignment_id,
        "record_type": record_type,
        "lecturer_id": lecturer.get("id", lecturer_id),
        "lecturer_name": lecturer.get("name", ""),
        "lecturer_code": lecturer.get("employee_id", ""),
        "lecturer_email": lecturer.get("email", ""),
        "lecturer_folder_name": lecturer_folder,
        "upload_status": "stored_on_server",
        "drive_sync_status": "not_configured",
        "drive_hierarchy": drive_hierarchy,
    }

    if sync_drive and google_drive_upload_enabled(drive_settings):
        if async_drive and background_tasks:
            file_doc.update({"drive_sync_status": "pending"})
            await db.stored_files.insert_one(file_doc)
            background_tasks.add_task(sync_stored_file_to_drive, file_id)
            return public_doc(file_doc.copy())
        try:
            drive_doc = await asyncio.to_thread(
                upload_to_drive,
                str(local_path),
                local_path.name,
                mime_type,
                drive_hierarchy,
                drive_settings,
                lecturer.get("email", ""),
            )
            file_doc.update(
                {
                    **drive_doc,
                    "storage_provider": "google_drive",
                    "upload_status": "uploaded_to_drive",
                    "drive_sync_status": "synced",
                }
            )
        except Exception as exc:
            logger.exception("Upload Google Drive gagal untuk %s: %s", original_filename, exc)
            if google_drive_upload_required(drive_settings):
                local_path.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=503,
                    detail="Upload ke Google Drive gagal. File belum disimpan sebagai submission.",
                )
            file_doc.update(
                {
                    "upload_status": "drive_upload_failed",
                    "drive_sync_status": "failed",
                    "drive_error": str(exc)[:500],
                }
            )

    await db.stored_files.insert_one(file_doc)
    return public_doc(file_doc.copy())


async def enrich_class_payload(class_doc: Dict[str, Any]) -> Dict[str, Any]:
    course = await db.courses.find_one({"id": class_doc.get("course_id")}, {"_id": 0})
    class_doc["course_name"] = course.get("name", class_doc.get("course_name", "")) if course else class_doc.get("course_name", "")
    class_doc["program_id"] = course.get("program_id", class_doc.get("program_id", "")) if course else class_doc.get("program_id", "")
    class_doc["program_name"] = course.get("program_name", class_doc.get("program_name", "")) if course else class_doc.get("program_name", "")
    class_doc["student_count"] = len(class_doc.get("student_ids", []))
    return class_doc


async def enrich_course_payload(course_doc: Dict[str, Any]) -> Dict[str, Any]:
    program = await db.programs.find_one({"id": course_doc.get("program_id")}, {"_id": 0})
    course_doc["program_name"] = program.get("name", course_doc.get("program_name", "")) if program else course_doc.get("program_name", "")
    return course_doc


async def enrich_material_payload(material_doc: Dict[str, Any]) -> Dict[str, Any]:
    if isinstance(material_doc.get("attachment"), dict):
        material_doc["attachment"] = enrich_file_urls(material_doc["attachment"])
    class_doc = await db.classes.find_one({"id": material_doc.get("class_id")}, {"_id": 0})
    class_doc = await enrich_class_payload(class_doc) if class_doc else {}
    linked_assignment = await db.assignments.find_one(
        {"material_id": material_doc.get("id")},
        {"_id": 0, "class_name": 1, "course_id": 1, "course_name": 1},
    ) or {}
    material_doc.update(
        {
            "class_name": class_doc.get("name") or linked_assignment.get("class_name", ""),
            "class_code": class_doc.get("class_code", ""),
            "class_status": class_doc.get("status", ""),
            "course_id": class_doc.get("course_id") or linked_assignment.get("course_id", ""),
            "course_name": class_doc.get("course_name") or linked_assignment.get("course_name", ""),
            "academic_year": class_doc.get("academic_year", ""),
            "semester": class_doc.get("semester", ""),
        }
    )
    return material_doc


async def enrich_materials_batch(materials: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not materials:
        return []
    class_ids = list(set(m.get("class_id", "") for m in materials if m.get("class_id")))
    material_ids = [m.get("id", "") for m in materials if m.get("id")]
    classes = await db.classes.find({"id": {"$in": class_ids}}, {"_id": 0}).to_list(1000) if class_ids else []
    classes_map: Dict[str, Dict[str, Any]] = {c["id"]: c for c in classes}
    course_ids = list(set(c.get("course_id", "") for c in classes if c.get("course_id")))
    courses = await db.courses.find({"id": {"$in": course_ids}}, {"_id": 0}).to_list(1000) if course_ids else []
    courses_map: Dict[str, Dict[str, Any]] = {c["id"]: c for c in courses}
    for class_doc in classes_map.values():
        course = courses_map.get(class_doc.get("course_id", ""))
        if course:
            class_doc["course_name"] = course.get("name", class_doc.get("course_name", ""))
            class_doc["program_id"] = course.get("program_id", class_doc.get("program_id", ""))
            class_doc["program_name"] = course.get("program_name", class_doc.get("program_name", ""))
        class_doc["student_count"] = len(class_doc.get("student_ids", []))
    linked_assignments = await db.assignments.find(
        {"material_id": {"$in": material_ids}},
        {"_id": 0, "material_id": 1, "class_name": 1, "course_id": 1, "course_name": 1},
    ).to_list(1000) if material_ids else []
    assignments_by_material: Dict[str, Dict[str, Any]] = {}
    for a in linked_assignments:
        mid = a.get("material_id", "")
        if mid and mid not in assignments_by_material:
            assignments_by_material[mid] = a
    for material in materials:
        if isinstance(material.get("attachment"), dict):
            material["attachment"] = enrich_file_urls(material["attachment"])
        class_doc = classes_map.get(material.get("class_id"), {})
        linked = assignments_by_material.get(material.get("id", ""), {})
        material.update({
            "class_name": class_doc.get("name") or linked.get("class_name", ""),
            "class_code": class_doc.get("class_code", ""),
            "class_status": class_doc.get("status", ""),
            "course_id": class_doc.get("course_id") or linked.get("course_id", ""),
            "course_name": class_doc.get("course_name") or linked.get("course_name", ""),
            "academic_year": class_doc.get("academic_year", ""),
            "semester": class_doc.get("semester", ""),
        })
    return materials


async def material_meeting_label_map(class_ids: Optional[List[str]] = None) -> Dict[str, str]:
    query: Dict[str, Any] = {}
    if class_ids is not None:
        query = {"class_id": {"$in": list(set(class_ids))}}
    material_docs = (
        await db.materials.find(query, {"_id": 0, "id": 1, "class_id": 1})
        .sort([("created_at", 1), ("id", 1)])
        .to_list(10000)
    )
    counts: Dict[str, int] = {}
    labels: Dict[str, str] = {}
    for material in material_docs:
        class_id = material.get("class_id", "")
        counts[class_id] = counts.get(class_id, 0) + 1
        labels[material["id"]] = f"Pertemuan {counts[class_id]}"
    return labels


def normalized_material_payload(payload: MaterialInput) -> Dict[str, Any]:
    doc = payload.model_dump()
    meeting_type = str(doc.get("meeting_type") or "offline").strip().lower()
    if meeting_type not in {"offline", "online"}:
        raise HTTPException(status_code=400, detail="Metode pertemuan harus offline atau online")
    meeting_url = str(doc.get("meeting_url") or "").strip()
    if meeting_url and not re.fullmatch(r"https://meet\.google\.com/[a-z0-9-]+", meeting_url, re.IGNORECASE):
        raise HTTPException(status_code=400, detail="Link Google Meet tidak valid")
    doc["meeting_type"] = meeting_type
    doc["meeting_url"] = meeting_url if meeting_type == "online" else ""
    return doc


@api_router.get("/")
async def root():
    return {
        "message": "E-Learning Dosen API aktif",
        "version": version_payload(),
        **await storage_status_summary(),
    }


@api_router.get("/version")
async def app_version():
    schema_versions: list[str] = []
    try:
        rows = await db.pool.fetch("SELECT version FROM app_schema_migrations ORDER BY version")
        schema_versions = [str(row["version"]) for row in rows]
    except Exception:
        # Version information should stay available while the database is starting.
        schema_versions = []
    return version_payload(schema_versions=schema_versions)


@api_router.get("/auth/sso/config")
async def sso_config():
    settings = oidc_settings()
    login_url = ""
    if settings["enabled"] and settings["redirect_uri"]:
        callback = urlsplit(settings["redirect_uri"])
        login_url = f"{callback.scheme}://{callback.netloc}/api/auth/sso/login"
    return {
        "enabled": bool(settings["enabled"]),
        "provider": "SCI-ID",
        "login_url": login_url,
        "local_login_enabled": bool(settings["local_login_enabled"]),
    }


@api_router.get("/auth/sso/login")
async def sso_login():
    settings = oidc_require_settings()
    metadata = await oidc_discovery()
    state = secrets.token_urlsafe(48)
    nonce = secrets.token_urlsafe(48)
    verifier = secrets.token_urlsafe(64)
    challenge = base64url_encode(hashlib.sha256(verifier.encode("ascii")).digest())
    await db.oidc_flows.insert_one(
        {
            "id": new_id(),
            "state_hash": hashlib.sha256(state.encode("utf-8")).hexdigest(),
            "nonce": nonce,
            "code_verifier": verifier,
            "created_at": now_iso(),
            "expires_at": datetime.now(timezone.utc) + timedelta(minutes=5),
        }
    )
    query = urlencode(
        {
            "client_id": settings["client_id"],
            "redirect_uri": settings["redirect_uri"],
            "response_type": "code",
            "scope": settings["scopes"],
            "state": state,
            "nonce": nonce,
            "code_challenge": challenge,
            "code_challenge_method": "S256",
        }
    )
    return RedirectResponse(f"{metadata['authorization_endpoint']}?{query}", status_code=307)


@api_router.get("/auth/sso/callback")
async def sso_callback(
    code: str = "",
    state: str = "",
    error: str = "",
    error_description: str = "",
):
    try:
        oidc_require_settings()
        if not state:
            raise HTTPException(status_code=400, detail="State login SCI-ID tidak tersedia")
        flow = await db.oidc_flows.find_one_and_delete(
            {"state_hash": hashlib.sha256(state.encode("utf-8")).hexdigest()},
            {"_id": 0},
        )
        if not flow:
            raise HTTPException(status_code=400, detail="State login SCI-ID tidak valid atau sudah digunakan")
        expires_at = flow.get("expires_at")
        if isinstance(expires_at, datetime):
            expires_at = expires_at.replace(tzinfo=timezone.utc) if expires_at.tzinfo is None else expires_at.astimezone(timezone.utc)
            if expires_at <= datetime.now(timezone.utc):
                raise HTTPException(status_code=400, detail="Permintaan login SCI-ID sudah kedaluwarsa")
        if error:
            raise HTTPException(status_code=401, detail=error_description or f"SCI-ID menolak login: {error}")
        if not code:
            raise HTTPException(status_code=400, detail="Authorization code SCI-ID tidak tersedia")
        settings = oidc_require_settings()
        metadata = await oidc_discovery()
        token_payload = {
            "grant_type": "authorization_code",
            "client_id": settings["client_id"],
            "redirect_uri": settings["redirect_uri"],
            "code": code,
            "code_verifier": flow["code_verifier"],
        }
        if settings["client_secret"]:
            token_payload["client_secret"] = settings["client_secret"]
        try:
            async with httpx.AsyncClient(timeout=10) as http:
                token_response = await http.post(
                    metadata["token_endpoint"],
                    data=token_payload,
                    headers={"Accept": "application/json"},
                )
                token_response.raise_for_status()
                token_data = token_response.json()
        except (httpx.HTTPError, ValueError) as exc:
            logger.warning("Penukaran authorization code SCI-ID gagal: %s", type(exc).__name__)
            raise HTTPException(status_code=401, detail="Authorization code SCI-ID tidak dapat ditukar") from exc
        id_token = token_data.get("id_token") if isinstance(token_data, dict) else None
        if not isinstance(id_token, str) or not id_token:
            raise HTTPException(status_code=401, detail="SCI-ID tidak mengembalikan ID token")
        claims = await validate_oidc_id_token(id_token, metadata, flow["nonce"])
        user = await provision_oidc_user(claims)
        ticket = secrets.token_urlsafe(48)
        await db.oidc_login_tickets.insert_one(
            {
                "ticket_hash": hashlib.sha256(ticket.encode("utf-8")).hexdigest(),
                "user_id": user["id"],
                "sso_subject": claims["sub"],
                "created_at": now_iso(),
                "expires_at": datetime.now(timezone.utc) + timedelta(minutes=1),
            }
        )
        return RedirectResponse(oidc_frontend_redirect(sso_ticket=ticket), status_code=303)
    except HTTPException as exc:
        logger.warning("Callback SCI-ID ditolak: %s", exc.detail)
        return RedirectResponse(oidc_frontend_redirect(sso_error=str(exc.detail)), status_code=303)
    except Exception:
        logger.exception("Callback SCI-ID gagal")
        return RedirectResponse(oidc_frontend_redirect(sso_error="Login SCI-ID gagal diproses"), status_code=303)


@api_router.post("/auth/sso/exchange")
async def sso_exchange(payload: SsoExchangeInput):
    ticket = await db.oidc_login_tickets.find_one_and_delete(
        {"ticket_hash": hashlib.sha256(payload.ticket.encode("utf-8")).hexdigest()},
        {"_id": 0},
    )
    if not ticket:
        raise HTTPException(status_code=401, detail="Tiket login SCI-ID tidak valid atau sudah digunakan")
    expires_at = ticket.get("expires_at")
    if isinstance(expires_at, datetime):
        expires_at = expires_at.replace(tzinfo=timezone.utc) if expires_at.tzinfo is None else expires_at.astimezone(timezone.utc)
        if expires_at <= datetime.now(timezone.utc):
            raise HTTPException(status_code=401, detail="Tiket login SCI-ID sudah kedaluwarsa")
    user = await find_user(ticket["user_id"])
    token = new_id() + new_id()
    await db.sessions.insert_one(
        {
            "token": token,
            "user_id": user["id"],
            "auth_source": "sso",
            "sso_subject": ticket.get("sso_subject", ""),
            "created_at": now_iso(),
        }
    )
    return {"token": token, "user": public_doc(user)}


@api_router.post("/auth/login")
async def login(payload: LoginInput):
    if not oidc_settings()["local_login_enabled"]:
        raise HTTPException(status_code=403, detail="Login lokal dinonaktifkan. Gunakan SCI-ID.")
    identifier = (payload.identifier or payload.email or "").strip().lower()
    if not identifier:
        raise HTTPException(status_code=400, detail="Username, NIM, nomor HP, atau email diperlukan")
    user = await db.users.find_one(identity_query(identifier), {"_id": 0})
    if not user or not user.get("password_hash") or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Identitas login atau password salah")
    if user.get("status", "active") != "active":
        raise HTTPException(status_code=403, detail="Akun tidak aktif. Hubungi admin kampus.")
    token = new_id() + new_id()
    await db.sessions.insert_one({"token": token, "user_id": user["id"], "created_at": now_iso()})
    await db.users.update_one({"id": user["id"]}, {"$set": {"last_login_at": now_iso()}})
    user = public_doc(user)
    return {"token": token, "user": user}


@api_router.post("/auth/register-student")
async def register_student(payload: RegisterStudentInput):
    email = payload.email.lower()
    username = (payload.username or payload.nim).strip().lower()
    whatsapp = payload.whatsapp.strip()
    existing = await db.users.find_one(
        {
            "$or": [
                {"email": email},
                {"username": username},
                {"nim": payload.nim},
                {"whatsapp": whatsapp} if whatsapp else {"_never": "_never"},
            ]
        },
        {"_id": 0},
    )
    if existing:
        raise HTTPException(status_code=409, detail="Email, username, NIM, atau WhatsApp sudah terdaftar")
    student_id = new_id()
    doc = {
        "id": student_id,
        "role": "student",
        "username": username,
        "nim": payload.nim,
        "name": payload.name,
        "email": email,
        "whatsapp": whatsapp,
        "password_hash": hash_password(payload.password),
        "status": "active",
        "class_ids": [],
        "created_at": now_iso(),
        "last_login_at": "",
    }
    await db.users.insert_one(doc)
    return await login(LoginInput(identifier=email, password=payload.password))


@api_router.post("/auth/forgot-password")
async def forgot_password(payload: ForgotPasswordInput, background_tasks: BackgroundTasks):
    identifier = payload.identifier.strip().lower()
    user = await db.users.find_one(identity_query(identifier), {"_id": 0})
    if not user:
        return {
            "ok": True,
            "message": "Jika akun ditemukan, permintaan reset password akan diproses.",
            "otp_delivery": {"status": "not_found", "message_id": ""},
            "email_delivery": {"status": "not_found", "error": ""},
        }
    response: Dict[str, Any] = {
        "ok": True,
        "message": "Jika akun ditemukan, permintaan reset password akan diproses.",
    }
    request_doc = {
        "id": new_id(),
        "identifier": identifier,
        "user_id": user["id"],
        "status": "otp_created",
        "requested_at": now_iso(),
    }
    otp = generate_otp()
    expires_at = (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat()
    reset_id = new_id()
    settings = await get_whatsapp_settings(mask=False)
    email_settings = await get_email_settings(mask=False)
    reset_link = build_password_reset_link(settings.get("app_url") or os.environ.get("APP_URL", ""), identifier)
    await db.password_reset_otps.insert_one(
        {
            "id": reset_id,
            "user_id": user["id"],
            "identifier": identifier,
            "otp": otp,
            "status": "active",
            "expires_at": expires_at,
            "created_at": now_iso(),
            "used_at": "",
        }
    )
    request_doc.update({"otp_id": reset_id, "expires_at": expires_at})
    email_delivery: Dict[str, Any] = {"status": "not_attempted", "error": ""}
    if email_settings.get("enabled") and user.get("email"):
        html_body = (
            f"<h3>Reset Password</h3>"
            f"<p>Halo {user.get('name', '')},</p>"
            f"<p>Kode OTP reset password Anda: <strong>{otp}</strong></p>"
            f"<p>Kode ini berlaku selama 10 menit.</p>"
            f"<p>Link reset: <a href='{reset_link}'>{reset_link}</a></p>"
            f"<p>Jika Anda tidak meminta reset password, abaikan email ini.</p>"
        )
        email_result = await send_email_message(user["email"], "Kode OTP Reset Password - E-Learning Dosen", html_body)
        email_delivery = {
            "status": "sent" if email_result["ok"] else "failed",
            "error": email_result.get("error", "") if not email_result["ok"] else "",
        }
        if email_result["ok"]:
            await db.email_messages.insert_one({
                "id": new_id(),
                "to": user["email"],
                "subject": "Kode OTP Reset Password - E-Learning Dosen",
                "message_type": "password_reset_otp",
                "ref_id": reset_id,
                "status": "sent",
                "created_at": now_iso(),
                "sent_at": now_iso(),
            })
        else:
            await db.email_messages.insert_one({
                "id": new_id(),
                "to": user["email"],
                "subject": "Kode OTP Reset Password - E-Learning Dosen",
                "message_type": "password_reset_otp",
                "ref_id": reset_id,
                "status": "failed",
                "error": email_result.get("error", ""),
                "created_at": now_iso(),
            })
    elif not user.get("email"):
        email_delivery = {"status": "no_email", "error": ""}
    response["email_delivery"] = email_delivery
    if user.get("whatsapp"):
        template = settings.get("otp_template") or "Kode OTP reset password Anda: {code}. Berlaku {minutes} menit. Link: {link}"
        message = template.format(code=otp, minutes=10, link=reset_link, name=user.get("name", ""))
        queued = await enqueue_whatsapp_message(user.get("whatsapp", ""), message, "password_reset_otp", reset_id)
        request_doc.update({"delivery_status": queued.get("status", ""), "message_id": queued.get("id", "")})
        delivery = public_whatsapp_delivery_status(queued)
        response["otp_delivery"] = delivery
        if queued.get("status") == "pending":
            background_tasks.add_task(send_whatsapp_message, queued["id"])
        response["message"] = forgot_password_response_message(delivery)
    else:
        request_doc.update({"delivery_status": "no_whatsapp", "message_id": ""})
        delivery = {"message_id": "", "status": "no_whatsapp", "provider": "", "created_at": now_iso(), "sent_at": "", "error": ""}
        response["otp_delivery"] = delivery
        if email_delivery.get("status") == "sent":
            response["message"] = "OTP sudah dikirim ke email Anda."
        elif email_delivery.get("status") == "failed":
            response["message"] = "Permintaan reset password diproses, tetapi gagal mengirim email."
        else:
            response["message"] = forgot_password_response_message(delivery)
    if local_reset_otp_enabled() and request_doc.get("delivery_status", "") in {"pending_config", "no_whatsapp", ""}:
        response["local_otp_available"] = True
    await db.password_reset_requests.insert_one(request_doc)
    return response


@api_router.get("/auth/forgot-password/messages/{message_id}")
async def forgot_password_message_status(message_id: str):
    message = await db.whatsapp_messages.find_one({"id": message_id, "message_type": "password_reset_otp"}, {"_id": 0})
    if not message:
        raise HTTPException(status_code=404, detail="Status antrean OTP tidak ditemukan")
    return public_whatsapp_delivery_status(message)


@api_router.post("/auth/reset-password-otp")
async def reset_password_otp(payload: ResetPasswordOtpInput):
    identifier = payload.identifier.strip().lower()
    user = await db.users.find_one(identity_query(identifier), {"_id": 0})
    if not user:
        raise HTTPException(status_code=400, detail="OTP atau akun tidak valid")
    otp_doc = await db.password_reset_otps.find_one(
        {"user_id": user["id"], "otp": payload.otp.strip(), "status": "active"}, {"_id": 0}
    )
    if not otp_doc:
        raise HTTPException(status_code=400, detail="OTP tidak valid")
    if datetime.fromisoformat(otp_doc["expires_at"].replace("Z", "+00:00")) < datetime.now(timezone.utc):
        await db.password_reset_otps.update_one({"id": otp_doc["id"]}, {"$set": {"status": "expired"}})
        raise HTTPException(status_code=400, detail="OTP sudah kedaluwarsa")
    await db.users.update_one({"id": user["id"]}, {"$set": {"password_hash": hash_password(payload.new_password), "password_changed_at": now_iso()}})
    await db.password_reset_otps.update_one({"id": otp_doc["id"]}, {"$set": {"status": "used", "used_at": now_iso()}})
    return {"ok": True}


@api_router.post("/auth/change-password")
async def change_password(payload: ChangePasswordInput, user: Dict[str, Any] = Depends(get_current_user)):
    full_user = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    if not full_user or not verify_password(payload.current_password, full_user["password_hash"]):
        raise HTTPException(status_code=400, detail="Password lama salah")
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"password_hash": hash_password(payload.new_password), "password_changed_at": now_iso()}},
    )
    return {"ok": True}


@api_router.post("/auth/join-class")
async def join_class(payload: JoinClassInput):
    class_code = clean_code(payload.class_code)
    class_doc = await db.classes.find_one({"class_code": class_code, "status": "active"}, {"_id": 0})
    if not class_doc:
        raise HTTPException(status_code=404, detail="Kode kelas tidak ditemukan")
    existing = await db.users.find_one({"email": payload.email.lower()}, {"_id": 0})
    if existing:
        student_id = existing["id"]
        await db.users.update_one({"id": student_id}, {"$addToSet": {"class_ids": class_doc["id"]}})
    else:
        student_id = new_id()
        await db.users.insert_one(
            {
                "id": student_id,
                "role": "student",
                "nim": payload.nim,
                "name": payload.name,
                "email": payload.email.lower(),
                "whatsapp": payload.whatsapp,
                "password_hash": hash_password(payload.password),
                "status": "active",
                "class_ids": [class_doc["id"]],
                "created_at": now_iso(),
                "last_login_at": "",
            }
        )
    await db.classes.update_one({"id": class_doc["id"]}, {"$addToSet": {"student_ids": student_id}})
    return await login(LoginInput(identifier=payload.email, password=payload.password))


@api_router.get("/auth/me")
async def me(user: Dict[str, Any] = Depends(get_current_user)):
    return public_doc(user)


@api_router.put("/auth/me")
async def update_profile(payload: ProfileInput, user: Dict[str, Any] = Depends(get_current_user)):
    name = payload.name.strip()
    username = payload.username.strip().lower()
    email = str(payload.email).strip().lower()
    whatsapp = payload.whatsapp.strip()
    if not name or not username:
        raise HTTPException(status_code=400, detail="Nama dan username wajib diisi")
    current_user = await db.users.find_one({"id": user["id"]}, {"_id": 0, "whatsapp": 1})
    current_whatsapp = (current_user or {}).get("whatsapp", "")
    candidates: List[Dict[str, Any]] = [{"email": email}, {"username": username}]
    if whatsapp and normalize_phone(whatsapp) != normalize_phone(current_whatsapp):
        candidates.append({"whatsapp": whatsapp})
    if len(candidates) > 2:
        existing = await db.users.find_one(
            {"id": {"$ne": user["id"]}, "$or": candidates},
            {"_id": 0, "id": 1},
        )
        if existing:
            raise HTTPException(status_code=409, detail="Email, username, atau WhatsApp sudah digunakan akun lain")
    await db.users.update_one(
        {"id": user["id"]},
        {
            "$set": {
                "name": name,
                "username": username,
                "email": email,
                "whatsapp": whatsapp,
                "profile_updated_at": now_iso(),
            }
        },
    )
    if user.get("role") in {"admin", "lecturer"} and name != user.get("name"):
        await db.classes.update_many({"lecturer_id": user["id"]}, {"$set": {"lecturer_name": name}})
        for collection in [db.materials, db.assignments, db.submissions, db.enrollment_requests, db.reminder_logs]:
            await collection.update_many({"lecturer_id": user["id"]}, {"$set": {"lecturer_name": name}})
    updated = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    return public_doc(updated)


@api_router.post("/auth/logout")
async def logout(authorization: Optional[str] = Header(None)):
    logout_url = ""
    if authorization and authorization.startswith("Bearer "):
        token = authorization.replace("Bearer ", "", 1).strip()
        session = await db.sessions.find_one_and_delete({"token": token}, {"_id": 0})
        if session and session.get("auth_source") == "sso" and oidc_settings()["enabled"]:
            try:
                settings = oidc_require_settings()
                metadata = await oidc_discovery()
                logout_url = f"{metadata['end_session_endpoint']}?{urlencode({'client_id': settings['client_id'], 'post_logout_redirect_uri': settings['frontend_url']})}"
            except HTTPException:
                logout_url = ""
    return {"ok": True, "logout_url": logout_url}


@api_router.get("/password-reset-requests")
async def list_password_reset_requests(_: Dict[str, Any] = Depends(require_admin)):
    return await db.password_reset_requests.find({}, {"_id": 0}).sort("requested_at", -1).to_list(1000)


@api_router.get("/lecturers")
async def list_lecturers(_: Dict[str, Any] = Depends(require_campus_admin)):
    lecturers = await db.users.find(
        {"role": "lecturer", "status": {"$ne": "deleted"}}, {"_id": 0, "password_hash": 0}
    ).sort("name", 1).to_list(2000)
    counts = await db.classes.aggregate(
        [
            {"$match": {"status": {"$ne": "deleted"}, "lecturer_id": {"$ne": ""}}},
            {"$group": {"_id": "$lecturer_id", "total": {"$sum": 1}}},
        ]
    ).to_list(2000)
    class_counts = {item["_id"]: item["total"] for item in counts}
    storage_counts = await db.stored_files.aggregate(
        [
            {"$match": {"lecturer_id": {"$nin": ["", None]}}},
            {
                "$group": {
                    "_id": "$lecturer_id",
                    "file_count": {"$sum": 1},
                    "storage_bytes": {"$sum": {"$ifNull": ["$size", 0]}},
                    "drive_synced_count": {
                        "$sum": {"$cond": [{"$eq": ["$drive_sync_status", "synced"]}, 1, 0]}
                    },
                    "drive_failed_count": {
                        "$sum": {"$cond": [{"$eq": ["$drive_sync_status", "failed"]}, 1, 0]}
                    },
                }
            },
        ]
    ).to_list(2000)
    lecturer_storage = {item["_id"]: item for item in storage_counts}
    for lecturer in lecturers:
        lecturer["class_count"] = class_counts.get(lecturer["id"], 0)
        storage = lecturer_storage.get(lecturer["id"], {})
        lecturer["storage_file_count"] = storage.get("file_count", 0)
        lecturer["storage_bytes"] = storage.get("storage_bytes", 0)
        lecturer["drive_synced_count"] = storage.get("drive_synced_count", 0)
        lecturer["drive_failed_count"] = storage.get("drive_failed_count", 0)
    return lecturers


@api_router.post("/lecturers")
async def create_lecturer(
    payload: LecturerInput,
    background_tasks: BackgroundTasks,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    if payload.status not in {"active", "inactive"}:
        raise HTTPException(status_code=400, detail="Status dosen harus active atau inactive")
    username = payload.username.strip().lower()
    email = payload.email.lower()
    duplicate = await db.users.find_one({"$or": [{"username": username}, {"email": email}]}, {"_id": 0, "id": 1})
    if duplicate:
        raise HTTPException(status_code=409, detail="Username atau email dosen sudah digunakan")
    doc = {
        "id": new_id(),
        "role": "lecturer",
        "employee_id": payload.employee_id.strip(),
        "username": username,
        "name": payload.name.strip(),
        "email": email,
        "whatsapp": payload.whatsapp.strip(),
        "password_hash": hash_password(payload.password),
        "status": payload.status,
        "created_at": now_iso(),
        "last_login_at": "",
    }
    await db.users.insert_one(doc)
    background_tasks.add_task(reconcile_lecturer_drive_access, doc["id"])
    return public_doc(doc)


@api_router.put("/lecturers/{lecturer_id}")
async def update_lecturer(
    lecturer_id: str,
    payload: LecturerUpdateInput,
    background_tasks: BackgroundTasks,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    existing = await db.users.find_one({"id": lecturer_id, "role": "lecturer", "status": {"$ne": "deleted"}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Dosen tidak ditemukan")
    if payload.status not in {"active", "inactive"}:
        raise HTTPException(status_code=400, detail="Status dosen harus active atau inactive")
    username = payload.username.strip().lower()
    email = payload.email.lower()
    duplicate = await db.users.find_one(
        {"id": {"$ne": lecturer_id}, "$or": [{"username": username}, {"email": email}]}, {"_id": 0, "id": 1}
    )
    if duplicate:
        raise HTTPException(status_code=409, detail="Username atau email dosen sudah digunakan")
    update = payload.model_dump()
    update.update({"username": username, "email": email, "updated_at": now_iso()})
    await db.users.update_one({"id": lecturer_id}, {"$set": update})
    if payload.name != existing.get("name"):
        await db.classes.update_many({"lecturer_id": lecturer_id}, {"$set": {"lecturer_name": payload.name}})
        for collection in [db.materials, db.assignments, db.submissions, db.enrollment_requests, db.reminder_logs]:
            await collection.update_many({"lecturer_id": lecturer_id}, {"$set": {"lecturer_name": payload.name}})
    background_tasks.add_task(reconcile_lecturer_drive_access, lecturer_id)
    return public_doc(await db.users.find_one({"id": lecturer_id}, {"_id": 0}))


@api_router.post("/lecturers/{lecturer_id}/reset-password")
async def reset_lecturer_password(
    lecturer_id: str, payload: ResetPasswordInput, _: Dict[str, Any] = Depends(require_campus_admin)
):
    lecturer = await db.users.find_one({"id": lecturer_id, "role": "lecturer", "status": {"$ne": "deleted"}}, {"_id": 0})
    if not lecturer:
        raise HTTPException(status_code=404, detail="Dosen tidak ditemukan")
    password = payload.password.strip() or "Dosen123!"
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="Password dosen minimal 6 karakter")
    await db.users.update_one(
        {"id": lecturer_id}, {"$set": {"password_hash": hash_password(password), "password_reset_at": now_iso()}}
    )
    await db.sessions.delete_many({"user_id": lecturer_id})
    return {"ok": True, "temporary_password": password}


@api_router.delete("/lecturers/{lecturer_id}")
async def delete_lecturer(
    lecturer_id: str,
    background_tasks: BackgroundTasks,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    active_classes = await db.classes.count_documents({"lecturer_id": lecturer_id, "status": "active"})
    if active_classes:
        raise HTTPException(status_code=400, detail="Dosen masih memiliki kelas aktif. Arsipkan atau pindahkan kelas terlebih dahulu.")
    result = await db.users.update_one(
        {"id": lecturer_id, "role": "lecturer"}, {"$set": {"status": "deleted", "deleted_at": now_iso()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dosen tidak ditemukan")
    await db.sessions.delete_many({"user_id": lecturer_id})
    background_tasks.add_task(reconcile_lecturer_drive_access, lecturer_id)
    return {"ok": True}


@api_router.get("/whatsapp/settings")
async def whatsapp_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    return await get_whatsapp_settings(mask=True)


@api_router.put("/whatsapp/settings")
async def update_whatsapp_settings(payload: WhatsAppSettingsInput, user: Dict[str, Any] = Depends(require_campus_admin)):
    existing = await get_whatsapp_settings(mask=False)
    doc = payload.model_dump()
    if not doc.get("fonnte_token") and existing.get("fonnte_token"):
        doc["fonnte_token"] = existing.get("fonnte_token", "")
    if not doc.get("waha_api_key") and existing.get("waha_api_key"):
        doc["waha_api_key"] = existing.get("waha_api_key", "")
    doc.update({"id": "main", "updated_at": now_iso(), "updated_by": user["id"]})
    await db.whatsapp_settings.update_one({"id": "main"}, {"$set": doc}, upsert=True)
    _invalidate_settings_cache("whatsapp_settings")
    return await get_whatsapp_settings(mask=True)


@api_router.get("/whatsapp/messages")
async def whatsapp_messages(_: Dict[str, Any] = Depends(require_campus_admin)):
    return await db.whatsapp_messages.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)


@api_router.get("/whatsapp/waha/status")
async def waha_connection_status(_: Dict[str, Any] = Depends(require_campus_admin)):
    settings = await get_whatsapp_settings(mask=False)
    base_url = normalize_http_base_url(settings.get("waha_base_url", ""))
    session = (settings.get("waha_session") or "default").strip() or "default"
    if settings.get("provider") != "waha":
        return {"ok": False, "provider": settings.get("provider", "disabled"), "detail": "Provider WAHA belum dipilih"}
    if not base_url:
        return {"ok": False, "provider": "waha", "session": session, "detail": "WAHA Base URL belum diisi"}
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            session_doc = await fetch_waha_session_status(client, base_url, session, waha_headers(settings.get("waha_api_key", "")))
        if not session_doc:
            return {"ok": False, "provider": "waha", "base_url": base_url, "session": session, "detail": "Sesi WAHA tidak ditemukan"}
        status = str(session_doc.get("status") or "")
        return {
            "ok": status.upper() == "WORKING",
            "provider": "waha",
            "base_url": base_url,
            "session": session,
            "status": status,
            "me": session_doc.get("me"),
            "detail": "Sesi WAHA siap mengirim pesan" if status.upper() == "WORKING" else f"Sesi WAHA belum siap: {status}",
        }
    except Exception as exc:
        return {"ok": False, "provider": "waha", "base_url": base_url, "session": session, "detail": str(exc)}


@api_router.post("/whatsapp/messages/{message_id}/retry")
async def retry_whatsapp_message(message_id: str, background_tasks: BackgroundTasks, _: Dict[str, Any] = Depends(require_campus_admin)):
    msg = await db.whatsapp_messages.find_one({"id": message_id}, {"_id": 0})
    if not msg:
        raise HTTPException(status_code=404, detail="Pesan tidak ditemukan")
    await db.whatsapp_messages.update_one({"id": message_id}, {"$set": {"status": "pending", "error": "", "response": ""}})
    background_tasks.add_task(send_whatsapp_message, message_id)
    return {"ok": True}


@api_router.get("/email/settings")
async def email_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    return await get_email_settings(mask=True)


@api_router.put("/email/settings")
async def update_email_settings(payload: EmailSettingsInput, user: Dict[str, Any] = Depends(require_campus_admin)):
    existing = await get_email_settings(mask=False)
    doc = payload.model_dump()
    if not doc.get("smtp_password") and existing.get("smtp_password"):
        doc["smtp_password"] = existing.get("smtp_password", "")
    doc.update({"id": "main", "updated_at": now_iso(), "updated_by": user["id"]})
    await db.email_settings.update_one({"id": "main"}, {"$set": doc}, upsert=True)
    _invalidate_settings_cache("email_settings")
    return await get_email_settings(mask=True)


@api_router.post("/email/settings/test")
async def test_email_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    settings = await get_email_settings(mask=False)
    if not settings.get("enabled"):
        raise HTTPException(status_code=400, detail="Email belum diaktifkan")
    test_email = settings.get("smtp_user") or settings.get("from_email", "")
    if not test_email:
        raise HTTPException(status_code=400, detail="Tidak ada email tujuan untuk tes. Isi SMTP user atau from email.")
    result = await send_email_message(
        test_email,
        "Tes Konfigurasi Email - E-Learning Dosen",
        "<h3>Email Berhasil Terkirim</h3><p>Konfigurasi SMTP pada aplikasi E-Learning Dosen sudah benar.</p>",
    )
    if not result["ok"]:
        raise HTTPException(status_code=400, detail=result.get("error", "Gagal mengirim email tes"))
    return {"ok": True, "message": f"Email tes berhasil dikirim ke {test_email}"}


@api_router.get("/sso/settings")
async def get_oidc_admin_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    doc = await db.oidc_settings.find_one({"id": "main"}, {"_id": 0}) or {}
    return oidc_admin_view(oidc_settings(), "admin_ui" if doc else "environment", doc)


@api_router.put("/sso/settings")
async def update_oidc_admin_settings(payload: OidcSettingsInput, user: Dict[str, Any] = Depends(require_campus_admin)):
    discovery_url = validate_oidc_url(payload.discovery_url, "Discovery URL", required=payload.enabled)
    issuer = validate_oidc_url(payload.issuer, "Issuer", required=False)
    redirect_uri = validate_oidc_url(payload.redirect_uri, "Redirect URI", required=payload.enabled)
    frontend_url = validate_oidc_url(payload.frontend_url, "Frontend URL", required=payload.enabled)
    client_id = payload.client_id.strip()
    scopes = " ".join(payload.scopes.split())
    if payload.enabled and not client_id:
        raise HTTPException(status_code=400, detail="Client ID wajib diisi saat SSO aktif")
    if payload.enabled and "openid" not in scopes.split():
        raise HTTPException(status_code=400, detail="Scope openid wajib disertakan")

    existing_secret = oidc_settings().get("client_secret", "")
    client_secret = "" if payload.clear_client_secret else (payload.client_secret.strip() or existing_secret)
    doc = {
        "id": "main",
        "enabled": bool(payload.enabled),
        "discovery_url": discovery_url,
        "issuer": issuer,
        "client_id": client_id,
        "client_secret_encrypted": encrypt_secret(client_secret) if client_secret else "",
        "redirect_uri": redirect_uri,
        "frontend_url": frontend_url,
        "scopes": scopes or "openid profile email roles",
        "local_login_enabled": bool(payload.local_login_enabled),
        "updated_at": now_iso(),
        "updated_by": user["id"],
    }
    await db.oidc_settings.update_one({"id": "main"}, {"$set": doc}, upsert=True)
    return await load_oidc_runtime_settings()


@api_router.post("/sso/settings/test")
async def test_oidc_admin_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    metadata = await oidc_discovery(force=True)
    settings = oidc_require_settings()
    return {
        "ok": True,
        "message": "Koneksi SCI-ID berhasil",
        "issuer": metadata.get("issuer", ""),
        "authorization_endpoint": metadata.get("authorization_endpoint", ""),
        "client_id": settings.get("client_id", ""),
        "client_secret_configured": bool(settings.get("client_secret")),
    }


@api_router.post("/classes/join-request")
async def request_join_class(payload: JoinRequestInput, user: Dict[str, Any] = Depends(get_current_user)):
    if user.get("role") != "student":
        raise HTTPException(status_code=403, detail="Hanya mahasiswa yang dapat meminta masuk kelas")
    class_code = clean_code(payload.class_code)
    class_doc = await db.classes.find_one({"class_code": class_code, "status": "active"}, {"_id": 0})
    if not class_doc:
        raise HTTPException(status_code=404, detail="Kode kelas tidak ditemukan")
    if class_doc["id"] in user.get("class_ids", []):
        return {"status": "approved", "message": "Mahasiswa sudah terdaftar di kelas ini", "class_id": class_doc["id"]}
    invited = await db.enrollment_requests.find_one(
        {"class_id": class_doc["id"], "student_id": user["id"], "status": "invited"}, {"_id": 0}
    )
    if invited:
        await db.users.update_one({"id": user["id"]}, {"$addToSet": {"class_ids": class_doc["id"]}})
        await db.classes.update_one({"id": class_doc["id"]}, {"$addToSet": {"student_ids": user["id"]}})
        await db.enrollment_requests.update_one(
            {"id": invited["id"]},
            {"$set": {"status": "approved", "approved_at": now_iso(), "approved_by": user["id"], "accepted_by_student": True}},
        )
        updated = await db.enrollment_requests.find_one({"id": invited["id"]}, {"_id": 0})
        return public_doc(updated)
    existing = await db.enrollment_requests.find_one(
        {"class_id": class_doc["id"], "student_id": user["id"], "status": "pending"}, {"_id": 0}
    )
    if existing:
        return public_doc(existing)
    request_doc = {
        "id": new_id(),
        "class_id": class_doc["id"],
        "class_name": class_doc["name"],
        "class_code": class_doc.get("class_code", ""),
        "lecturer_id": class_doc.get("lecturer_id", ""),
        "lecturer_name": class_doc.get("lecturer_name", ""),
        "student_id": user["id"],
        "student_name": user["name"],
        "student_nim": user.get("nim", ""),
        "student_email": user.get("email", ""),
        "status": "pending",
        "requested_at": now_iso(),
    }
    await db.enrollment_requests.insert_one(request_doc)
    return public_doc(request_doc)


@api_router.get("/enrollment-requests")
async def list_enrollment_requests(user: Dict[str, Any] = Depends(get_current_user)):
    if user.get("role") == "admin":
        return await db.enrollment_requests.find({}, {"_id": 0}).sort("requested_at", -1).to_list(1000)
    if user.get("role") == "lecturer":
        return await db.enrollment_requests.find(
            {"class_id": {"$in": await lecturer_class_ids(user)}}, {"_id": 0}
        ).sort("requested_at", -1).to_list(1000)
    return await db.enrollment_requests.find({"student_id": user["id"]}, {"_id": 0}).sort("requested_at", -1).to_list(1000)


@api_router.post("/enrollment-requests/{request_id}/approve")
async def approve_enrollment_request(request_id: str, user: Dict[str, Any] = Depends(require_admin)):
    request_doc = await db.enrollment_requests.find_one({"id": request_id}, {"_id": 0})
    if not request_doc:
        raise HTTPException(status_code=404, detail="Request tidak ditemukan")
    await require_class_access(request_doc.get("class_id", ""), user)
    await db.users.update_one({"id": request_doc["student_id"]}, {"$addToSet": {"class_ids": request_doc["class_id"]}})
    await db.classes.update_one({"id": request_doc["class_id"]}, {"$addToSet": {"student_ids": request_doc["student_id"]}})
    await db.enrollment_requests.update_one(
        {"id": request_id},
        {"$set": {"status": "approved", "approved_at": now_iso(), "approved_by": user["id"]}},
    )
    updated = await db.enrollment_requests.find_one({"id": request_id}, {"_id": 0})
    return updated


@api_router.post("/enrollment-requests/{request_id}/reject")
async def reject_enrollment_request(request_id: str, user: Dict[str, Any] = Depends(require_admin)):
    request_doc = await db.enrollment_requests.find_one({"id": request_id}, {"_id": 0})
    if not request_doc:
        raise HTTPException(status_code=404, detail="Request tidak ditemukan")
    await require_class_access(request_doc.get("class_id", ""), user)
    await db.enrollment_requests.update_one(
        {"id": request_id},
        {"$set": {"status": "rejected", "rejected_at": now_iso(), "rejected_by": user["id"]}},
    )
    updated = await db.enrollment_requests.find_one({"id": request_id}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=404, detail="Request tidak ditemukan")
    return updated


@api_router.get("/dashboard")
async def dashboard(user: Dict[str, Any] = Depends(require_admin)):
    class_ids = await lecturer_class_ids(user)
    class_query = {"id": {"$in": class_ids}}
    active_classes = await db.classes.find({**class_query, "status": "active"}, {"_id": 0}).to_list(500)
    active_courses_count = len({item.get("course_id") for item in active_classes if item.get("course_id")})
    active_classes_count = len(active_classes)
    assignments = await db.assignments.find({"class_id": {"$in": class_ids}, "is_active": True}, {"_id": 0}).to_list(500)
    submissions = await db.submissions.find({"class_id": {"$in": class_ids}}, {"_id": 0}).to_list(2000)
    material_ids = [item["id"] for item in await db.materials.find({"class_id": {"$in": class_ids}}, {"_id": 0, "id": 1}).to_list(5000)]
    comments = await db.comments.find({"material_id": {"$in": material_ids}}, {"_id": 0}).sort("created_at", -1).to_list(5)
    student_ids = list({student_id for item in active_classes for student_id in item.get("student_ids", [])})
    students = await db.users.find({"id": {"$in": student_ids}, "role": "student"}, {"_id": 0, "password_hash": 0}).to_list(2000)
    now = datetime.now(timezone.utc)
    soon = now + timedelta(days=3)
    published_assignments = [assignment for assignment in assignments if assignment_is_published(assignment, now)]
    near_deadline = 0
    for assignment in published_assignments:
        try:
            deadline = datetime.fromisoformat(assignment["deadline"].replace("Z", "+00:00"))
            if now <= deadline <= soon:
                near_deadline += 1
        except Exception:
            continue
    submitted_pairs = {(s.get("assignment_id"), s.get("student_id")) for s in submissions}
    missing = 0
    classes_for_missing = active_classes
    for assignment in published_assignments:
        class_doc = next((c for c in classes_for_missing if c["id"] == assignment.get("class_id")), None)
        if class_doc:
            for student_id in class_doc.get("student_ids", []):
                if (assignment["id"], student_id) not in submitted_pairs:
                    missing += 1
    ungraded = len([s for s in submissions if s.get("status") in ["Sudah Submit", "Terlambat", "Direvisi"]])
    avg_grade_values = [s.get("grade", 0) for s in submissions if isinstance(s.get("grade"), (int, float))]
    avg_grade = round(sum(avg_grade_values) / len(avg_grade_values), 1) if avg_grade_values else 0
    progress_map = await calculate_student_progress_many([s["id"] for s in students], class_ids)
    risk_high = sum(1 for p in progress_map.values() if p.get("risk_label") == "Risiko Tinggi")
    return {
        "summary": {
            "active_courses": active_courses_count,
            "active_classes": active_classes_count,
            "active_assignments": len(assignments),
            "near_deadline": near_deadline,
            "missing_submissions": missing,
            "ungraded_submissions": ungraded,
            "latest_comments": len(comments),
            "avg_grade": avg_grade,
            "high_risk_students": risk_high,
            **await storage_status_summary(),
        },
        "latest_comments": comments,
    }


CLEAN_DATA_MODULES = {
    "academic": {
        "label": "Akademik, mata kuliah & kelas",
        "description": "Menghapus prodi, mata kuliah, kelas, materi, tugas, submission, diskusi, enrollment, dan file akademik terkait.",
    },
    "students": {
        "label": "Mahasiswa",
        "description": "Menghapus akun mahasiswa, sesi mahasiswa, enrollment, submission, komentar, chat, dan file upload mahasiswa.",
    },
    "materials": {
        "label": "Materi & diskusi",
        "description": "Menghapus materi, file materi, komentar diskusi, dan lampiran komentar. Tugas yang terhubung akan dilepas dari materi.",
    },
    "assignments": {
        "label": "Tugas & submission",
        "description": "Menghapus tugas, submission, lampiran soal, file submission, dan reminder tugas.",
    },
    "grades": {
        "label": "Nilai",
        "description": "Mengosongkan nilai, feedback, histori nilai pada submission, serta menghapus range predikat nilai.",
    },
    "chat": {
        "label": "Chat",
        "description": "Menghapus semua pesan chat dan foto yang dikirim lewat chat.",
    },
    "notifications": {
        "label": "Notifikasi & reset password",
        "description": "Menghapus reminder in-app, histori WhatsApp, email, OTP, dan request reset password.",
    },
    "all": {
        "label": "Semua data percobaan",
        "description": "Membersihkan semua modul percobaan tanpa menghapus akun dosen, settings aplikasi, dan konfigurasi Drive.",
    },
}


async def delete_stored_files(query: Dict[str, Any]) -> int:
    files = await db.stored_files.find(query, {"_id": 0, "id": 1, "local_path": 1}).to_list(5000)
    for file_doc in files:
        local_path = file_doc.get("local_path")
        if local_path:
            try:
                Path(local_path).unlink(missing_ok=True)
            except Exception as exc:
                logger.warning("Gagal menghapus file lokal %s: %s", local_path, exc)
    if files:
        await db.stored_files.delete_many({"id": {"$in": [item["id"] for item in files]}})
    return len(files)


async def clean_data_module_counts() -> Dict[str, int]:
    student_ids = [
        item["id"]
        for item in await db.users.find({"role": "student"}, {"_id": 0, "id": 1}).to_list(5000)
    ]
    return {
        "academic": (
            await db.programs.count_documents({})
            + await db.courses.count_documents({})
            + await db.classes.count_documents({})
            + await db.enrollment_requests.count_documents({})
        ),
        "students": (
            await db.users.count_documents({"role": "student"})
            + await db.sessions.count_documents({"user_id": {"$in": student_ids}} if student_ids else {"_never": "_never"})
        ),
        "materials": (
            await db.materials.count_documents({})
            + await db.comments.count_documents({})
            + await db.stored_files.count_documents({"record_type": {"$in": ["material_attachment", "comment_attachment"]}})
        ),
        "assignments": (
            await db.assignments.count_documents({})
            + await db.submissions.count_documents({})
            + await db.stored_files.count_documents({"record_type": {"$in": ["assignment_attachment", "submission"]}})
        ),
        "grades": (
            await db.submissions.count_documents({"grade": {"$ne": None}})
            + await db.grade_predicates.count_documents({})
        ),
        "chat": (
            await db.chat_messages.count_documents({})
            + await db.stored_files.count_documents({"record_type": "chat_image"})
        ),
        "notifications": (
            await db.reminder_logs.count_documents({})
            + await db.whatsapp_messages.count_documents({})
            + await db.email_messages.count_documents({})
            + await db.password_reset_requests.count_documents({})
            + await db.password_reset_otps.count_documents({})
        ),
    }


def deleted_count(result: Any) -> int:
    return int(getattr(result, "deleted_count", 0) or 0)


def modified_count(result: Any) -> int:
    return int(getattr(result, "modified_count", 0) or 0)


async def execute_clean_data_module(module: str) -> Dict[str, Any]:
    if module not in CLEAN_DATA_MODULES:
        raise HTTPException(status_code=404, detail="Modul clean data tidak ditemukan")
    affected: Dict[str, int] = {}

    if module == "all":
        for child in ["chat", "materials", "assignments", "students", "academic", "grades", "notifications"]:
            child_result = await execute_clean_data_module(child)
            for key, value in child_result.get("affected", {}).items():
                affected[f"{child}.{key}"] = value
        affected["stored_files_remaining_deleted"] = await delete_stored_files({})
        return {"module": module, "label": CLEAN_DATA_MODULES[module]["label"], "affected": affected}

    if module == "chat":
        affected["chat_messages"] = deleted_count(await db.chat_messages.delete_many({}))
        affected["stored_files"] = await delete_stored_files({"record_type": "chat_image"})

    elif module == "materials":
        affected["materials"] = deleted_count(await db.materials.delete_many({}))
        affected["comments"] = deleted_count(await db.comments.delete_many({}))
        affected["stored_files"] = await delete_stored_files({"record_type": {"$in": ["material_attachment", "comment_attachment"]}})
        affected["assignments_unlinked"] = modified_count(await db.assignments.update_many({}, {"$set": {"material_id": ""}}))

    elif module == "assignments":
        affected["assignments"] = deleted_count(await db.assignments.delete_many({}))
        affected["submissions"] = deleted_count(await db.submissions.delete_many({}))
        affected["reminder_logs"] = deleted_count(await db.reminder_logs.delete_many({"assignment_id": {"$ne": ""}}))
        affected["stored_files"] = await delete_stored_files({"record_type": {"$in": ["assignment_attachment", "submission"]}})

    elif module == "grades":
        affected["submissions_reset"] = modified_count(
            await db.submissions.update_many(
                {},
                {
                    "$set": {
                        "grade": None,
                        "feedback": "",
                        "grade_predicate": "",
                        "grade_history": [],
                        "review_status": "submitted",
                    },
                    "$unset": {"graded_at": "", "graded_by": "", "rubric_scores": ""},
                },
            )
        )
        affected["grade_predicates"] = deleted_count(await db.grade_predicates.delete_many({}))

    elif module == "notifications":
        affected["reminder_logs"] = deleted_count(await db.reminder_logs.delete_many({}))
        affected["whatsapp_messages"] = deleted_count(await db.whatsapp_messages.delete_many({}))
        affected["email_messages"] = deleted_count(await db.email_messages.delete_many({}))
        affected["password_reset_requests"] = deleted_count(await db.password_reset_requests.delete_many({}))
        affected["password_reset_otps"] = deleted_count(await db.password_reset_otps.delete_many({}))

    elif module == "students":
        student_ids = [
            item["id"]
            for item in await db.users.find({"role": "student"}, {"_id": 0, "id": 1}).to_list(5000)
        ]
        student_query = {"$in": student_ids} if student_ids else {"$in": ["__none__"]}
        affected["student_sessions"] = deleted_count(await db.sessions.delete_many({"user_id": student_query}))
        affected["student_users"] = deleted_count(await db.users.delete_many({"role": "student"}))
        affected["class_memberships_reset"] = modified_count(await db.classes.update_many({}, {"$set": {"student_ids": []}}))
        affected["enrollment_requests"] = deleted_count(await db.enrollment_requests.delete_many({}))
        affected["submissions"] = deleted_count(await db.submissions.delete_many({"student_id": student_query}))
        affected["reminder_logs"] = deleted_count(await db.reminder_logs.delete_many({"student_id": student_query}))
        affected["comments"] = deleted_count(await db.comments.delete_many({"author_id": student_query}))
        affected["chat_messages"] = deleted_count(await db.chat_messages.delete_many({"participant_ids": {"$in": student_ids}}))
        affected["stored_files"] = await delete_stored_files({"uploaded_by": student_query})

    elif module == "academic":
        affected["programs"] = deleted_count(await db.programs.delete_many({}))
        affected["courses"] = deleted_count(await db.courses.delete_many({}))
        affected["classes"] = deleted_count(await db.classes.delete_many({}))
        affected["enrollment_requests"] = deleted_count(await db.enrollment_requests.delete_many({}))
        affected["materials"] = deleted_count(await db.materials.delete_many({}))
        affected["comments"] = deleted_count(await db.comments.delete_many({}))
        affected["assignments"] = deleted_count(await db.assignments.delete_many({}))
        affected["submissions"] = deleted_count(await db.submissions.delete_many({}))
        affected["reminder_logs"] = deleted_count(await db.reminder_logs.delete_many({}))
        affected["student_class_links"] = modified_count(await db.users.update_many({"role": "student"}, {"$set": {"class_ids": []}}))
        affected["stored_files"] = await delete_stored_files({"record_type": {"$in": ["assignment_attachment", "submission", "material_attachment", "comment_attachment"]}})

    return {"module": module, "label": CLEAN_DATA_MODULES[module]["label"], "affected": affected}


@api_router.get("/clean-data/summary")
async def clean_data_summary(_: Dict[str, Any] = Depends(require_campus_admin)):
    counts = await clean_data_module_counts()
    return [
        {"key": key, **meta, "count": counts.get(key, sum(counts.values()) if key == "all" else 0)}
        for key, meta in CLEAN_DATA_MODULES.items()
    ]


@api_router.post("/clean-data/{module}")
async def clean_data(module: str, payload: CleanDataInput, _: Dict[str, Any] = Depends(require_campus_admin)):
    if payload.confirmation != "HAPUS":
        raise HTTPException(status_code=400, detail="Konfirmasi clean data tidak valid")
    result = await execute_clean_data_module(module)
    result["cleaned_at"] = now_iso()
    return result


@api_router.get("/storage/status")
async def storage_status(_: Dict[str, Any] = Depends(require_campus_admin)):
    return {
        **await storage_status_summary(),
        "files": {
            "total": await db.stored_files.count_documents({}),
            "google_drive": await db.stored_files.count_documents({"storage_provider": "google_drive"}),
            "server_local": await db.stored_files.count_documents({"storage_provider": "server_local"}),
            "drive_pending": await db.stored_files.count_documents({"drive_sync_status": "pending"}),
            "drive_failed": await db.stored_files.count_documents({"drive_sync_status": "failed"}),
        },
    }


async def drive_sync_overview(limit: int = 50) -> Dict[str, Any]:
    query = {"record_type": {"$in": ["submission", "assignment_attachment"]}}
    summary = {
        "total": await db.stored_files.count_documents(query),
        "pending": await db.stored_files.count_documents({**query, "drive_sync_status": "pending"}),
        "synced": await db.stored_files.count_documents({**query, "drive_sync_status": "synced"}),
        "failed": await db.stored_files.count_documents({**query, "drive_sync_status": "failed"}),
        "not_configured": await db.stored_files.count_documents({**query, "drive_sync_status": "not_configured"}),
    }
    files = await db.stored_files.find(query, {"_id": 0}).sort("uploaded_at", -1).to_list(limit)
    submission_ids = [item.get("submission_id") for item in files if item.get("submission_id")]
    assignment_ids = [item.get("assignment_id") for item in files if item.get("assignment_id")]
    submissions = await db.submissions.find({"id": {"$in": submission_ids}}, {"_id": 0}).to_list(limit) if submission_ids else []
    assignments = await db.assignments.find({"id": {"$in": assignment_ids}}, {"_id": 0}).to_list(limit) if assignment_ids else []
    submissions_by_id = {item["id"]: item for item in submissions}
    assignments_by_id = {item["id"]: item for item in assignments}
    items = []
    for item in files:
        submission = submissions_by_id.get(item.get("submission_id", ""), {})
        assignment = assignments_by_id.get(item.get("assignment_id", ""), {})
        items.append(
            {
                "id": item.get("id", ""),
                "file_id": item.get("file_id", item.get("id", "")),
                "file_name": item.get("file_name", ""),
                "size": item.get("size", 0),
                "record_type": item.get("record_type", ""),
                "upload_status": item.get("upload_status", ""),
                "drive_sync_status": item.get("drive_sync_status", ""),
                "drive_error": item.get("drive_error", ""),
                "drive_file_url": item.get("drive_file_url", ""),
                "uploaded_at": item.get("uploaded_at", ""),
                "updated_at": item.get("updated_at", ""),
                "assignment_id": item.get("assignment_id", ""),
                "assignment_title": assignment.get("title") or submission.get("assignment_title", ""),
                "course_name": assignment.get("course_name", ""),
                "class_name": assignment.get("class_name", ""),
                "submission_id": item.get("submission_id", ""),
                "student_name": submission.get("student_name", ""),
                "student_nim": submission.get("student_nim", ""),
            }
        )
    return {"summary": summary, "items": items}


@api_router.get("/database-backups")
async def list_database_backups(_: Dict[str, Any] = Depends(require_campus_admin)):
    settings = await get_database_backup_settings()
    backups = await db.database_backups.find({}, {"_id": 0, "local_path": 0}).sort("created_at", -1).to_list(100)
    drive_settings = await get_google_drive_settings(mask=True)
    return {
        "settings": settings,
        "backups": backups,
        "drive_ready": google_drive_upload_enabled(drive_settings),
        "drive_account": drive_settings.get("service_account_email", ""),
        "drive_folder": f"{drive_settings.get('root_folder_name', 'E-Learning Dosen')} / Database Backups",
    }


@api_router.put("/database-backups/settings")
async def update_database_backup_settings(
    payload: DatabaseBackupSettingsInput,
    user: Dict[str, Any] = Depends(require_campus_admin),
):
    if payload.frequency not in {"daily", "weekly"}:
        raise HTTPException(status_code=400, detail="Frekuensi backup harus harian atau mingguan")
    if not re.fullmatch(r"(?:[01]\d|2[0-3]):[0-5]\d", payload.run_time):
        raise HTTPException(status_code=400, detail="Waktu backup harus memakai format HH:MM")
    if payload.enabled and payload.upload_to_drive:
        drive_settings = await get_google_drive_settings(mask=True)
        if not google_drive_upload_enabled(drive_settings):
            raise HTTPException(
                status_code=400,
                detail="Aktifkan dan simpan konfigurasi Google Drive sebelum menjadwalkan backup ke Drive",
            )
    doc = {
        "id": "main",
        **payload.model_dump(),
        "timezone": str(BACKUP_TIMEZONE),
        "next_run_at": next_database_backup_at(payload.model_dump()) if payload.enabled else "",
        "updated_at": now_iso(),
        "updated_by": user["id"],
        "running": False,
    }
    await db.database_backup_settings.update_one({"id": "main"}, {"$set": doc}, upsert=True)
    return await get_database_backup_settings()


@api_router.post("/database-backups/run")
async def run_database_backup(user: Dict[str, Any] = Depends(require_campus_admin)):
    running = await db.database_backups.find_one({"status": "running"}, {"_id": 0, "id": 1})
    if running:
        raise HTTPException(status_code=409, detail="Proses backup lain masih berjalan")
    return await create_database_backup("manual", user["id"])


@api_router.get("/database-backups/{backup_id}/download")
async def download_database_backup(
    backup_id: str, _: Dict[str, Any] = Depends(require_campus_admin)
):
    backup = await db.database_backups.find_one({"id": backup_id}, {"_id": 0})
    if not backup or not backup.get("local_available") or not backup.get("local_path"):
        raise HTTPException(status_code=404, detail="File backup lokal tidak tersedia")
    path = Path(backup["local_path"]).resolve()
    if BACKUP_ROOT.resolve() not in path.parents or not path.exists():
        raise HTTPException(status_code=404, detail="File backup lokal tidak ditemukan")
    return FileResponse(path, filename=backup.get("file_name", path.name), media_type="application/gzip")


@api_router.get("/drive/settings")
async def get_drive_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    return {**await get_google_drive_settings(mask=True), **await storage_status_summary(), **await drive_sync_overview()}


@api_router.get("/drive/sync-status")
async def get_drive_sync_status(_: Dict[str, Any] = Depends(require_campus_admin)):
    return await drive_sync_overview()


@api_router.post("/drive/sync/{file_id}/retry")
async def retry_drive_sync_file(file_id: str, background_tasks: BackgroundTasks, _: Dict[str, Any] = Depends(require_campus_admin)):
    file_doc = await db.stored_files.find_one({"id": file_id}, {"_id": 0})
    if not file_doc:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    await db.stored_files.update_one(
        {"id": file_id},
        {"$set": {"drive_sync_status": "pending", "upload_status": "stored_on_server", "updated_at": now_iso()}, "$unset": {"drive_error": ""}},
    )
    await refresh_embedded_file_references(file_id)
    background_tasks.add_task(sync_stored_file_to_drive, file_id)
    return {"ok": True, "file_id": file_id, "status": "pending"}


@api_router.post("/drive/sync/retry-failed")
async def retry_failed_drive_sync(background_tasks: BackgroundTasks, _: Dict[str, Any] = Depends(require_campus_admin)):
    failed = await db.stored_files.find({"drive_sync_status": "failed"}, {"_id": 0, "id": 1}).to_list(100)
    for item in failed:
        file_id = item["id"]
        await db.stored_files.update_one(
            {"id": file_id},
            {"$set": {"drive_sync_status": "pending", "upload_status": "stored_on_server", "updated_at": now_iso()}, "$unset": {"drive_error": ""}},
        )
        await refresh_embedded_file_references(file_id)
        background_tasks.add_task(sync_stored_file_to_drive, file_id)
    return {"ok": True, "queued": len(failed)}


@api_router.put("/drive/settings")
async def update_drive_settings(
    payload: GoogleDriveSettingsInput,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_campus_admin),
):
    if payload.lecturer_folder_role not in {"reader", "writer"}:
        raise HTTPException(status_code=400, detail="Akses folder dosen harus reader atau writer")
    delegated_user = payload.google_workspace_delegated_user.strip().lower()
    if delegated_user and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", delegated_user):
        raise HTTPException(status_code=400, detail="Email Google Workspace untuk Meet tidak valid")
    existing = await db.google_drive_settings.find_one({"id": "main"}, {"_id": 0}) or {}
    service_account_json_encrypted = existing.get("service_account_json_encrypted", "")
    service_account_email = existing.get("service_account_email", "")
    if payload.clear_service_account:
        service_account_json_encrypted = ""
        service_account_email = ""
    if payload.service_account_json.strip():
        try:
            info = normalize_service_account_payload(payload.service_account_json)
        except Exception:
            raise HTTPException(status_code=400, detail="Service account JSON tidak valid")
        normalized = json.dumps(info, separators=(",", ":"), ensure_ascii=False)
        service_account_json_encrypted = encrypt_secret(normalized)
        service_account_email = info.get("client_email", "")
    doc = {
        "id": "main",
        "enabled": payload.enabled,
        "root_folder_id": payload.root_folder_id.strip(),
        "root_folder_name": safe_path_segment(payload.root_folder_name or "E-Learning Dosen"),
        "require_upload": payload.require_upload,
        "lecturer_folder_sharing_enabled": payload.lecturer_folder_sharing_enabled,
        "lecturer_folder_role": payload.lecturer_folder_role,
        "google_meet_enabled": payload.google_meet_enabled,
        "google_workspace_delegated_user": delegated_user,
        "service_account_json_encrypted": service_account_json_encrypted,
        "service_account_email": service_account_email,
        "updated_at": now_iso(),
        "updated_by": user["id"],
    }
    await db.google_drive_settings.update_one({"id": "main"}, {"$set": doc}, upsert=True)
    _invalidate_settings_cache("google_drive_settings")
    background_tasks.add_task(reconcile_all_lecturer_drive_access)
    return {**await get_google_drive_settings(mask=True), **await storage_status_summary(), **await drive_sync_overview()}


def test_drive_connection_sync(settings: Dict[str, Any]) -> Dict[str, Any]:
    service = get_drive_service(settings)
    root_folder_id = settings.get("root_folder_id", "")
    folder_name = ""
    if root_folder_id:
        folder = (
            service.files()
            .get(fileId=root_folder_id, fields="id,name,mimeType", supportsAllDrives=True)
            .execute()
        )
        parent_id = folder.get("id", "")
        folder_name = folder.get("name", "")
    else:
        root_folder_name = safe_path_segment(settings.get("root_folder_name") or "E-Learning Dosen")
        parent_id = drive_find_or_create_folder(service, root_folder_name, None)
        folder_name = root_folder_name
    test_name = f".nugas-upload-test-{uuid.uuid4().hex[:8]}.txt"
    media = MediaIoBaseUpload(io.BytesIO(b"nugas google drive upload test"), mimetype="text/plain", resumable=True)
    uploaded = (
        service.files()
        .create(
            body={"name": test_name, "parents": [parent_id]},
            media_body=media,
            fields="id,name",
            supportsAllDrives=True,
        )
        .execute()
    )
    uploaded_id = uploaded.get("id", "")
    if uploaded_id:
        service.files().delete(fileId=uploaded_id, supportsAllDrives=True).execute()
    return {"ok": True, "folder_id": parent_id, "folder_name": folder_name, "upload_test": "ok"}


@api_router.post("/drive/settings/test")
async def test_drive_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    settings = await get_google_drive_settings(mask=False)
    if not google_drive_upload_enabled(settings):
        raise HTTPException(status_code=400, detail="Google Drive belum aktif atau credential belum tersimpan")
    try:
        return await asyncio.to_thread(test_drive_connection_sync, settings)
    except Exception as exc:
        logger.exception("Tes koneksi Google Drive gagal: %s", exc)
        raise HTTPException(status_code=400, detail=google_drive_error_message(exc))


@api_router.post("/drive/settings/test-meet")
async def test_google_meet_settings(_: Dict[str, Any] = Depends(require_campus_admin)):
    settings = await get_google_drive_settings(mask=False)
    if not settings.get("google_meet_enabled"):
        raise HTTPException(status_code=400, detail="Google Meet REST API belum diaktifkan pada konfigurasi admin")
    if not settings.get("service_account_configured"):
        raise HTTPException(status_code=400, detail="Service Account JSON belum tersimpan")
    delegated_user = settings.get("google_workspace_delegated_user", "")
    if not delegated_user:
        raise HTTPException(status_code=400, detail="Email akun Google Workspace untuk pengujian belum diisi")
    try:
        meet = await asyncio.to_thread(create_google_meet_space_sync, settings, delegated_user)
    except Exception as exc:
        logger.warning("Tes koneksi Google Meet gagal: %s", exc)
        raise HTTPException(status_code=400, detail=str(exc)[:500]) from exc
    return {
        "ok": True,
        "message": "Koneksi Google Meet berhasil dan ruang uji telah dibuat",
        "delegated_user": delegated_user,
        **meet,
    }


@api_router.get("/programs")
async def list_programs(_: Dict[str, Any] = Depends(get_current_user)):
    return await db.programs.find({"status": {"$ne": "deleted"}}, {"_id": 0}).sort("created_at", -1).to_list(500)


@api_router.post("/programs")
async def create_program(payload: ProgramInput, _: Dict[str, Any] = Depends(require_admin)):
    doc = payload.model_dump()
    doc.update({"id": new_id(), "status": "active", "created_at": now_iso()})
    await db.programs.insert_one(doc)
    return public_doc(doc)


@api_router.put("/programs/{program_id}")
async def update_program(program_id: str, payload: ProgramInput, _: Dict[str, Any] = Depends(require_admin)):
    existing = await db.programs.find_one({"id": program_id, "status": {"$ne": "deleted"}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Prodi tidak ditemukan")
    update = payload.model_dump()
    update["updated_at"] = now_iso()
    await db.programs.update_one({"id": program_id}, {"$set": update})
    await db.courses.update_many({"program_id": program_id}, {"$set": {"program_name": payload.name}})
    await db.classes.update_many({"program_id": program_id}, {"$set": {"program_name": payload.name}})
    doc = await db.programs.find_one({"id": program_id}, {"_id": 0})
    return public_doc(doc)


@api_router.delete("/programs/{program_id}")
async def delete_program(program_id: str, _: Dict[str, Any] = Depends(require_admin)):
    linked_course = await db.courses.find_one({"program_id": program_id, "status": {"$ne": "deleted"}}, {"_id": 0, "id": 1})
    if linked_course:
        raise HTTPException(status_code=400, detail="Prodi masih dipakai mata kuliah. Hapus atau pindahkan mata kuliah dulu.")
    result = await db.programs.update_one({"id": program_id}, {"$set": {"status": "deleted", "deleted_at": now_iso()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Prodi tidak ditemukan")
    return {"ok": True}


@api_router.get("/courses")
async def list_courses(_: Dict[str, Any] = Depends(get_current_user)):
    courses = await db.courses.find({"status": {"$ne": "deleted"}}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [await enrich_course_payload(item) for item in courses]


@api_router.post("/courses")
async def create_course(payload: CourseInput, _: Dict[str, Any] = Depends(require_admin)):
    program_query = {"id": payload.program_id, "status": {"$ne": "deleted"}} if payload.program_id else {"status": {"$ne": "deleted"}}
    program = await db.programs.find_one(program_query, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Prodi tidak ditemukan")
    doc = payload.model_dump()
    doc["program_id"] = program["id"]
    doc.update({"id": new_id(), "program_name": program["name"], "status": "active", "created_at": now_iso()})
    await db.courses.insert_one(doc)
    return public_doc(doc)


@api_router.put("/courses/{course_id}")
async def update_course(course_id: str, payload: CourseInput, _: Dict[str, Any] = Depends(require_admin)):
    existing = await db.courses.find_one({"id": course_id, "status": {"$ne": "deleted"}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Mata kuliah tidak ditemukan")
    selected_program_id = payload.program_id or existing.get("program_id", "")
    program = await db.programs.find_one({"id": selected_program_id, "status": {"$ne": "deleted"}}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Prodi tidak ditemukan")
    update = payload.model_dump()
    update["program_id"] = program["id"]
    update.update({"program_name": program["name"], "updated_at": now_iso()})
    await db.courses.update_one({"id": course_id}, {"$set": update})
    await db.classes.update_many(
        {"course_id": course_id},
        {"$set": {"course_name": payload.name, "program_id": payload.program_id, "program_name": program["name"]}},
    )
    doc = await db.courses.find_one({"id": course_id}, {"_id": 0})
    return await enrich_course_payload(public_doc(doc))


@api_router.delete("/courses/{course_id}")
async def delete_course(course_id: str, _: Dict[str, Any] = Depends(require_admin)):
    linked_class = await db.classes.find_one({"course_id": course_id, "status": {"$ne": "deleted"}}, {"_id": 0, "id": 1})
    if linked_class:
        raise HTTPException(status_code=400, detail="Mata kuliah masih dipakai kelas. Hapus kelas dulu.")
    result = await db.courses.update_one({"id": course_id}, {"$set": {"status": "deleted", "deleted_at": now_iso()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Mata kuliah tidak ditemukan")
    return {"ok": True}


@api_router.get("/classes")
async def list_classes(user: Dict[str, Any] = Depends(get_current_user)):
    query: Dict[str, Any] = {"status": {"$ne": "deleted"}}
    if user["role"] == "student":
        query = {"id": {"$in": user.get("class_ids", [])}, "status": {"$ne": "deleted"}}
    elif user["role"] == "lecturer":
        query["lecturer_id"] = user["id"]
    classes = await db.classes.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [await enrich_class_payload(item) for item in classes]


@api_router.post("/classes")
async def create_class(payload: ClassInput, user: Dict[str, Any] = Depends(require_admin)):
    course = await db.courses.find_one({"id": payload.course_id, "status": {"$ne": "deleted"}}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Mata kuliah tidak ditemukan")
    code_seed = f"{course.get('code', 'KLS')}{payload.name}{uuid.uuid4().hex[:4]}"
    doc = payload.model_dump()
    doc.update(
        {
            "id": new_id(),
            "course_name": course["name"],
            "program_id": course.get("program_id", ""),
            "program_name": course.get("program_name", ""),
            "class_code": clean_code(code_seed),
            "lecturer_id": user["id"],
            "lecturer_name": user.get("name", "Dosen"),
            "status": "active",
            "student_ids": [],
            "created_at": now_iso(),
        }
    )
    await db.classes.insert_one(doc)
    return await enrich_class_payload(public_doc(doc))


@api_router.put("/classes/{class_id}")
async def update_class(class_id: str, payload: ClassInput, user: Dict[str, Any] = Depends(require_admin)):
    existing = await require_class_access(class_id, user)
    course = await db.courses.find_one({"id": payload.course_id, "status": {"$ne": "deleted"}}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Mata kuliah tidak ditemukan")
    update = payload.model_dump()
    update.update(
        {
            "course_name": course["name"],
            "program_id": course.get("program_id", ""),
            "program_name": course.get("program_name", ""),
            "updated_at": now_iso(),
        }
    )
    await db.classes.update_one({"id": class_id}, {"$set": update})
    doc = await db.classes.find_one({"id": class_id}, {"_id": 0})
    return await enrich_class_payload(public_doc(doc))


@api_router.delete("/classes/{class_id}")
async def delete_class(class_id: str, user: Dict[str, Any] = Depends(require_admin)):
    await require_class_access(class_id, user)
    result = await db.classes.update_one({"id": class_id}, {"$set": {"status": "deleted", "deleted_at": now_iso()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan")
    await db.users.update_many({"class_ids": class_id}, {"$pull": {"class_ids": class_id}})
    return {"ok": True}


@api_router.post("/classes/{class_id}/archive")
async def archive_class(class_id: str, user: Dict[str, Any] = Depends(require_admin)):
    await require_class_access(class_id, user)
    await db.classes.update_one({"id": class_id}, {"$set": {"status": "archived", "archived_at": now_iso()}})
    return {"ok": True}


@api_router.post("/classes/{class_id}/end")
async def end_class(class_id: str, user: Dict[str, Any] = Depends(require_admin)):
    await require_class_access(class_id, user)
    await db.classes.update_one({"id": class_id}, {"$set": {"status": "ended", "ended_at": now_iso()}})
    return {"ok": True}


@api_router.get("/classes/{class_id}/students")
async def class_students(class_id: str, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_access(class_id, user)
    students = await db.users.find(
        {"role": "student", "class_ids": class_id}, {"_id": 0, "password_hash": 0}
    ).sort("name", 1).to_list(1000)
    progress_map = await calculate_student_progress_many([s["id"] for s in students], [class_id])
    for student in students:
        student["progress"] = progress_map.get(student["id"], {})
    return {"class": await enrich_class_payload(class_doc), "students": students}


@api_router.post("/classes/{class_id}/students/{student_id}/remove")
async def remove_student_from_class(class_id: str, student_id: str, user: Dict[str, Any] = Depends(require_admin)):
    await require_class_access(class_id, user)
    await db.classes.update_one({"id": class_id}, {"$pull": {"student_ids": student_id}})
    await db.users.update_one({"id": student_id}, {"$pull": {"class_ids": class_id}})
    return {"ok": True}


@api_router.post("/classes/{class_id}/students/{student_id}/add")
async def add_existing_student_to_class(class_id: str, student_id: str, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await get_manageable_class(class_id, user=user)
    student = await get_active_student(student_id)
    result = await add_student_to_class_record(class_doc, student, user["id"])
    return {"ok": True, **result}


@api_router.post("/classes/{class_id}/students/bulk-add")
async def bulk_add_existing_students_to_class(class_id: str, payload: StudentIdsInput, user: Dict[str, Any] = Depends(require_admin)):
    student_ids = unique_ids(payload.student_ids)
    if not student_ids:
        raise HTTPException(status_code=400, detail="Pilih minimal satu mahasiswa")
    class_doc = await get_manageable_class(class_id, user=user)
    results = []
    for student_id in student_ids:
        try:
            student = await get_active_student(student_id)
            results.append(await add_student_to_class_record(class_doc, student, user["id"]))
        except HTTPException as exc:
            results.append({"student_id": student_id, "status": "skipped", "detail": exc.detail})
    added = len([item for item in results if item.get("status") == "approved"])
    already_joined = len([item for item in results if item.get("status") == "already_joined"])
    skipped = len([item for item in results if item.get("status") == "skipped"])
    return {"ok": True, "added": added, "already_joined": already_joined, "skipped": skipped, "results": results}


@api_router.post("/classes/{class_id}/students/{student_id}/invite")
async def invite_student_to_class(
    class_id: str,
    student_id: str,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    class_doc = await get_manageable_class(class_id, active_only=True, user=user)
    student = await get_active_student(student_id)
    result = await invite_student_to_class_record(class_doc, student, background_tasks, user["id"])
    return {"ok": True, **result}


@api_router.post("/classes/{class_id}/students/bulk-invite")
async def bulk_invite_students_to_class(
    class_id: str,
    payload: StudentIdsInput,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    student_ids = unique_ids(payload.student_ids)
    if not student_ids:
        raise HTTPException(status_code=400, detail="Pilih minimal satu mahasiswa")
    class_doc = await get_manageable_class(class_id, active_only=True, user=user)
    results = []
    for student_id in student_ids:
        try:
            student = await get_active_student(student_id)
            results.append(await invite_student_to_class_record(class_doc, student, background_tasks, user["id"]))
        except HTTPException as exc:
            results.append({"student_id": student_id, "status": "skipped", "detail": exc.detail, "delivery_status": ""})
    invited = len([item for item in results if item.get("status") in ["invited", "pending"]])
    already_joined = len([item for item in results if item.get("status") == "already_joined"])
    skipped = len([item for item in results if item.get("status") == "skipped"])
    queued = len([item for item in results if item.get("delivery_status") == "pending"])
    pending_config = len([item for item in results if item.get("delivery_status") == "pending_config"])
    no_whatsapp = len([item for item in results if item.get("delivery_status") == "no_whatsapp"])
    return {
        "ok": True,
        "invited": invited,
        "already_joined": already_joined,
        "skipped": skipped,
        "queued": queued,
        "pending_config": pending_config,
        "no_whatsapp": no_whatsapp,
        "results": results,
    }


@api_router.post("/students/{student_id}/status")
async def update_student_status(
    student_id: str,
    payload: Dict[str, str],
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    status = payload.get("status")
    if status not in ["active", "inactive"]:
        raise HTTPException(status_code=400, detail="Status harus active atau inactive")
    await db.users.update_one({"id": student_id, "role": "student"}, {"$set": {"status": status, "status_updated_at": now_iso()}})
    return {"ok": True, "status": status}


@api_router.post("/students/{student_id}/reset-password")
async def reset_student_password(
    student_id: str,
    payload: ResetPasswordInput,
    _: Dict[str, Any] = Depends(require_campus_admin),
):
    student = await db.users.find_one({"id": student_id, "role": "student"}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Mahasiswa tidak ditemukan")
    new_password = payload.password.strip() if payload.password.strip() else (student.get("nim") or "Mahasiswa123!")
    if len(new_password) < 3:
        raise HTTPException(status_code=400, detail="Password minimal 3 karakter")
    await db.users.update_one(
        {"id": student_id},
        {"$set": {"password_hash": hash_password(new_password), "password_reset_at": now_iso()}},
    )
    return {"ok": True, "temporary_password": new_password}


@api_router.get("/students")
async def list_students(user: Dict[str, Any] = Depends(require_admin)):
    query: Dict[str, Any] = {"role": "student"}
    progress_class_ids: Optional[List[str]] = None
    if not is_campus_admin(user):
        class_ids = await lecturer_class_ids(user)
        progress_class_ids = class_ids
        # Dosen dapat memilih mahasiswa aktif dari seluruh katalog kampus.
        # Mahasiswa nonaktif tetap dikembalikan hanya bila sudah ada di kelas
        # yang dikelola dosen agar keanggotaannya tetap terlihat.
        query["$or"] = [
            {"status": "active"},
            {"status": {"$exists": False}},
            {"class_ids": {"$in": class_ids}},
        ]
    students = await db.users.find(query, {"_id": 0, "password_hash": 0}).sort("name", 1).to_list(2000)
    progress_map = await calculate_student_progress_many(
        [s["id"] for s in students], progress_class_ids
    )
    for student in students:
        student["progress"] = progress_map.get(student["id"], {})
    return students


@api_router.post("/students")
async def create_student(payload: StudentInput, user: Dict[str, Any] = Depends(require_campus_admin)):
    class_doc = await require_class_access(payload.class_id, user)
    existing = await db.users.find_one({"email": payload.email.lower()}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=409, detail="Email mahasiswa sudah digunakan")
    student_id = new_id()
    username = payload.nim.strip().lower()
    if await db.users.find_one({"username": username}, {"_id": 0, "id": 1}):
        username = f"{username}-{student_id[:6]}"
    doc = {
        "id": student_id,
        "role": "student",
        "username": username,
        "nim": payload.nim,
        "name": payload.name,
        "email": payload.email.lower(),
        "whatsapp": payload.whatsapp,
        "password_hash": hash_password(payload.password),
        "status": payload.status,
        "class_ids": [payload.class_id],
        "created_at": now_iso(),
        "last_login_at": "",
    }
    await db.users.insert_one(doc)
    await db.classes.update_one({"id": payload.class_id}, {"$addToSet": {"student_ids": student_id}})
    return public_doc(doc)


@api_router.post("/classes/{class_id}/students/import")
async def import_students(
    class_id: str,
    file: UploadFile = File(...),
    default_password: str = Form(""),
    user: Dict[str, Any] = Depends(require_campus_admin),
):
    class_doc = await require_class_access(class_id, user)
    content = await file.read()
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active
    headers = [str(cell.value or "").strip().lower() for cell in sheet[1]]
    mapping = {name: idx for idx, name in enumerate(headers)}
    default_import_password = default_password.strip()

    def row_value(row: tuple[Any, ...], keys: List[str], fallback_index: Optional[int] = None) -> str:
        index = next((mapping[key] for key in keys if key in mapping), fallback_index)
        if index is None or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    created = 0
    skipped = 0
    password_from_column = 0
    password_from_default = 0
    password_from_nim = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        nim = row_value(row, ["nim"], 0)
        name = row_value(row, ["nama", "name"], 1)
        email = row_value(row, ["email"], 2).lower()
        whatsapp = row_value(row, ["whatsapp", "wa", "no hp", "nomor hp"], 3)
        if not nim or not name or not email or await db.users.find_one({"email": email}, {"_id": 0}):
            skipped += 1
            continue
        row_password = row_value(row, ["password", "pass", "sandi", "kata sandi", "kata_sandi"])
        password = row_password or default_import_password or nim
        if len(password) < 3:
            skipped += 1
            continue
        if row_password:
            password_from_column += 1
        elif default_import_password:
            password_from_default += 1
        else:
            password_from_nim += 1
        student_id = new_id()
        username = nim.lower()
        if await db.users.find_one({"username": username}, {"_id": 0, "id": 1}):
            username = f"{username}-{student_id[:6]}"
        await db.users.insert_one(
            {
                "id": student_id,
                "role": "student",
                "username": username,
                "nim": nim,
                "name": name,
                "email": email,
                "whatsapp": whatsapp,
                "password_hash": hash_password(password),
                "status": "active",
                "class_ids": [class_id],
                "created_at": now_iso(),
                "last_login_at": "",
            }
        )
        await db.classes.update_one({"id": class_id}, {"$addToSet": {"student_ids": student_id}})
        created += 1
    return {
        "created": created,
        "skipped": skipped,
        "password_from_column": password_from_column,
        "password_from_default": password_from_default,
        "password_from_nim": password_from_nim,
    }


@api_router.get("/materials")
async def list_materials(user: Dict[str, Any] = Depends(get_current_user)):
    query: Dict[str, Any] = {}
    if user["role"] == "student":
        query = {"class_id": {"$in": user.get("class_ids", [])}, "is_active": True}
    elif user["role"] == "lecturer":
        query = {"class_id": {"$in": await lecturer_class_ids(user)}}
    materials = await db.materials.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    class_ids = None if user["role"] != "student" else [material.get("class_id", "") for material in materials]
    meeting_labels = await material_meeting_label_map(class_ids)
    for material in materials:
        material["meeting"] = meeting_labels.get(material.get("id", ""), material.get("meeting", ""))
    return await enrich_materials_batch(materials)


@api_router.post("/materials")
async def create_material(payload: MaterialInput, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_access(payload.class_id, user)
    async with _material_creation_lock:
        meeting_number = await db.materials.count_documents({"class_id": payload.class_id}) + 1
        doc = normalized_material_payload(payload)
        doc.update(
            {
                "id": new_id(),
                "meeting": f"Pertemuan {meeting_number}",
                "created_at": now_iso(),
                "created_by": user["id"],
                "lecturer_id": class_doc.get("lecturer_id", user["id"]),
                "lecturer_name": class_doc.get("lecturer_name", user.get("name", "Dosen")),
            }
        )
        await db.materials.insert_one(doc)
    return public_doc(await enrich_material_payload(doc))


@api_router.post("/materials/google-meet")
async def generate_material_google_meet(payload: GoogleMeetInput, user: Dict[str, Any] = Depends(require_admin)):
    await require_class_access(payload.class_id, user)
    settings = await get_google_drive_settings(mask=False)
    if not settings.get("google_meet_enabled"):
        raise HTTPException(status_code=503, detail="Google Meet belum diaktifkan oleh Admin Kampus")
    try:
        meet = await asyncio.to_thread(create_google_meet_space_sync, settings, user.get("email", ""))
    except Exception as exc:
        logger.warning("Pembuatan Google Meet gagal: %s", exc)
        raise HTTPException(status_code=503, detail=str(exc)[:500]) from exc
    return {"ok": True, **meet}


@api_router.put("/materials/{material_id}")
async def update_material(material_id: str, payload: MaterialInput, user: Dict[str, Any] = Depends(require_admin)):
    existing = await db.materials.find_one({"id": material_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Pertemuan tidak ditemukan")
    await require_class_access(existing.get("class_id", ""), user)
    class_doc = await require_class_access(payload.class_id, user)
    existing_attachment = existing.get("attachment") if isinstance(existing.get("attachment"), dict) else {}
    attachment_id = existing_attachment.get("file_id") or existing_attachment.get("id")
    attachment_url = local_file_urls(attachment_id)["file_url"] if attachment_id else ""
    remove_attachment = bool(attachment_id and payload.file_url != attachment_url)
    update = normalized_material_payload(payload)
    update.pop("meeting", None)
    update["updated_at"] = now_iso()
    operation: Dict[str, Any] = {"$set": update}
    if remove_attachment:
        operation["$unset"] = {"attachment": ""}
    await db.materials.update_one({"id": material_id}, operation)
    if remove_attachment:
        await delete_stored_files({"id": attachment_id, "record_type": "material_attachment"})
    elif attachment_id:
        await db.stored_files.update_one(
            {"id": attachment_id, "record_type": "material_attachment"},
            {"$set": {"material_class_id": payload.class_id}},
        )
    updated = await db.materials.find_one({"id": material_id}, {"_id": 0})
    meeting_labels = await material_meeting_label_map([updated.get("class_id", "")])
    updated["meeting"] = meeting_labels.get(material_id, updated.get("meeting", ""))
    return public_doc(await enrich_material_payload(updated))


@api_router.post("/materials/{material_id}/attachment")
async def upload_material_attachment(
    material_id: str,
    background_tasks: BackgroundTasks,
    attachment: UploadFile = File(...),
    user: Dict[str, Any] = Depends(require_admin),
):
    material = await db.materials.find_one({"id": material_id}, {"_id": 0})
    if not material:
        raise HTTPException(status_code=404, detail="Pertemuan tidak ditemukan")
    await require_class_access(material.get("class_id", ""), user)
    ext = attachment.filename.rsplit(".", 1)[-1].lower() if attachment.filename and "." in attachment.filename else ""
    allowed = {"pdf", "doc", "docx", "ppt", "pptx", "xls", "xlsx", "csv", "txt", "zip", "png", "jpg", "jpeg", "webp"}
    if ext not in allowed:
        raise HTTPException(status_code=400, detail="File materi hanya dokumen, gambar, atau ZIP")
    content = await attachment.read(25 * 1024 * 1024 + 1)
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Ukuran file materi maksimal 25 MB")
    await attachment.seek(0)
    class_doc = await db.classes.find_one({"id": material.get("class_id", "")}, {"_id": 0}) or {}
    class_doc = await enrich_class_payload(class_doc) if class_doc else {}
    hierarchy = [
        class_doc.get("academic_year", "Tahun Akademik"),
        class_doc.get("semester", "Semester"),
        class_doc.get("course_name", "Mata Kuliah"),
        class_doc.get("name", "Kelas"),
        material.get("title", "Materi"),
        "Lampiran Materi",
    ]
    file_doc = await save_uploaded_file_record(
        attachment,
        hierarchy,
        "DOSEN",
        user.get("name", "Dosen"),
        user["id"],
        record_type="material_attachment",
        background_tasks=background_tasks,
        async_drive=True,
        lecturer_id=class_doc.get("lecturer_id", ""),
    )
    await db.stored_files.update_one(
        {"id": file_doc["id"]},
        {"$set": {"material_id": material_id, "material_class_id": material.get("class_id", "")}},
    )
    old_attachment = material.get("attachment") if isinstance(material.get("attachment"), dict) else {}
    old_file_id = old_attachment.get("file_id") or old_attachment.get("id")
    await db.materials.update_one(
        {"id": material_id},
        {"$set": {"attachment": file_doc, "file_url": file_doc["file_url"], "updated_at": now_iso()}},
    )
    if old_file_id and old_file_id != file_doc["id"]:
        await delete_stored_files({"id": old_file_id, "record_type": "material_attachment"})
    updated = await db.materials.find_one({"id": material_id}, {"_id": 0})
    meeting_labels = await material_meeting_label_map([updated.get("class_id", "")])
    updated["meeting"] = meeting_labels.get(material_id, updated.get("meeting", ""))
    return public_doc(await enrich_material_payload(updated))


@api_router.delete("/materials/{material_id}")
async def delete_material(material_id: str, user: Dict[str, Any] = Depends(require_admin)):
    existing = await db.materials.find_one({"id": material_id}, {"_id": 0, "id": 1, "attachment": 1})
    if not existing:
        raise HTTPException(status_code=404, detail="Pertemuan tidak ditemukan")
    full_material = await db.materials.find_one({"id": material_id}, {"_id": 0, "class_id": 1}) or {}
    await require_class_access(full_material.get("class_id", ""), user)
    comments = await db.comments.find({"material_id": material_id}, {"_id": 0, "attachment.id": 1, "attachment.file_id": 1}).to_list(5000)
    attachment_ids = []
    for comment in comments:
        attachment = comment.get("attachment") or {}
        file_id = attachment.get("id") or attachment.get("file_id")
        if file_id:
            attachment_ids.append(file_id)
    stored_file_query: Dict[str, Any] = {"record_type": "comment_attachment", "discussion_material_id": material_id}
    if attachment_ids:
        stored_file_query = {
            "record_type": "comment_attachment",
            "$or": [{"discussion_material_id": material_id}, {"id": {"$in": attachment_ids}}],
        }
    deleted_files = await delete_stored_files(stored_file_query)
    material_attachment = existing.get("attachment") if isinstance(existing.get("attachment"), dict) else {}
    material_attachment_id = material_attachment.get("file_id") or material_attachment.get("id")
    material_file_query: Dict[str, Any] = {"record_type": "material_attachment", "material_id": material_id}
    if material_attachment_id:
        material_file_query = {
            "record_type": "material_attachment",
            "$or": [{"material_id": material_id}, {"id": material_attachment_id}],
        }
    deleted_files += await delete_stored_files(material_file_query)
    deleted_comments = deleted_count(await db.comments.delete_many({"material_id": material_id}))
    unlinked_assignments = modified_count(
        await db.assignments.update_many(
            {"material_id": material_id},
            {"$set": {"material_id": "", "material_link_removed_at": now_iso()}},
        )
    )
    await db.materials.delete_one({"id": material_id})
    return {
        "ok": True,
        "comments_deleted": deleted_comments,
        "attachments_deleted": deleted_files,
        "assignments_unlinked": unlinked_assignments,
    }


async def discussion_material_for_user(material_id: str, user: Dict[str, Any]) -> Dict[str, Any]:
    material = await db.materials.find_one({"id": material_id}, {"_id": 0})
    if not material:
        raise HTTPException(status_code=404, detail="Materi tidak ditemukan")
    if user.get("role") == "student":
        if material.get("class_id") not in user.get("class_ids", []) or not material.get("is_active", True):
            raise HTTPException(status_code=403, detail="Diskusi hanya dapat diakses anggota kelas")
    elif user.get("role") == "lecturer":
        await require_class_access(material.get("class_id", ""), user)
    return material


async def validate_comment_parent(parent_id: str, material_id: str) -> None:
    if not parent_id:
        return
    parent = await db.comments.find_one({"id": parent_id, "material_id": material_id}, {"_id": 0, "id": 1})
    if not parent:
        raise HTTPException(status_code=400, detail="Komentar balasan tidak valid")


@api_router.get("/materials/{material_id}/comments")
async def list_comments(material_id: str, user: Dict[str, Any] = Depends(get_current_user)):
    await discussion_material_for_user(material_id, user)
    comments = await db.comments.find({"material_id": material_id}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    admin_author_ids = list(set(
        c.get("author_id", "") for c in comments
        if c.get("author_role") in {"admin", "lecturer"} and c.get("author_id")
    ))
    admin_users_map: Dict[str, Dict[str, Any]] = {}
    if admin_author_ids:
        admin_docs = await db.users.find({"id": {"$in": admin_author_ids}}, {"_id": 0}).to_list(100)
        admin_users_map = {u["id"]: u for u in admin_docs}
    displayed_admin_names: Dict[str, str] = {}
    for comment in comments:
        if comment.get("author_role") not in {"admin", "lecturer"}:
            continue
        author_id = comment.get("author_id", "")
        if author_id not in displayed_admin_names:
            author = admin_users_map.get(author_id) or {
                "id": author_id,
                "role": "admin",
                "name": comment.get("author_name", ""),
            }
            displayed_admin_names[author_id] = (await chat_contact_payload(author))["name"]
        comment["author_name"] = displayed_admin_names[author_id]
    return comments


@api_router.post("/comments")
async def create_comment(payload: CommentInput, user: Dict[str, Any] = Depends(get_current_user)):
    await discussion_material_for_user(payload.material_id, user)
    await validate_comment_parent(payload.parent_id, payload.material_id)
    if not payload.content.strip():
        raise HTTPException(status_code=400, detail="Isi komentar diperlukan")
    author = await chat_contact_payload(user)
    doc = payload.model_dump()
    doc["content"] = doc["content"].strip()
    doc.update(
        {
            "id": new_id(),
            "author_id": user["id"],
            "author_name": author["name"],
            "author_role": user["role"],
            "is_pinned": False,
            "created_at": now_iso(),
        }
    )
    await db.comments.insert_one(doc)
    return public_doc(doc)


@api_router.post("/materials/{material_id}/comments")
async def create_material_comment_with_image(
    material_id: str,
    content: str = Form(""),
    parent_id: str = Form(""),
    attachment: Optional[UploadFile] = File(None),
    image: Optional[UploadFile] = File(None),
    user: Dict[str, Any] = Depends(get_current_user),
):
    material = await discussion_material_for_user(material_id, user)
    await validate_comment_parent(parent_id, material_id)
    upload = attachment or image
    text = content.strip()
    if not text and not upload:
        raise HTTPException(status_code=400, detail="Isi komentar atau lampiran diperlukan")
    author = await chat_contact_payload(user)
    attachment_doc = None
    if upload:
        ext = upload.filename.rsplit(".", 1)[-1].lower() if upload.filename and "." in upload.filename else ""
        allowed = {"jpg", "jpeg", "png", "webp", "gif", "pdf", "doc", "docx", "xls", "xlsx", "txt", "csv"}
        if ext not in allowed:
            raise HTTPException(status_code=400, detail="Lampiran diskusi hanya gambar atau dokumen")
        content_bytes = await upload.read(10 * 1024 * 1024 + 1)
        if len(content_bytes) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Ukuran lampiran diskusi maksimal 10 MB")
        await upload.seek(0)
        hierarchy = ["Diskusi", safe_path_segment(material.get("title", "Materi"))]
        attachment_doc = await save_uploaded_file_record(
            upload,
            hierarchy,
            user.get("nim", user.get("role", "user")),
            user.get("name", "Pengguna"),
            user["id"],
            record_type="comment_attachment",
            sync_drive=False,
            lecturer_id=material.get("lecturer_id", ""),
        )
        await db.stored_files.update_one(
            {"id": attachment_doc["id"]},
            {"$set": {"discussion_class_id": material.get("class_id", ""), "discussion_material_id": material_id}},
        )
    doc = {
        "id": new_id(),
        "material_id": material_id,
        "content": text,
        "parent_id": parent_id,
        "author_id": user["id"],
        "author_name": author["name"],
        "author_role": user["role"],
        "is_pinned": False,
        "attachment": attachment_doc,
        "created_at": now_iso(),
    }
    await db.comments.insert_one(doc)
    return public_doc(doc)


@api_router.post("/comments/{comment_id}/pin")
async def pin_comment(comment_id: str, user: Dict[str, Any] = Depends(require_admin)):
    comment = await db.comments.find_one({"id": comment_id}, {"_id": 0, "material_id": 1}) or {}
    await discussion_material_for_user(comment.get("material_id", ""), user)
    await db.comments.update_one({"id": comment_id}, {"$set": {"is_pinned": True}})
    return {"ok": True}


@api_router.get("/assignments")
async def list_assignments(background_tasks: BackgroundTasks, user: Dict[str, Any] = Depends(get_current_user)):
    await dispatch_due_assignment_notifications(background_tasks)
    query: Dict[str, Any] = {}
    if user["role"] == "student":
        query = {"class_id": {"$in": user.get("class_ids", [])}, "is_active": True}
    elif user["role"] == "lecturer":
        query = {"class_id": {"$in": await lecturer_class_ids(user)}}
    assignments = await db.assignments.find(query, {"_id": 0}).sort("deadline", 1).to_list(1000)
    settings = await get_app_settings_cached()
    creator_ids = [item.get("created_by") for item in assignments if item.get("created_by")]
    creators = await db.users.find({"id": {"$in": creator_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(200) if creator_ids else []
    creator_names = {item["id"]: item.get("name", "") for item in creators}
    if user["role"] == "student":
        assignments = [assignment for assignment in assignments if assignment_is_published(assignment)]
        assignment_ids = [a["id"] for a in assignments]
        my_submissions = await db.submissions.find(
            {"assignment_id": {"$in": assignment_ids}, "student_id": user["id"]}, {"_id": 0}
        ).to_list(1000) if assignment_ids else []
        subs_by_assignment: Dict[str, Dict[str, Any]] = {}
        for s in my_submissions:
            enrich_submission_file_urls(s)
            subs_by_assignment[s["assignment_id"]] = s
        for assignment in assignments:
            assignment["my_submission"] = subs_by_assignment.get(assignment["id"])
    for assignment in assignments:
        assignment["publish_status"] = assignment_publish_status(assignment)
        assignment["max_file_size_mb"] = assignment_max_file_size_mb(assignment)
        assignment["max_submission_size_mb"] = assignment["max_file_size_mb"]
        lecturer_name = str(assignment.get("lecturer_name") or "").strip()
        if not lecturer_name or lecturer_name.lower() == "dosen admin":
            lecturer_name = creator_names.get(assignment.get("created_by", ""), "") or settings.get("lecturer_name", "") or "Dosen"
        if lecturer_name.lower() == "dosen admin":
            lecturer_name = "Dosen"
        assignment["lecturer_name"] = lecturer_name
    return assignments


@api_router.post("/assignments")
async def create_assignment(payload: AssignmentInput, background_tasks: BackgroundTasks, user: Dict[str, Any] = Depends(require_admin)):
    class_doc = await require_class_access(payload.class_id, user)
    doc = payload.model_dump()
    doc["attachment_link"] = str(doc.get("attachment_link") or "").strip()
    doc["max_file_size_mb"] = assignment_max_file_size_mb(doc)
    doc["deadline"] = normalize_optional_datetime(doc.get("deadline", ""), "Deadline")
    if not doc["deadline"]:
        raise HTTPException(status_code=400, detail="Deadline wajib diisi")
    doc["published_at"] = normalize_optional_datetime(doc.get("published_at", ""), "Jadwal tayang")
    publish_at = parse_iso_datetime(doc["published_at"])
    deadline = parse_iso_datetime(doc["deadline"])
    if publish_at and deadline and publish_at > deadline:
        raise HTTPException(status_code=400, detail="Jadwal tayang tidak boleh setelah deadline")
    lecturer_name = str(user.get("name") or class_doc.get("lecturer_name") or "Dosen").strip()
    if lecturer_name.lower() == "dosen admin":
        lecturer_name = str(user.get("name") or "Dosen").strip()
    if lecturer_name.lower() == "dosen admin":
        lecturer_name = "Dosen"
    doc.update(
        {
            "id": new_id(),
            "course_id": class_doc["course_id"],
            "course_name": class_doc.get("course_name", ""),
            "class_name": class_doc["name"],
            "lecturer_id": class_doc.get("lecturer_id", user["id"]),
            "lecturer_name": lecturer_name,
            "created_at": now_iso(),
            "created_by": user["id"],
        }
    )
    await db.assignments.insert_one(doc)
    if assignment_is_published(doc):
        sent_at = now_iso()
        await db.assignments.update_one({"id": doc["id"]}, {"$set": {"published_notification_sent_at": sent_at}})
        doc["published_notification_sent_at"] = sent_at
        await send_assignment_publication_notifications(doc, class_doc, background_tasks)
    response = doc.copy()
    response["publish_status"] = assignment_publish_status(response)
    return public_doc(response)


@api_router.put("/assignments/{assignment_id}")
async def update_assignment(
    assignment_id: str,
    payload: AssignmentInput,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    existing = await db.assignments.find_one({"id": assignment_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Tugas tidak ditemukan")
    await require_class_access(existing.get("class_id", ""), user)
    class_doc = await require_class_access(payload.class_id, user)
    doc = payload.model_dump()
    doc["attachment_link"] = str(doc.get("attachment_link") or "").strip()
    doc["max_file_size_mb"] = assignment_max_file_size_mb(doc)
    doc["deadline"] = normalize_optional_datetime(doc.get("deadline", ""), "Deadline")
    if not doc["deadline"]:
        raise HTTPException(status_code=400, detail="Deadline wajib diisi")
    doc["published_at"] = normalize_optional_datetime(doc.get("published_at", ""), "Jadwal tayang")
    publish_at = parse_iso_datetime(doc["published_at"])
    deadline = parse_iso_datetime(doc["deadline"])
    if publish_at and deadline and publish_at > deadline:
        raise HTTPException(status_code=400, detail="Jadwal tayang tidak boleh setelah deadline")

    if existing.get("deadline"):
        await db.submissions.update_many(
            {"assignment_id": assignment_id, "assignment_deadline": {"$exists": False}},
            {"$set": {"assignment_deadline": existing["deadline"]}},
        )
    await db.submissions.update_many(
        {"assignment_id": assignment_id, "assignment_late_penalty_per_day": {"$exists": False}},
        {"$set": {"assignment_late_penalty_per_day": existing.get("late_penalty_per_day", 0)}},
    )

    doc.update(
        {
            "course_id": class_doc["course_id"],
            "course_name": class_doc.get("course_name", ""),
            "class_name": class_doc["name"],
            "updated_at": now_iso(),
            "updated_by": user["id"],
        }
    )
    await db.assignments.update_one({"id": assignment_id}, {"$set": doc})
    updated = await db.assignments.find_one({"id": assignment_id}, {"_id": 0})
    if (
        updated
        and not assignment_is_published(existing)
        and assignment_is_published(updated)
        and not existing.get("published_notification_sent_at")
    ):
        sent_at = now_iso()
        await db.assignments.update_one({"id": assignment_id}, {"$set": {"published_notification_sent_at": sent_at}})
        updated["published_notification_sent_at"] = sent_at
        await send_assignment_publication_notifications(updated, class_doc, background_tasks)
    response = updated or doc
    response["publish_status"] = assignment_publish_status(response)
    response["max_file_size_mb"] = assignment_max_file_size_mb(response)
    response["max_submission_size_mb"] = response["max_file_size_mb"]
    return public_doc(response)


@api_router.post("/assignments/{assignment_id}/attachments")
async def upload_assignment_attachments(
    assignment_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    assignment = await db.assignments.find_one({"id": assignment_id}, {"_id": 0})
    if not assignment:
        raise HTTPException(status_code=404, detail="Tugas tidak ditemukan")
    await require_class_access(assignment.get("class_id", ""), user)
    _, files = await multipart_uploads(request, ["files", "file", "files[]"])
    if not files:
        raise HTTPException(status_code=400, detail="Minimal satu lampiran soal harus diunggah")
    class_doc = await db.classes.find_one({"id": assignment["class_id"]}, {"_id": 0}) or {}
    allowed = {"pdf", "doc", "docx", "ppt", "pptx", "xls", "xlsx", "csv", "txt", "zip", "png", "jpg", "jpeg", "webp"}
    saved = []
    hierarchy = [
        class_doc.get("academic_year", "Tahun Akademik"),
        class_doc.get("semester", "Semester"),
        assignment.get("course_name", "Mata Kuliah"),
        assignment.get("class_name", "Kelas"),
        assignment.get("title", "Tugas"),
        "Lampiran Soal",
    ]
    for upload in files:
        ext = upload.filename.rsplit(".", 1)[-1].lower() if upload.filename and "." in upload.filename else ""
        if ext not in allowed:
            raise HTTPException(status_code=400, detail="Lampiran soal hanya dokumen, gambar, atau ZIP")
        content = await upload.read(25 * 1024 * 1024 + 1)
        if len(content) > 25 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Ukuran setiap lampiran soal maksimal 25 MB")
        await upload.seek(0)
    for upload in files:
        file_doc = await save_uploaded_file_record(
            upload,
            hierarchy,
            "DOSEN",
            user.get("name", "Dosen"),
            user["id"],
            assignment_id=assignment_id,
            record_type="assignment_attachment",
            background_tasks=background_tasks,
            async_drive=True,
            lecturer_id=assignment.get("lecturer_id", class_doc.get("lecturer_id", "")),
        )
        saved.append(file_doc)
    await db.assignments.update_one({"id": assignment_id}, {"$push": {"attachments": {"$each": saved}}})
    updated = await db.assignments.find_one({"id": assignment_id}, {"_id": 0})
    return {"attachments": saved, "assignment": updated}


@api_router.get("/submissions")
async def list_submissions(user: Dict[str, Any] = Depends(get_current_user)):
    query: Dict[str, Any] = {}
    if user["role"] == "student":
        query = {"student_id": user["id"]}
    elif user["role"] == "lecturer":
        query = {"class_id": {"$in": await lecturer_class_ids(user)}}
    submissions = await db.submissions.find(query, {"_id": 0}).sort("submitted_at", -1).to_list(2000)
    for item in submissions:
        if isinstance(item.get("grade"), (int, float)) and not item.get("grade_predicate"):
            item["grade_predicate"] = await calculate_grade_predicate(item["grade"], item.get("class_id", ""))
        enrich_submission_file_urls(item)
    return submissions


@api_router.get("/chat/contacts")
async def chat_contacts(q: str = "", user: Dict[str, Any] = Depends(get_current_user)):
    query = q.strip().lower()
    users: List[Dict[str, Any]] = []
    if query:
        pattern = re.compile(f"^{re.escape(query)}$", re.IGNORECASE)
        users = await db.users.find(
            {
                "id": {"$ne": user["id"]},
                "status": {"$ne": "deleted"},
                "$or": [{"username": pattern}, {"email": pattern}],
            },
            {"_id": 0, "password_hash": 0},
        ).sort("name", 1).to_list(20)
    else:
        messages = await db.chat_messages.find(
            {"participant_ids": user["id"]}, {"_id": 0, "participant_ids": 1}
        ).sort("created_at", -1).to_list(100)
        recent_ids: List[str] = []
        for message in messages:
            for participant_id in message.get("participant_ids", []):
                if participant_id != user["id"] and participant_id not in recent_ids:
                    recent_ids.append(participant_id)
        if recent_ids:
            user_docs = await db.users.find(
                {"id": {"$in": recent_ids}}, {"_id": 0, "password_hash": 0}
            ).to_list(50)
            by_id = {item["id"]: item for item in user_docs}
            users = [by_id[item_id] for item_id in recent_ids if item_id in by_id]
    return [
        {
            **await chat_contact_payload(item),
            "online": chat_connections.is_online(item["id"]),
            "viewing_chat": chat_connections.is_viewing(item["id"], user["id"]),
        }
        for item in users
    ]


@api_router.get("/chat/lecturers")
async def chat_lecturers(user: Dict[str, Any] = Depends(get_current_user)):
    if user.get("role") != "student":
        return []
    class_docs = await db.classes.find(
        {"id": {"$in": user.get("class_ids", [])}, "status": {"$ne": "deleted"}}, {"_id": 0, "lecturer_id": 1}
    ).to_list(1000)
    lecturer_ids = list({item.get("lecturer_id") for item in class_docs if item.get("lecturer_id")})
    lecturers = await db.users.find(
        {"id": {"$in": lecturer_ids}, "role": {"$in": ["admin", "lecturer"]}, "status": {"$ne": "deleted"}},
        {"_id": 0, "password_hash": 0},
    ).sort("name", 1).to_list(100)
    return [
        {
            **await chat_contact_payload(item),
            "online": chat_connections.is_online(item["id"]),
            "viewing_chat": chat_connections.is_viewing(item["id"], user["id"]),
        }
        for item in lecturers
    ]


@api_router.get("/chat/users/{other_user_id}/messages")
async def list_chat_messages(other_user_id: str, user: Dict[str, Any] = Depends(get_current_user)):
    other = await db.users.find_one({"id": other_user_id}, {"_id": 0, "password_hash": 0})
    if not other or other_user_id == user["id"]:
        raise HTTPException(status_code=404, detail="Pengguna chat tidak ditemukan")
    conversation_id = chat_conversation_id(user["id"], other_user_id)
    messages = await db.chat_messages.find(
        {"conversation_id": conversation_id}, {"_id": 0}
    ).sort("created_at", 1).to_list(500)
    return {
        "contact": {
            **await chat_contact_payload(other),
            "online": chat_connections.is_online(other_user_id),
            "viewing_chat": chat_connections.is_viewing(other_user_id, user["id"]),
        },
        "messages": messages,
    }


@api_router.post("/chat/messages")
async def send_chat_message(
    recipient_id: str = Form(...),
    content: str = Form(""),
    attachment: Optional[UploadFile] = File(None),
    user: Dict[str, Any] = Depends(get_current_user),
):
    recipient = await db.users.find_one({"id": recipient_id}, {"_id": 0, "password_hash": 0})
    if not recipient or recipient_id == user["id"]:
        raise HTTPException(status_code=404, detail="Penerima chat tidak ditemukan")
    text = content.strip()
    if len(text) > 4000:
        raise HTTPException(status_code=400, detail="Pesan maksimal 4000 karakter")
    if not text and not attachment:
        raise HTTPException(status_code=400, detail="Isi pesan atau foto diperlukan")
    participant_ids = sorted([user["id"], recipient_id])
    conversation_id = chat_conversation_id(*participant_ids)
    attachment_doc: Optional[Dict[str, Any]] = None
    if attachment:
        if not (attachment.content_type or "").lower().startswith("image/"):
            raise HTTPException(status_code=400, detail="Lampiran chat hanya berupa foto")
        attachment_content = await attachment.read(5 * 1024 * 1024 + 1)
        if len(attachment_content) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Ukuran foto maksimal 5 MB")
        await attachment.seek(0)
        attachment_doc = await save_uploaded_file_record(
            attachment,
            ["Chat", conversation_id],
            user.get("username") or user.get("nim", "USER"),
            user.get("name", "Pengguna"),
            user["id"],
            record_type="chat_image",
            sync_drive=False,
        )
        await db.stored_files.update_one(
            {"id": attachment_doc["id"]}, {"$set": {"chat_participant_ids": participant_ids}}
        )
    message = {
        "id": new_id(),
        "conversation_id": conversation_id,
        "participant_ids": participant_ids,
        "sender_id": user["id"],
        "recipient_id": recipient_id,
        "content": text,
        "attachment": attachment_doc,
        "created_at": now_iso(),
    }
    await db.chat_messages.insert_one(message)
    response = public_doc(message.copy())
    await chat_connections.send_to_users(set(participant_ids), {"type": "message", "message": response})
    return response


async def stored_file_context(file_id: str, token: str = "", authorization: Optional[str] = Header(None)) -> tuple[Dict[str, Any], Dict[str, Any], Path]:
    auth_token = token.strip()
    if not auth_token and authorization and authorization.startswith("Bearer "):
        auth_token = authorization.replace("Bearer ", "", 1).strip()
    if not auth_token:
        raise HTTPException(status_code=401, detail="Token diperlukan")
    session = await db.sessions.find_one({"token": auth_token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Sesi tidak ditemukan")
    user = await find_user(session["user_id"])
    file_doc = await db.stored_files.find_one({"id": file_id}, {"_id": 0})
    if not file_doc:
        file_doc = await db.drive_files.find_one({"id": file_id}, {"_id": 0})
    if not file_doc:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    if user.get("role") == "student" and file_doc.get("submission_id") and file_doc.get("uploaded_by") != user.get("id"):
        raise HTTPException(status_code=403, detail="Tidak punya akses ke file ini")
    if file_doc.get("record_type") == "chat_image" and user.get("id") not in file_doc.get("chat_participant_ids", []):
        raise HTTPException(status_code=403, detail="Tidak punya akses ke foto chat ini")
    if file_doc.get("record_type") == "comment_attachment" and user.get("role") == "student":
        discussion_class_id = file_doc.get("discussion_class_id", "")
        if not discussion_class_id:
            comment = await db.comments.find_one({"attachment.file_id": file_id}, {"_id": 0, "material_id": 1}) or {}
            material = await db.materials.find_one({"id": comment.get("material_id", "")}, {"_id": 0, "class_id": 1}) or {}
            discussion_class_id = material.get("class_id", "")
        if not discussion_class_id or discussion_class_id not in user.get("class_ids", []):
            raise HTTPException(status_code=403, detail="Lampiran diskusi hanya untuk anggota kelas")
    if file_doc.get("record_type") == "material_attachment" and user.get("role") == "student":
        material_class_id = file_doc.get("material_class_id", "")
        if not material_class_id:
            material = await db.materials.find_one({"attachment.file_id": file_id}, {"_id": 0, "class_id": 1}) or {}
            material_class_id = material.get("class_id", "")
        if not material_class_id or material_class_id not in user.get("class_ids", []):
            raise HTTPException(status_code=403, detail="Lampiran materi hanya untuk anggota kelas")
    if user.get("role") == "lecturer" and file_doc.get("record_type") != "chat_image":
        academic_class_id = file_doc.get("material_class_id") or file_doc.get("discussion_class_id") or file_doc.get("class_id", "")
        if not academic_class_id and file_doc.get("submission_id"):
            submission = await db.submissions.find_one({"id": file_doc["submission_id"]}, {"_id": 0, "class_id": 1}) or {}
            academic_class_id = submission.get("class_id", "")
        if not academic_class_id and file_doc.get("assignment_id"):
            assignment = await db.assignments.find_one({"id": file_doc["assignment_id"]}, {"_id": 0, "class_id": 1}) or {}
            academic_class_id = assignment.get("class_id", "")
        if academic_class_id:
            await require_class_access(academic_class_id, user)
    local_path = file_doc.get("local_path", "")
    if not local_path:
        raise HTTPException(status_code=404, detail="File lokal belum tersedia")
    path = Path(local_path).resolve()
    storage_root = STORAGE_ROOT.resolve()
    if storage_root not in path.parents:
        raise HTTPException(status_code=403, detail="Path file tidak valid")
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="File tidak ada di server")
    return user, file_doc, path


@api_router.get("/files/{file_id}/download")
async def download_stored_file(file_id: str, token: str = "", authorization: Optional[str] = Header(None)):
    _, file_doc, path = await stored_file_context(file_id, token, authorization)
    return FileResponse(path, media_type=file_doc.get("mime_type") or "application/octet-stream", filename=file_doc.get("file_name") or path.name)


@api_router.get("/files/{file_id}/inline")
async def inline_stored_file(file_id: str, token: str = "", authorization: Optional[str] = Header(None)):
    _, file_doc, path = await stored_file_context(file_id, token, authorization)
    return FileResponse(
        path,
        media_type=file_doc.get("mime_type") or "application/octet-stream",
        filename=file_doc.get("file_name") or path.name,
        content_disposition_type="inline",
    )


@api_router.get("/files/{file_id}/preview")
async def preview_stored_file(file_id: str, token: str = "", authorization: Optional[str] = Header(None)):
    _, file_doc, path = await stored_file_context(file_id, token, authorization)
    kind = preview_kind(file_doc, path)
    response: Dict[str, Any] = {
        "id": file_id,
        "file_name": file_doc.get("file_name") or path.name,
        "mime_type": file_doc.get("mime_type") or "application/octet-stream",
        "size": file_doc.get("size", 0),
        "kind": kind,
        **local_file_urls(file_id),
    }
    if kind in {"pdf", "image"}:
        response["render"] = "inline"
        return response
    try:
        if kind == "docx":
            response.update({"render": "html", "html": preview_docx_html(path)})
            return response
        if kind == "xlsx":
            response.update({"render": "html", "html": preview_xlsx_html(path)})
            return response
        if kind == "text":
            response.update({"render": "html", "html": preview_text_html(path)})
            return response
    except Exception as exc:
        logger.warning("Preview dokumen gagal untuk %s: %s", file_id, exc)
        response.update({"render": "unsupported", "message": "Preview dokumen gagal dibaca. File mungkin rusak atau formatnya tidak valid."})
        return response
    response.update(
        {
            "render": "unsupported",
            "message": "Format file ini belum bisa dipreview langsung. Gunakan file PDF, DOCX, XLSX, TXT, CSV, atau gambar.",
        }
    )
    return response


@api_router.post("/assignments/{assignment_id}/submit")
async def submit_assignment(
    assignment_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(get_current_user),
):
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="Pengumpulan hanya untuk mahasiswa")
    assignment = await db.assignments.find_one({"id": assignment_id}, {"_id": 0})
    if not assignment:
        raise HTTPException(status_code=404, detail="Tugas tidak ditemukan")
    if assignment["class_id"] not in user.get("class_ids", []):
        raise HTTPException(status_code=403, detail="Mahasiswa tidak terdaftar pada kelas tugas ini")
    if not assignment_is_published(assignment):
        raise HTTPException(status_code=403, detail="Tugas belum tayang")
    submission = await db.submissions.find_one({"assignment_id": assignment_id, "student_id": user["id"]}, {"_id": 0})
    revision_open = bool(submission and (submission.get("status") == "Direvisi" or submission.get("review_status") == "revision_requested"))
    if submission and not revision_open:
        raise HTTPException(
            status_code=409,
            detail="Tugas sudah dikumpulkan. Mahasiswa hanya bisa mengirim ulang jika dosen mengembalikan tugas sebagai revisi.",
        )
    form, upload_files = await multipart_uploads(request, ["files", "file", "files[]"])
    note = str(form.get("note") or "")
    if not upload_files:
        raise HTTPException(status_code=400, detail="Minimal satu file tugas harus dilampirkan")
    allowed = [item.lower().replace(".", "") for item in assignment.get("allowed_formats", [])]
    allowed = sorted(set(allowed + ["jpg", "jpeg", "png", "webp", "pdf", "doc", "docx", "xls", "xlsx", "zip"]))
    for upload in upload_files:
        extension = upload.filename.rsplit(".", 1)[-1].lower() if upload.filename and "." in upload.filename else ""
        if extension and extension not in allowed:
            raise HTTPException(status_code=400, detail=f"Format file {upload.filename} tidak diizinkan")
    max_file_size_mb = assignment_max_file_size_mb(assignment)
    await validate_upload_file_sizes(upload_files, max_file_size_mb, "file jawaban")
    deadline = datetime.fromisoformat(assignment["deadline"].replace("Z", "+00:00"))
    submit_time = datetime.now(timezone.utc)
    if submit_time > deadline and assignment.get("close_after_deadline") and not revision_open:
        raise HTTPException(status_code=400, detail="Deadline sudah ditutup oleh dosen")
    late_delta = submit_time - deadline if submit_time > deadline else timedelta(0)
    late_hours = round(late_delta.total_seconds() / 3600, 2) if late_delta.total_seconds() > 0 else 0
    late_days = late_delta.days if late_delta.total_seconds() > 0 else 0
    status = "Terlambat" if late_hours > 0 else "Sudah Submit"
    class_doc = await db.classes.find_one({"id": assignment["class_id"]}, {"_id": 0}) or {}
    hierarchy = [
        class_doc.get("academic_year", "Tahun Akademik"),
        class_doc.get("semester", "Semester"),
        assignment.get("course_name", "Mata Kuliah"),
        assignment.get("class_name", "Kelas"),
        assignment.get("title", "Tugas"),
    ]
    submission_id = submission["id"] if submission else new_id()
    saved_files = []
    for upload in upload_files:
        file_doc = await save_uploaded_file_record(
            upload,
            hierarchy,
            user.get("nim", ""),
            user.get("name", "Mahasiswa"),
            user["id"],
            submission_id,
            assignment_id,
            "submission",
            background_tasks=background_tasks,
            async_drive=True,
            lecturer_id=assignment.get("lecturer_id", class_doc.get("lecturer_id", "")),
        )
        saved_files.append(file_doc)
    revision_count = int(submission.get("revision_count", 0)) + 1 if submission else 0
    doc = {
        "id": submission_id,
        "assignment_id": assignment_id,
        "assignment_title": assignment["title"],
        "assignment_deadline": assignment.get("deadline", ""),
        "assignment_late_penalty_per_day": assignment.get("late_penalty_per_day", 0),
        "student_id": user["id"],
        "student_name": user["name"],
        "student_nim": user.get("nim", ""),
        "class_id": assignment["class_id"],
        "lecturer_id": assignment.get("lecturer_id", class_doc.get("lecturer_id", "")),
        "lecturer_name": assignment.get("lecturer_name", class_doc.get("lecturer_name", "")),
        "status": status,
        "review_status": "submitted",
        "note": note,
        "file": saved_files[0],
        "files": saved_files,
        "submitted_at": now_iso(),
        "late_hours": late_hours,
        "late_days": late_days,
        "late_text": f"Terlambat {late_days} hari {round(late_hours % 24, 2)} jam" if late_hours > 0 else "Tepat waktu",
        "revision_count": revision_count,
        "grade": submission.get("grade") if submission else None,
        "feedback": submission.get("feedback", "") if submission else "",
        "revision_note": submission.get("revision_note", "") if submission else "",
        "grade_history": submission.get("grade_history", []) if submission else [],
    }
    await db.submissions.update_one({"id": submission_id}, {"$set": doc}, upsert=True)
    await db.reminder_logs.insert_one(
        {
            "id": new_id(),
            "assignment_id": assignment_id,
            "class_id": assignment["class_id"],
            "lecturer_id": assignment.get("lecturer_id", class_doc.get("lecturer_id", "")),
            "student_id": user["id"],
            "reminder_type": "konfirmasi_submit",
            "sent_at": now_iso(),
            "status": "in_app",
            "response": f"Tugas diterima dengan status {status}",
        }
    )
    return public_doc(doc)


@api_router.post("/submissions/{submission_id}/grade")
async def grade_submission(
    submission_id: str,
    payload: GradeInput,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    submission = await require_submission_access(submission_id, user)
    assignment = await db.assignments.find_one({"id": submission["assignment_id"]}, {"_id": 0}) or {}
    weighted = sum((item.score * item.weight) / 100 for item in payload.rubric_scores)
    penalty = 0.0
    penalty_rate = submission.get("assignment_late_penalty_per_day", assignment.get("late_penalty_per_day", 0))
    if submission.get("status") == "Terlambat" and penalty_rate:
        submitted_at = parse_iso_datetime(submission.get("submitted_at", ""))
        deadline = parse_iso_datetime(submission.get("assignment_deadline", "") or assignment.get("deadline", ""))
        if submitted_at and deadline:
            days_late = max(1, (submitted_at - deadline).days + 1)
            penalty = min(100, float(penalty_rate) * days_late)
    final_grade = min(100, max(0, round(weighted - penalty, 2)))
    grade_predicate = await calculate_grade_predicate(final_grade, submission.get("class_id", assignment.get("class_id", "")))
    history = submission.get("grade_history", [])
    history.append(
        {
            "grade": final_grade,
            "grade_predicate": grade_predicate,
            "rubric_scores": [item.model_dump() for item in payload.rubric_scores],
            "feedback": payload.feedback,
            "revision_note": payload.revision_note,
            "penalty": penalty,
            "graded_by": user["id"],
            "graded_at": now_iso(),
        }
    )
    await db.submissions.update_one(
        {"id": submission_id},
        {
            "$set": {
                "grade": final_grade,
                "grade_predicate": grade_predicate,
                "feedback": payload.feedback,
                "revision_note": payload.revision_note,
                "rubric_scores": [item.model_dump() for item in payload.rubric_scores],
                "late_penalty": penalty,
                "status": payload.status,
                "review_status": "graded",
                "graded_at": now_iso(),
                "grade_history": history,
            }
        },
    )
    await db.reminder_logs.insert_one(
        {
            "id": new_id(),
            "assignment_id": submission["assignment_id"],
            "student_id": submission["student_id"],
            "reminder_type": "nilai_tersedia",
            "sent_at": now_iso(),
            "status": "in_app",
            "response": "Nilai dan feedback sudah tersedia",
        }
    )
    updated = await db.submissions.find_one({"id": submission_id}, {"_id": 0})
    await notify_submission_status_whatsapp(updated or submission, assignment, "nilai_tersedia", background_tasks)
    return updated


@api_router.post("/submissions/{submission_id}/review")
async def review_submission(submission_id: str, user: Dict[str, Any] = Depends(require_admin)):
    await require_submission_access(submission_id, user)
    await db.submissions.update_one(
        {"id": submission_id},
        {"$set": {"review_status": "reviewed", "reviewed_at": now_iso(), "reviewed_by": user["id"]}},
    )
    updated = await db.submissions.find_one({"id": submission_id}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=404, detail="Submission tidak ditemukan")
    return updated


@api_router.post("/submissions/{submission_id}/request-revision")
async def request_revision(
    submission_id: str,
    payload: Dict[str, str],
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    submission = await require_submission_access(submission_id, user)
    await db.submissions.update_one(
        {"id": submission_id},
        {
            "$set": {
                "status": "Direvisi",
                "review_status": "revision_requested",
                "revision_note": payload.get("revision_note", "Perlu revisi"),
                "revision_requested_at": now_iso(),
                "revision_requested_by": user["id"],
            }
        },
    )
    await db.reminder_logs.insert_one(
        {
            "id": new_id(),
            "assignment_id": submission["assignment_id"],
            "student_id": submission["student_id"],
            "reminder_type": "revisi_tugas",
            "sent_at": now_iso(),
            "status": "in_app",
            "response": "Dosen meminta revisi tugas",
        }
    )
    updated = await db.submissions.find_one({"id": submission_id}, {"_id": 0})
    assignment = await db.assignments.find_one({"id": submission["assignment_id"]}, {"_id": 0}) or {}
    await notify_submission_status_whatsapp(updated or submission, assignment, "revisi_tugas", background_tasks)
    return updated


@api_router.post("/submissions/bulk-grade")
async def bulk_grade_submissions(
    payload: BulkGradeInput,
    background_tasks: BackgroundTasks,
    user: Dict[str, Any] = Depends(require_admin),
):
    results = []
    for item in payload.grades:
        try:
            submission = await require_submission_access(item.submission_id, user)
        except HTTPException:
            results.append({"submission_id": item.submission_id, "status": "not_found"})
            continue
        assignment = await db.assignments.find_one({"id": submission["assignment_id"]}, {"_id": 0}) or {}
        rubric = assignment.get("rubric") or [{"criterion": "Nilai total", "weight": 100}]
        rubric_scores = [
            {"criterion": r.get("criterion", "Nilai"), "weight": float(r.get("weight", 0)), "score": float(item.score)}
            for r in rubric
        ]
        weighted = sum((score["score"] * score["weight"]) / 100 for score in rubric_scores)
        penalty = 0.0
        penalty_rate = submission.get("assignment_late_penalty_per_day", assignment.get("late_penalty_per_day", 0))
        if submission.get("status") == "Terlambat" and penalty_rate:
            penalty = min(100, float(penalty_rate) * max(1, int(submission.get("late_days", 0)) + 1))
        final_grade = min(100, max(0, round(weighted - penalty, 2)))
        grade_predicate = await calculate_grade_predicate(final_grade, submission.get("class_id", assignment.get("class_id", "")))
        history = submission.get("grade_history", [])
        history.append(
            {
                "grade": final_grade,
                "grade_predicate": grade_predicate,
                "rubric_scores": rubric_scores,
                "feedback": item.feedback,
                "revision_note": item.revision_note,
                "penalty": penalty,
                "graded_by": user["id"],
                "graded_at": now_iso(),
                "bulk": True,
            }
        )
        await db.submissions.update_one(
            {"id": item.submission_id},
            {
                "$set": {
                    "grade": final_grade,
                    "grade_predicate": grade_predicate,
                    "feedback": item.feedback,
                    "revision_note": item.revision_note,
                    "rubric_scores": rubric_scores,
                    "late_penalty": penalty,
                    "status": "Dinilai",
                    "review_status": "graded",
                    "graded_at": now_iso(),
                    "grade_history": history,
                }
            },
        )
        await db.reminder_logs.insert_one(
            {
                "id": new_id(),
                "assignment_id": submission["assignment_id"],
                "student_id": submission["student_id"],
                "reminder_type": "nilai_tersedia",
                "sent_at": now_iso(),
                "status": "in_app",
                "response": "Nilai dan feedback sudah tersedia",
            }
        )
        updated = await db.submissions.find_one({"id": item.submission_id}, {"_id": 0}) or submission
        await notify_submission_status_whatsapp(updated, assignment, "nilai_tersedia", background_tasks)
        results.append({"submission_id": item.submission_id, "status": "graded", "grade": final_grade, "grade_predicate": grade_predicate})
    return {"updated": len([r for r in results if r["status"] == "graded"]), "results": results}


async def calculate_student_progress(student_id: str) -> Dict[str, Any]:
    student = await db.users.find_one({"id": student_id}, {"_id": 0}) or {}
    class_ids = student.get("class_ids", [])
    assignments = await db.assignments.find({"class_id": {"$in": class_ids}, "is_active": True}, {"_id": 0}).to_list(1000)
    assignments = [assignment for assignment in assignments if assignment_is_published(assignment)]
    submissions = await db.submissions.find({"student_id": student_id}, {"_id": 0}).to_list(1000)
    submitted_ids = {item["assignment_id"] for item in submissions}
    graded = [item for item in submissions if isinstance(item.get("grade"), (int, float))]
    late = len([item for item in submissions if item.get("status") == "Terlambat"])
    missing = len([assignment for assignment in assignments if assignment["id"] not in submitted_ids])
    avg_grade = round(sum(item["grade"] for item in graded) / len(graded), 1) if graded else 0
    inactive_days = 0
    if student.get("last_login_at"):
        try:
            inactive_days = (datetime.now(timezone.utc) - datetime.fromisoformat(student["last_login_at"].replace("Z", "+00:00"))).days
        except Exception:
            inactive_days = 0
    risk_score = 0
    risk_score += min(4, missing) * 2
    risk_score += min(3, late) * 2
    if graded and avg_grade < 60:
        risk_score += 3
    if inactive_days > 14:
        risk_score += 2
    if risk_score >= 8:
        label = "Risiko Tinggi"
    elif risk_score >= 5:
        label = "Perlu Perhatian"
    elif risk_score >= 2:
        label = "Risiko Rendah"
    else:
        label = "Aman"
    return {
        "submitted": len(submitted_ids),
        "missing": missing,
        "late": late,
        "avg_grade": avg_grade,
        "last_login_at": student.get("last_login_at", ""),
        "status": student.get("status", "active"),
        "risk_label": label,
    }


def _compute_student_progress(
    student: Dict[str, Any],
    published_assignments: List[Dict[str, Any]],
    student_submissions: List[Dict[str, Any]],
) -> Dict[str, Any]:
    submitted_ids = {item["assignment_id"] for item in student_submissions}
    graded = [item for item in student_submissions if isinstance(item.get("grade"), (int, float))]
    late = len([item for item in student_submissions if item.get("status") == "Terlambat"])
    missing = len([assignment for assignment in published_assignments if assignment["id"] not in submitted_ids])
    avg_grade = round(sum(item["grade"] for item in graded) / len(graded), 1) if graded else 0
    inactive_days = 0
    if student.get("last_login_at"):
        try:
            inactive_days = (datetime.now(timezone.utc) - datetime.fromisoformat(student["last_login_at"].replace("Z", "+00:00"))).days
        except Exception:
            inactive_days = 0
    risk_score = 0
    risk_score += min(4, missing) * 2
    risk_score += min(3, late) * 2
    if graded and avg_grade < 60:
        risk_score += 3
    if inactive_days > 14:
        risk_score += 2
    if risk_score >= 8:
        label = "Risiko Tinggi"
    elif risk_score >= 5:
        label = "Perlu Perhatian"
    elif risk_score >= 2:
        label = "Risiko Rendah"
    else:
        label = "Aman"
    return {
        "submitted": len(submitted_ids),
        "missing": missing,
        "late": late,
        "avg_grade": avg_grade,
        "last_login_at": student.get("last_login_at", ""),
        "status": student.get("status", "active"),
        "risk_label": label,
    }


async def calculate_student_progress_many(
    student_ids: List[str], scoped_class_ids: Optional[List[str]] = None
) -> Dict[str, Dict[str, Any]]:
    if not student_ids:
        return {}
    students = await db.users.find({"id": {"$in": student_ids}, "role": "student"}, {"_id": 0}).to_list(len(student_ids) * 2)
    students_map: Dict[str, Dict[str, Any]] = {s["id"]: s for s in students}
    all_class_ids: set = set()
    for s in students:
        student_classes = s.get("class_ids", [])
        if scoped_class_ids is not None:
            student_classes = [class_id for class_id in student_classes if class_id in scoped_class_ids]
        all_class_ids.update(student_classes)
    all_assignments: List[Dict[str, Any]] = []
    if all_class_ids:
        all_assignments = await db.assignments.find(
            {"class_id": {"$in": list(all_class_ids)}, "is_active": True}, {"_id": 0}
        ).to_list(5000)
    published_assignments = [a for a in all_assignments if assignment_is_published(a)]
    class_to_assignments: Dict[str, List[Dict[str, Any]]] = {}
    for a in published_assignments:
        cid = a.get("class_id", "")
        class_to_assignments.setdefault(cid, []).append(a)
    all_submissions: List[Dict[str, Any]] = []
    if student_ids:
        all_submissions = await db.submissions.find(
            {"student_id": {"$in": student_ids}}, {"_id": 0}
        ).to_list(5000)
    subs_by_student: Dict[str, List[Dict[str, Any]]] = {}
    for s in all_submissions:
        sid = s.get("student_id", "")
        subs_by_student.setdefault(sid, []).append(s)
    result: Dict[str, Dict[str, Any]] = {}
    for sid in student_ids:
        student = students_map.get(sid, {})
        student_class_ids = student.get("class_ids", [])
        if scoped_class_ids is not None:
            student_class_ids = [class_id for class_id in student_class_ids if class_id in scoped_class_ids]
        student_assignments = []
        for cid in student_class_ids:
            student_assignments.extend(class_to_assignments.get(cid, []))
        student_subs = subs_by_student.get(sid, [])
        if scoped_class_ids is not None:
            student_subs = [
                submission
                for submission in student_subs
                if submission.get("class_id") in scoped_class_ids
            ]
        result[sid] = _compute_student_progress(student, student_assignments, student_subs)
    return result


@api_router.get("/progress")
async def progress(user: Dict[str, Any] = Depends(get_current_user)):
    if user["role"] == "student":
        return {"student": public_doc(user.copy()), "progress": await calculate_student_progress(user["id"])}
    class_ids = await lecturer_class_ids(user)
    student_ids = list({
        student_id
        for item in await db.classes.find({"id": {"$in": class_ids}}, {"_id": 0, "student_ids": 1}).to_list(5000)
        for student_id in item.get("student_ids", [])
    })
    students = await db.users.find({"id": {"$in": student_ids}, "role": "student"}, {"_id": 0, "password_hash": 0}).to_list(2000)
    progress_map = await calculate_student_progress_many([s["id"] for s in students], class_ids)
    for student in students:
        student["progress"] = progress_map.get(student["id"], {})
    return students


@api_router.get("/grade-predicates")
async def get_grade_predicates(class_id: str = "", user: Dict[str, Any] = Depends(require_admin)):
    if class_id:
        await require_class_access(class_id, user)
    elif user.get("role") == "lecturer":
        return {"class_id": "", "predicates": DEFAULT_GRADE_PREDICATES}
    predicates = await get_grade_predicates_for_class(class_id)
    return {"class_id": class_id, "predicates": predicates}


@api_router.put("/grade-predicates")
async def save_grade_predicates(payload: GradePredicateInput, user: Dict[str, Any] = Depends(require_admin)):
    if payload.class_id:
        await require_class_access(payload.class_id, user)
    elif user.get("role") == "lecturer":
        raise HTTPException(status_code=400, detail="Dosen harus memilih kelas untuk menyimpan predikat")
    predicates = validate_grade_predicates([item.model_dump() for item in payload.predicates])
    doc = {
        "id": payload.class_id or "global",
        "class_id": payload.class_id,
        "predicates": predicates,
        "updated_at": now_iso(),
        "updated_by": user["id"],
    }
    await db.grade_predicates.update_one({"class_id": payload.class_id}, {"$set": doc}, upsert=True)
    return doc


@api_router.get("/calendar")
async def calendar(user: Dict[str, Any] = Depends(get_current_user)):
    class_filter: Dict[str, Any] = {}
    if user["role"] == "student":
        class_filter = {"class_id": {"$in": user.get("class_ids", [])}}
    elif user["role"] == "lecturer":
        class_filter = {"class_id": {"$in": await lecturer_class_ids(user)}}
    assignments = await db.assignments.find({**class_filter, "is_active": True}, {"_id": 0}).to_list(1000)
    if user["role"] == "student":
        assignments = [assignment for assignment in assignments if assignment_is_published(assignment)]
    materials = await db.materials.find(class_filter, {"_id": 0}).to_list(1000)
    events = []
    for assignment in assignments:
        if user["role"] in {"admin", "lecturer"} and assignment.get("published_at") and not assignment_is_published(assignment):
            events.append(
                {
                    "id": f"{assignment['id']}-publish",
                    "type": "tayang",
                    "title": f"Tayang: {assignment['title']}",
                    "date": assignment["published_at"],
                    "class_name": assignment.get("class_name", ""),
                }
            )
        events.append(
            {
                "id": assignment["id"],
                "type": "deadline",
                "title": assignment["title"],
                "date": assignment["deadline"],
                "class_name": assignment.get("class_name", ""),
            }
        )
    for material in materials:
        if material.get("locked_until"):
            events.append(
                {"id": material["id"], "type": "materi", "title": material["title"], "date": material["locked_until"]}
            )
    return sorted(events, key=lambda item: item.get("date", ""))


@api_router.post("/reminders/send")
async def send_reminder(payload: ReminderInput, user: Dict[str, Any] = Depends(require_admin)):
    assignment = await db.assignments.find_one({"id": payload.assignment_id}, {"_id": 0})
    if not assignment:
        raise HTTPException(status_code=404, detail="Tugas tidak ditemukan")
    class_doc = await require_class_access(assignment.get("class_id", ""), user)
    doc = payload.model_dump()
    doc.update({
        "id": new_id(),
        "class_id": assignment.get("class_id", ""),
        "lecturer_id": class_doc.get("lecturer_id", user["id"]),
        "sent_at": now_iso(),
        "status": "in_app",
        "response": "Reminder tersimpan di aplikasi",
    })
    await db.reminder_logs.insert_one(doc)
    return public_doc(doc)


@api_router.get("/reminders")
async def list_reminders(user: Dict[str, Any] = Depends(get_current_user)):
    query: Dict[str, Any] = {}
    if user["role"] == "student":
        query = {"student_id": user["id"]}
    elif user["role"] == "lecturer":
        query = {"class_id": {"$in": await lecturer_class_ids(user)}}
    reminders = await db.reminder_logs.find(query, {"_id": 0}).sort("sent_at", -1).to_list(1000)
    if user["role"] == "student":
        assignment_ids = [item.get("assignment_id") for item in reminders if item.get("assignment_id")]
        if assignment_ids:
            assignments = await db.assignments.find({"id": {"$in": assignment_ids}}, {"_id": 0}).to_list(1000)
            visible_ids = {item["id"] for item in assignments if assignment_is_published(item)}
            reminders = [item for item in reminders if not item.get("assignment_id") or item.get("assignment_id") in visible_ids]
    return reminders


@api_router.get("/reports/grades.xlsx")
async def export_grades(user: Dict[str, Any] = Depends(require_admin)):
    submissions = await db.submissions.find(
        {"class_id": {"$in": await lecturer_class_ids(user)}}, {"_id": 0}
    ).sort("student_name", 1).to_list(5000)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rekap Nilai"
    sheet.append(["NIM", "Nama", "Tugas", "Status", "Nilai", "Predikat", "Feedback", "Waktu Submit"])
    for item in submissions:
        sheet.append(
            [
                item.get("student_nim", ""),
                item.get("student_name", ""),
                item.get("assignment_title", ""),
                item.get("status", ""),
                item.get("grade", ""),
                item.get("grade_predicate", ""),
                item.get("feedback", ""),
                item.get("submitted_at", ""),
            ]
        )
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    headers = {"Content-Disposition": "attachment; filename=rekap-nilai.xlsx"}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@api_router.get("/reports/summary")
async def report_summary(user: Dict[str, Any] = Depends(require_admin)):
    class_ids = await lecturer_class_ids(user)
    class_docs = await db.classes.find({"id": {"$in": class_ids}}, {"_id": 0, "student_ids": 1}).to_list(5000)
    student_ids = list({student_id for item in class_docs for student_id in item.get("student_ids", [])})
    students = await db.users.find({"id": {"$in": student_ids}, "role": "student"}, {"_id": 0, "password_hash": 0}).to_list(2000)
    submissions = await db.submissions.find({"class_id": {"$in": class_ids}}, {"_id": 0}).to_list(5000)
    assignments = await db.assignments.find({"class_id": {"$in": class_ids}}, {"_id": 0}).to_list(1000)
    return {
        "total_students": len(students),
        "total_assignments": len(assignments),
        "total_submissions": len(submissions),
        "graded_submissions": len([item for item in submissions if item.get("status") == "Dinilai"]),
        "late_submissions": len([item for item in submissions if item.get("status") == "Terlambat"]),
    }


@api_router.get("/reports/grade-recap")
async def grade_recap(class_id: Optional[str] = None, user: Dict[str, Any] = Depends(require_admin)):
    class_ids = await lecturer_class_ids(user)
    if class_id and class_id not in class_ids:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan")
    query = {"class_id": class_id} if class_id else {"class_id": {"$in": class_ids}}
    submissions = await db.submissions.find({**query, "grade": {"$ne": None}}, {"_id": 0}).to_list(5000)
    classes = await db.classes.find({"id": {"$in": class_ids}}, {"_id": 0}).to_list(500)
    class_map = {c["id"]: c for c in classes}

    class_groups: Dict[str, Any] = {}
    for s in submissions:
        cid = s.get("class_id", "unknown")
        if cid not in class_groups:
            entry = class_map.get(cid, {})
            class_groups[cid] = {
                "class_id": cid,
                "course_name": s.get("course_name") or entry.get("course_name", ""),
                "class_name": s.get("class_name") or entry.get("name", ""),
                "course_id": s.get("course_id") or entry.get("course_id", ""),
                "students": {},
                "assignment_ids": set(),
            }
        g = class_groups[cid]
        sid = s["student_id"]
        if sid not in g["students"]:
            g["students"][sid] = {
                "student_id": sid,
                "student_name": s["student_name"],
                "student_nim": s.get("student_nim", ""),
                "scores": [],
            }
        g["students"][sid]["scores"].append({
            "assignment_id": s["assignment_id"],
            "assignment_title": s.get("assignment_title", ""),
            "grade": s["grade"],
            "grade_predicate": s.get("grade_predicate", ""),
        })
        g["assignment_ids"].add(s["assignment_id"])

    result = []
    for cid, g in class_groups.items():
        students_list = list(g["students"].values())
        all_scores = []
        for student in students_list:
            scores_list = [sc["grade"] for sc in student["scores"]]
            student["average"] = round(sum(scores_list) / len(scores_list), 1) if scores_list else 0
            student["total_graded"] = len(student["scores"])
            all_scores.extend(scores_list)

        class_avg = round(sum(all_scores) / len(all_scores), 1) if all_scores else 0

        dist = {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0}
        for score in all_scores:
            if score >= 85: dist["A"] += 1
            elif score >= 70: dist["B"] += 1
            elif score >= 60: dist["C"] += 1
            elif score >= 50: dist["D"] += 1
            else: dist["E"] += 1

        result.append({
            "class_id": cid,
            "course_name": g["course_name"],
            "class_name": g["class_name"],
            "course_id": g["course_id"],
            "student_count": len(students_list),
            "total_assignments": len(g["assignment_ids"]),
            "class_average": class_avg,
            "grade_distribution": dist,
            "students": sorted(students_list, key=lambda x: x["student_name"]),
        })

    return sorted(result, key=lambda x: (x["course_name"], x["class_name"]))


@api_router.get("/settings")
async def get_settings(_: Dict[str, Any] = Depends(get_current_user)):
    settings = await db.app_settings.find_one({"id": "main"}, {"_id": 0})
    return settings or default_app_settings()


@api_router.get("/settings/public")
async def get_public_settings():
    settings = await db.app_settings.find_one({"id": "main"}, {"_id": 0}) or default_app_settings()
    return {
        "app_name": settings.get("app_name") or default_app_settings()["app_name"],
        "campus_name": settings.get("campus_name", ""),
        "campus_logo_url": settings.get("campus_logo_url", ""),
    }


@api_router.put("/settings")
async def update_settings(payload: AppSettingsInput, user: Dict[str, Any] = Depends(require_campus_admin)):
    doc = payload.model_dump()
    doc.update({"id": "main", "updated_at": now_iso(), "updated_by": user["id"]})
    await db.app_settings.update_one({"id": "main"}, {"$set": doc}, upsert=True)
    _invalidate_settings_cache("app_settings")
    return await db.app_settings.find_one({"id": "main"}, {"_id": 0})


@api_router.post("/academic-years/rollover-preview")
async def rollover_preview(user: Dict[str, Any] = Depends(require_admin)):
    classes = await db.classes.find(
        {"id": {"$in": await lecturer_class_ids(user)}, "status": "active"}, {"_id": 0}
    ).to_list(500)
    return {
        "recommended_flow": [
            "Export rekap nilai dan arsip kelas semester lama.",
            "Ubah active_academic_year dan active_semester pada Settings.",
            "Duplikasi kelas yang masih dipakai ke tahun ajaran baru dengan kode kelas baru.",
            "Mahasiswa lama tidak otomatis ikut kelas baru; mereka mengajukan kode kelas dan dosen approve.",
            "Materi/tugas semester lama tetap bisa dilihat di arsip, sedangkan submission baru mengikuti deadline kelas baru.",
        ],
        "active_classes_to_archive": len(classes),
        "classes": classes,
    }


@app.websocket("/api/chat/ws")
async def chat_websocket(websocket: WebSocket, token: str = ""):
    session = await db.sessions.find_one({"token": token.strip()}, {"_id": 0}) if token.strip() else None
    if not session:
        await websocket.close(code=1008)
        return
    user = await db.users.find_one({"id": session["user_id"]}, {"_id": 0, "password_hash": 0})
    if not user:
        await websocket.close(code=1008)
        return
    if not await chat_connections.connect(user["id"], websocket):
        return
    try:
        while True:
            payload = json.loads(await websocket.receive_text())
            if payload.get("type") == "viewing":
                target_id = str(payload.get("user_id", ""))
                if target_id and target_id != user["id"]:
                    await chat_connections.set_viewing(user["id"], websocket, target_id)
                else:
                    await chat_connections.set_viewing(user["id"], websocket, "")
            elif payload.get("type") == "ping":
                await websocket.send_json({"type": "pong"})
    except (WebSocketDisconnect, json.JSONDecodeError):
        await chat_connections.disconnect(user["id"], websocket)


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(GZipMiddleware, minimum_size=1000)


@app.on_event("startup")
async def on_startup():
    global _database_backup_scheduler_task
    await db.connect()
    await seed_data()
    await load_oidc_runtime_settings()
    await ensure_program_course_links()
    await ensure_multi_lecturer_schema()
    await db.users.update_one({"email": "dosen@demo.id", "username": {"$exists": False}}, {"$set": {"username": "dosenadmin", "whatsapp": "628000000001"}})
    async for student in db.users.find({"role": "student", "username": {"$exists": False}}, {"_id": 0, "id": 1, "nim": 1}):
        if student.get("nim"):
            await db.users.update_one({"id": student["id"]}, {"$set": {"username": str(student["nim"]).lower()}})
    async for assignment in db.assignments.find(
        {
            "$or": [{"material_id": {"$exists": False}}, {"material_id": ""}, {"material_id": None}],
            "material_link_removed_at": {"$exists": False},
        },
        {"_id": 0, "id": 1, "class_id": 1},
    ):
        material = await db.materials.find_one({"class_id": assignment.get("class_id")}, {"_id": 0})
        if material:
            await db.assignments.update_one({"id": assignment["id"]}, {"$set": {"material_id": material["id"]}})
    await db.users.create_index("email", unique=True)
    await db.users.create_index("username", unique=True, sparse=True)
    await db.users.create_index("name")
    await db.users.create_index([("sso_issuer", 1), ("sso_subject", 1)], unique=True, sparse=True)
    await db.sessions.create_index("token", unique=True)
    await db.database_backups.create_index("created_at")
    await db.oidc_flows.create_index("state_hash", unique=True)
    await db.oidc_flows.create_index("expires_at", expireAfterSeconds=0)
    await db.oidc_login_tickets.create_index("ticket_hash", unique=True)
    await db.oidc_login_tickets.create_index("expires_at", expireAfterSeconds=0)
    await db.chat_messages.create_index([("conversation_id", 1), ("created_at", 1)])
    await db.assignments.create_index("class_id")
    await db.assignments.create_index([("is_active", 1), ("published_at", 1)])
    await db.assignments.create_index("deadline")
    await db.submissions.create_index([("assignment_id", 1), ("student_id", 1)])
    await db.submissions.create_index("student_id")
    await db.submissions.create_index("submitted_at")
    await db.submissions.create_index("student_name")
    await db.materials.create_index("class_id")
    await db.materials.create_index("created_at")
    await db.classes.create_index("class_code")
    await db.classes.create_index("status")
    await db.classes.create_index([("lecturer_id", 1), ("status", 1)])
    await db.classes.create_index("created_at")
    await db.programs.create_index("created_at")
    await db.courses.create_index("created_at")
    await db.materials.create_index("lecturer_id")
    await db.submissions.create_index("lecturer_id")
    await db.stored_files.create_index("drive_sync_status")
    await db.stored_files.create_index("record_type")
    await db.stored_files.create_index("lecturer_id")
    await db.stored_files.create_index("uploaded_at")
    await db.comments.create_index("material_id")
    await db.comments.create_index("created_at")
    await db.enrollment_requests.create_index([("class_id", 1), ("student_id", 1)])
    await db.enrollment_requests.create_index("requested_at")
    await db.reminder_logs.create_index("assignment_id")
    await db.reminder_logs.create_index("sent_at")
    await db.whatsapp_messages.create_index("status")
    await db.whatsapp_messages.create_index("created_at")
    await db.password_reset_requests.create_index("requested_at")
    _database_backup_scheduler_task = asyncio.create_task(database_backup_scheduler())


@app.on_event("shutdown")
async def on_shutdown():
    global _database_backup_scheduler_task
    if _database_backup_scheduler_task:
        _database_backup_scheduler_task.cancel()
        try:
            await _database_backup_scheduler_task
        except asyncio.CancelledError:
            pass
        _database_backup_scheduler_task = None


@app.on_event("shutdown")
async def shutdown_db_client():
    await db.close()
