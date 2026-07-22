"""Iteration 7 regression: grade predicate defaults, validation, grading integration, and export."""

import io
import os
import uuid
import asyncio
from datetime import datetime, timedelta, timezone

import asyncpg
import pytest
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook


load_dotenv("/app/frontend/.env")
load_dotenv("/app/backend/.env")

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL")
assert BASE_URL, "REACT_APP_BACKEND_URL is required"
API_BASE = f"{BASE_URL.rstrip('/')}/api"

DATABASE_URL = os.environ.get("DATABASE_URL")
assert DATABASE_URL, "DATABASE_URL is required"
DATABASE_URL = DATABASE_URL.replace("postgresql+asyncpg://", "postgresql://", 1)

DEFAULT_PREDICATES = [
    {"label": "A", "min_score": 85, "max_score": 100},
    {"label": "B", "min_score": 70, "max_score": 84.99},
    {"label": "C", "min_score": 60, "max_score": 69.99},
    {"label": "D", "min_score": 50, "max_score": 59.99},
    {"label": "E", "min_score": 0, "max_score": 49.99},
]


def _postgres_execute(sql: str, *params) -> None:
    async def execute():
        connection = await asyncpg.connect(DATABASE_URL)
        try:
            await connection.execute(sql, *params)
        finally:
            await connection.close()

    asyncio.run(execute())


def _auth(identifier: str, password: str) -> dict:
    # Auth module: unified login helper for admin/student tests.
    resp = requests.post(
        f"{API_BASE}/auth/login",
        json={"identifier": identifier, "password": password},
        timeout=45,
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert isinstance(body.get("token"), str) and body["token"]
    return body


def _headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _normalize_predicates(rows: list[dict]) -> list[tuple[str, float, float]]:
    return sorted(
        [
            (
                str(item["label"]).upper(),
                float(item["min_score"]),
                float(item["max_score"]),
            )
            for item in rows
        ],
        key=lambda x: x[1],
    )


def _create_assignment(admin_headers: dict, class_id: str, title_prefix: str = "IT7") -> dict:
    # Assignment module: helper to create assignment for seeded class.
    suffix = uuid.uuid4().hex[:6]
    payload = {
        "class_id": class_id,
        "title": f"{title_prefix}_Predicate_{suffix}",
        "description": "Iteration 7 predicate test assignment",
        "deadline": (datetime.now(timezone.utc) + timedelta(days=2)).isoformat(),
        "tolerance_hours": 0,
        "allowed_formats": ["pdf"],
        "rubric": [{"criterion": "Nilai total", "weight": 100}],
        "assignment_type": "individu",
        "allow_revision": True,
        "is_active": True,
        "is_practicum": False,
        "practicum_goal": "",
        "practicum_tools": "",
        "practicum_steps": [],
        "required_screenshot": False,
        "late_penalty_per_day": 0,
        "close_after_deadline": False,
        "material_id": "",
    }
    resp = requests.post(f"{API_BASE}/assignments", headers=admin_headers, json=payload, timeout=45)
    assert resp.status_code == 200, resp.text
    doc = resp.json()
    assert doc.get("class_id") == class_id
    return doc


def _submit_assignment(student_headers: dict, assignment_id: str, note: str) -> dict:
    # Submission module: helper to upload one file to generate submission.
    resp = requests.post(
        f"{API_BASE}/assignments/{assignment_id}/submit",
        headers=student_headers,
        files=[("files", ("it7_submission.pdf", b"%PDF-it7", "application/pdf"))],
        data={"note": note},
        timeout=45,
    )
    assert resp.status_code == 200, resp.text
    doc = resp.json()
    assert isinstance(doc.get("id"), str)
    return doc


@pytest.fixture(scope="module")
def auth_state():
    admin = _auth("dosenadmin", "Dosen123!")
    admin_headers = _headers(admin["token"])

    # Auth bootstrap module: recover seeded student password if prior tests changed it.
    student_login = requests.post(
        f"{API_BASE}/auth/login",
        json={"identifier": "230001001", "password": "Mahasiswa123!"},
        timeout=45,
    )
    if student_login.status_code != 200:
        students_resp = requests.get(f"{API_BASE}/students", headers=admin_headers, timeout=45)
        assert students_resp.status_code == 200, students_resp.text
        seeded = next((s for s in students_resp.json() if s.get("nim") == "230001001"), None)
        assert seeded is not None, "Seeded student 230001001 not found"
        reset_resp = requests.post(
            f"{API_BASE}/students/{seeded['id']}/reset-password",
            headers=admin_headers,
            json={"password": "Mahasiswa123!"},
            timeout=45,
        )
        assert reset_resp.status_code == 200, reset_resp.text
        student = _auth("230001001", "Mahasiswa123!")
    else:
        student = student_login.json()

    student_headers = _headers(student["token"])

    classes_resp = requests.get(f"{API_BASE}/classes", headers=student_headers, timeout=45)
    assert classes_resp.status_code == 200, classes_resp.text
    classes = classes_resp.json()
    assert classes, "Seeded student has no class"

    return {
        "admin_headers": admin_headers,
        "student_headers": student_headers,
        "class_id": classes[0]["id"],
    }


@pytest.fixture(autouse=True)
def restore_global_defaults(auth_state):
    # Predicate module: keep global predicate config deterministic after each test.
    yield
    requests.put(
        f"{API_BASE}/grade-predicates",
        headers=auth_state["admin_headers"],
        json={"class_id": "", "predicates": DEFAULT_PREDICATES},
        timeout=45,
    )


def test_default_grade_predicates_returned_when_global_override_missing(auth_state):
    # Predicate module: default A-E ranges are used when global db override is absent.
    _postgres_execute(
        "DELETE FROM app_doc_grade_predicates WHERE data->>'class_id' = ''"
    )

    resp = requests.get(f"{API_BASE}/grade-predicates", headers=auth_state["admin_headers"], timeout=45)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body.get("class_id") == ""
    assert _normalize_predicates(body.get("predicates", [])) == _normalize_predicates(DEFAULT_PREDICATES)


def test_admin_can_put_and_get_grade_predicates_global_and_class(auth_state):
    # Predicate module: admin can update/fetch global and per-class ranges.
    admin_headers = auth_state["admin_headers"]
    class_id = auth_state["class_id"]

    global_payload = {
        "class_id": "",
        "predicates": [
            {"label": "A", "min_score": 90, "max_score": 100},
            {"label": "B", "min_score": 80, "max_score": 89.99},
            {"label": "C", "min_score": 70, "max_score": 79.99},
            {"label": "D", "min_score": 60, "max_score": 69.99},
            {"label": "E", "min_score": 0, "max_score": 59.99},
        ],
    }
    put_global = requests.put(f"{API_BASE}/grade-predicates", headers=admin_headers, json=global_payload, timeout=45)
    assert put_global.status_code == 200, put_global.text
    assert put_global.json().get("class_id") == ""

    get_global = requests.get(f"{API_BASE}/grade-predicates", headers=admin_headers, timeout=45)
    assert get_global.status_code == 200, get_global.text
    assert _normalize_predicates(get_global.json()["predicates"]) == _normalize_predicates(global_payload["predicates"])

    class_payload = {
        "class_id": class_id,
        "predicates": [
            {"label": "A", "min_score": 88, "max_score": 100},
            {"label": "B", "min_score": 75, "max_score": 87.99},
            {"label": "C", "min_score": 60, "max_score": 74.99},
            {"label": "D", "min_score": 45, "max_score": 59.99},
            {"label": "E", "min_score": 0, "max_score": 44.99},
        ],
    }
    put_class = requests.put(f"{API_BASE}/grade-predicates", headers=admin_headers, json=class_payload, timeout=45)
    assert put_class.status_code == 200, put_class.text
    assert put_class.json().get("class_id") == class_id

    get_class = requests.get(
        f"{API_BASE}/grade-predicates",
        headers=admin_headers,
        params={"class_id": class_id},
        timeout=45,
    )
    assert get_class.status_code == 200, get_class.text
    assert _normalize_predicates(get_class.json()["predicates"]) == _normalize_predicates(class_payload["predicates"])


def test_grade_predicate_put_validates_overlap_and_invalid_ranges(auth_state):
    # Validation module: overlapping and invalid min/max ranges are rejected.
    admin_headers = auth_state["admin_headers"]

    overlap_payload = {
        "class_id": "",
        "predicates": [
            {"label": "A", "min_score": 80, "max_score": 100},
            {"label": "B", "min_score": 70, "max_score": 85},
        ],
    }
    overlap_resp = requests.put(f"{API_BASE}/grade-predicates", headers=admin_headers, json=overlap_payload, timeout=45)
    assert overlap_resp.status_code == 400, overlap_resp.text
    assert "tumpang tindih" in overlap_resp.json().get("detail", "").lower()

    invalid_payload = {
        "class_id": "",
        "predicates": [
            {"label": "A", "min_score": 95, "max_score": 85},
            {"label": "B", "min_score": 0, "max_score": 84.99},
        ],
    }
    invalid_resp = requests.put(f"{API_BASE}/grade-predicates", headers=admin_headers, json=invalid_payload, timeout=45)
    assert invalid_resp.status_code == 400, invalid_resp.text
    assert "minimum <= maksimum" in invalid_resp.json().get("detail", "")


def test_single_grading_stores_grade_predicate_automatically(auth_state):
    # Grading module: single grade endpoint stores grade_predicate from score.
    admin_headers = auth_state["admin_headers"]
    student_headers = auth_state["student_headers"]
    class_id = auth_state["class_id"]

    requests.put(
        f"{API_BASE}/grade-predicates",
        headers=admin_headers,
        json={"class_id": class_id, "predicates": DEFAULT_PREDICATES},
        timeout=45,
    )

    assignment = _create_assignment(admin_headers, class_id, title_prefix="IT7_SINGLE")
    submission = _submit_assignment(student_headers, assignment["id"], "single grade predicate")

    grade_payload = {
        "rubric_scores": [{"criterion": "Nilai total", "weight": 100, "score": 92}],
        "feedback": "Nilai single grading",
        "revision_note": "",
        "status": "Dinilai",
    }
    grade_resp = requests.post(
        f"{API_BASE}/submissions/{submission['id']}/grade",
        headers=admin_headers,
        json=grade_payload,
        timeout=45,
    )
    assert grade_resp.status_code == 200, grade_resp.text
    graded = grade_resp.json()
    assert graded.get("grade") == 92
    assert graded.get("grade_predicate") == "A"


def test_bulk_grading_stores_and_returns_grade_predicate(auth_state):
    # Grading module: bulk-grade writes predicate and returns it in API results.
    admin_headers = auth_state["admin_headers"]
    student_headers = auth_state["student_headers"]
    class_id = auth_state["class_id"]

    requests.put(
        f"{API_BASE}/grade-predicates",
        headers=admin_headers,
        json={"class_id": class_id, "predicates": DEFAULT_PREDICATES},
        timeout=45,
    )

    assignment = _create_assignment(admin_headers, class_id, title_prefix="IT7_BULK")
    submission = _submit_assignment(student_headers, assignment["id"], "bulk grade predicate")

    bulk_payload = {
        "grades": [
            {
                "submission_id": submission["id"],
                "score": 72,
                "feedback": "Nilai bulk grading",
                "revision_note": "",
            }
        ]
    }
    bulk_resp = requests.post(f"{API_BASE}/submissions/bulk-grade", headers=admin_headers, json=bulk_payload, timeout=45)
    assert bulk_resp.status_code == 200, bulk_resp.text
    body = bulk_resp.json()
    assert body.get("updated") == 1
    assert body["results"][0].get("grade_predicate") == "B"

    list_resp = requests.get(f"{API_BASE}/submissions", headers=admin_headers, timeout=45)
    assert list_resp.status_code == 200, list_resp.text
    rows = list_resp.json()
    target = next((r for r in rows if r.get("id") == submission["id"]), None)
    assert target is not None
    assert target.get("grade_predicate") == "B"


def test_list_submissions_enriches_missing_grade_predicate(auth_state):
    # Submission module: graded docs missing grade_predicate are enriched in /submissions response.
    admin_headers = auth_state["admin_headers"]
    student_headers = auth_state["student_headers"]
    class_id = auth_state["class_id"]

    requests.put(
        f"{API_BASE}/grade-predicates",
        headers=admin_headers,
        json={"class_id": class_id, "predicates": DEFAULT_PREDICATES},
        timeout=45,
    )

    assignment = _create_assignment(admin_headers, class_id, title_prefix="IT7_ENRICH")
    submission = _submit_assignment(student_headers, assignment["id"], "enrich predicate")
    grade_payload = {
        "rubric_scores": [{"criterion": "Nilai total", "weight": 100, "score": 65}],
        "feedback": "Nilai untuk enrich",
        "revision_note": "",
        "status": "Dinilai",
    }
    grade_resp = requests.post(
        f"{API_BASE}/submissions/{submission['id']}/grade",
        headers=admin_headers,
        json=grade_payload,
        timeout=45,
    )
    assert grade_resp.status_code == 200, grade_resp.text
    assert grade_resp.json().get("grade_predicate") == "C"

    _postgres_execute(
        "UPDATE app_doc_submissions SET data = data - 'grade_predicate', updated_at = NOW() "
        "WHERE data->>'id' = $1",
        submission["id"],
    )

    list_resp = requests.get(f"{API_BASE}/submissions", headers=admin_headers, timeout=45)
    assert list_resp.status_code == 200, list_resp.text
    rows = list_resp.json()
    target = next((r for r in rows if r.get("id") == submission["id"]), None)
    assert target is not None
    assert target.get("grade") == 65
    assert target.get("grade_predicate") == "C"


def test_export_excel_contains_predikat_column(auth_state):
    # Report module: exported grade sheet includes Predikat header column.
    resp = requests.get(f"{API_BASE}/reports/grades.xlsx", headers=auth_state["admin_headers"], timeout=45)
    assert resp.status_code == 200, resp.text
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp.headers.get("content-type", "")

    workbook = load_workbook(io.BytesIO(resp.content))
    sheet = workbook.active
    headers = [sheet.cell(row=1, column=idx).value for idx in range(1, 9)]
    assert headers[0] == "NIM"
    assert "Predikat" in headers
