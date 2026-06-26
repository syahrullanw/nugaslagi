"""MVP regression tests for e-learning lecturer/student workflows."""

import io
import os
import uuid
from datetime import datetime, timedelta, timezone

import pytest
import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook


load_dotenv("/app/frontend/.env")

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL").rstrip("/")
API_BASE = f"{BASE_URL}/api"


@pytest.fixture(scope="module")
def state():
    """Shared IDs and credentials created during this suite run."""
    return {
        "course_id": "",
        "program_id": "",
        "class_id": "",
        "class_code": "",
        "student_email": f"test.student.{uuid.uuid4().hex[:8]}@demo.id",
        "student_password": "Mahasiswa123!",
        "joined_email": f"test.join.{uuid.uuid4().hex[:8]}@demo.id",
        "joined_password": "Mahasiswa123!",
        "material_id": "",
        "assignment_id": "",
        "submission_id": "",
    }


@pytest.fixture(scope="module")
def admin_headers():
    """Admin auth header for protected endpoints."""
    response = requests.post(
        f"{API_BASE}/auth/login",
        json={"email": "dosen@demo.id", "password": "Dosen123!"},
        timeout=30,
    )
    assert response.status_code == 200, response.text
    data = response.json()
    assert data["user"]["role"] == "admin"
    return {"Authorization": f"Bearer {data['token']}"}


def login_student(email: str, password: str) -> dict:
    """Utility to obtain student auth header."""
    response = requests.post(
        f"{API_BASE}/auth/login",
        json={"email": email, "password": password},
        timeout=30,
    )
    assert response.status_code == 200, response.text
    data = response.json()
    assert data["user"]["role"] == "student"
    return {"Authorization": f"Bearer {data['token']}"}


# --- Core API health and authentication ---
def test_api_root_and_drive_not_configured():
    response = requests.get(f"{API_BASE}/", timeout=30)
    assert response.status_code == 200
    data = response.json()
    assert data["message"] == "E-Learning Dosen API aktif"
    assert data["drive_configured"] is False


def test_admin_login_credentials_work():
    response = requests.post(
        f"{API_BASE}/auth/login",
        json={"email": "dosen@demo.id", "password": "Dosen123!"},
        timeout=30,
    )
    assert response.status_code == 200
    data = response.json()
    assert isinstance(data.get("token"), str)
    assert data["user"]["email"] == "dosen@demo.id"


def test_seeded_student_login_credentials_work():
    response = requests.post(
        f"{API_BASE}/auth/login",
        json={"email": "alya@demo.id", "password": "Mahasiswa123!"},
        timeout=30,
    )
    assert response.status_code == 200
    data = response.json()
    assert isinstance(data.get("token"), str)
    assert data["user"]["email"] == "alya@demo.id"


def test_student_join_class_flow_with_unique_email(state):
    payload = {
        "class_code": "WEB4A1",
        "nim": f"TEST{uuid.uuid4().hex[:7]}",
        "name": "TEST Join Student",
        "email": state["joined_email"],
        "whatsapp": "628111111111",
        "password": state["joined_password"],
    }
    response = requests.post(f"{API_BASE}/auth/join-class", json=payload, timeout=30)
    assert response.status_code == 200, response.text
    data = response.json()
    assert data["user"]["email"] == state["joined_email"]
    assert data["user"]["role"] == "student"


# --- Admin dashboard and course/class management ---
def test_admin_dashboard_stats_and_drive_warning(admin_headers):
    response = requests.get(f"{API_BASE}/dashboard", headers=admin_headers, timeout=30)
    assert response.status_code == 200
    summary = response.json()["summary"]
    assert isinstance(summary["active_courses"], int)
    assert isinstance(summary["active_classes"], int)
    assert summary["drive_configured"] is False


def test_admin_create_course_and_class_generate_code(admin_headers, state):
    program_payload = {
        "code": f"TST{uuid.uuid4().hex[:4].upper()}",
        "name": "TEST Program Automation",
        "description": "Created by pytest",
    }
    program_resp = requests.post(f"{API_BASE}/programs", json=program_payload, headers=admin_headers, timeout=30)
    assert program_resp.status_code == 200, program_resp.text
    program = program_resp.json()
    state["program_id"] = program["id"]

    course_payload = {
        "program_id": state["program_id"],
        "code": f"TEST{uuid.uuid4().hex[:4].upper()}",
        "name": "TEST Course Automation",
        "credits": 3,
        "description": "Created by pytest",
    }
    course_resp = requests.post(f"{API_BASE}/courses", json=course_payload, headers=admin_headers, timeout=30)
    assert course_resp.status_code == 200, course_resp.text
    course = course_resp.json()
    assert course["name"] == course_payload["name"]
    state["course_id"] = course["id"]

    class_payload = {
        "academic_year": "2025/2026",
        "semester": "Ganjil",
        "course_id": state["course_id"],
        "name": "TEST-4A",
        "schedule": "Rabu 10.00",
    }
    class_resp = requests.post(f"{API_BASE}/classes", json=class_payload, headers=admin_headers, timeout=30)
    assert class_resp.status_code == 200, class_resp.text
    class_doc = class_resp.json()
    assert class_doc["course_id"] == state["course_id"]
    assert isinstance(class_doc["class_code"], str) and len(class_doc["class_code"]) > 0
    state["class_id"] = class_doc["id"]
    state["class_code"] = class_doc["class_code"]


# --- Student management: manual and Excel import ---
def test_admin_add_student_manually(admin_headers, state):
    payload = {
        "nim": f"TEST{uuid.uuid4().hex[:6]}",
        "name": "TEST Student Manual",
        "email": state["student_email"],
        "whatsapp": "628222222222",
        "class_id": state["class_id"],
        "status": "active",
        "password": state["student_password"],
    }
    response = requests.post(f"{API_BASE}/students", json=payload, headers=admin_headers, timeout=30)
    assert response.status_code == 200, response.text
    student = response.json()
    assert student["email"] == state["student_email"]
    assert state["class_id"] in student["class_ids"]


def test_admin_import_students_from_excel(admin_headers, state):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["nim", "nama", "email", "whatsapp", "password"])
    default_password = "ImportDefault123!"
    row_password = "ImportExcel123!"
    import_default_email = f"import1.{uuid.uuid4().hex[:6]}@demo.id"
    import_excel_email = f"import2.{uuid.uuid4().hex[:6]}@demo.id"
    sheet.append([f"TEST{uuid.uuid4().hex[:6]}", "Import One", import_default_email, "628111", ""])
    sheet.append([f"TEST{uuid.uuid4().hex[:6]}", "Import Two", import_excel_email, "628112", row_password])
    sheet.append(["", "Missing Nim", f"invalid.{uuid.uuid4().hex[:6]}@demo.id", "628222", "Skip123!"])
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)

    files = {
        "file": (
            "students.xlsx",
            stream.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    response = requests.post(
        f"{API_BASE}/classes/{state['class_id']}/students/import",
        headers=admin_headers,
        files=files,
        data={"default_password": default_password},
        timeout=30,
    )
    assert response.status_code == 200, response.text
    result = response.json()
    assert result["created"] >= 2
    assert result["skipped"] >= 1
    assert result["password_from_default"] >= 1
    assert result["password_from_column"] >= 1

    assert login_student(import_default_email, default_password)
    assert login_student(import_excel_email, row_password)


# --- Materials/discussion and assignment creation ---
def test_admin_create_material_and_comment(admin_headers, state):
    material_payload = {
        "class_id": state["class_id"],
        "title": "TEST Material",
        "description": "Material for discussion",
        "meeting": "Pertemuan 9",
        "file_url": "",
        "video_url": "",
        "is_active": True,
        "locked_until": "",
    }
    material_resp = requests.post(f"{API_BASE}/materials", json=material_payload, headers=admin_headers, timeout=30)
    assert material_resp.status_code == 200, material_resp.text
    material = material_resp.json()
    state["material_id"] = material["id"]
    assert material["title"] == "TEST Material"
    assert material["meeting"] == "Pertemuan 1"

    next_material_resp = requests.post(
        f"{API_BASE}/materials",
        json={**material_payload, "title": "TEST Material Lanjutan"},
        headers=admin_headers,
        timeout=30,
    )
    assert next_material_resp.status_code == 200, next_material_resp.text
    assert next_material_resp.json()["meeting"] == "Pertemuan 2"

    comment_payload = {"material_id": state["material_id"], "content": "Komentar uji otomatis", "parent_id": ""}
    comment_resp = requests.post(f"{API_BASE}/comments", json=comment_payload, headers=admin_headers, timeout=30)
    assert comment_resp.status_code == 200, comment_resp.text
    comment = comment_resp.json()
    assert comment["material_id"] == state["material_id"]

    list_resp = requests.get(f"{API_BASE}/materials/{state['material_id']}/comments", headers=admin_headers, timeout=30)
    assert list_resp.status_code == 200
    comments = list_resp.json()
    assert any(item["id"] == comment["id"] for item in comments)


def test_admin_create_assignment_with_rubric_and_practicum(admin_headers, state):
    payload = {
        "class_id": state["class_id"],
        "title": "TEST Assignment Practicum",
        "description": "Create API and upload report",
        "deadline": (datetime.now(timezone.utc) + timedelta(days=2)).isoformat(),
        "tolerance_hours": 6,
        "allowed_formats": ["pdf", "docx", "zip", "xlsx"],
        "rubric": [
            {"criterion": "Ketepatan", "weight": 40},
            {"criterion": "Kerapian", "weight": 30},
            {"criterion": "Kreativitas", "weight": 30},
        ],
        "assignment_type": "individu",
        "allow_revision": True,
        "is_active": True,
        "is_practicum": True,
        "practicum_goal": "Menguji endpoint",
        "practicum_tools": "FastAPI, MongoDB",
        "practicum_steps": ["Buat", "Uji", "Lapor"],
        "required_screenshot": True,
        "late_penalty_per_day": 5,
    }
    response = requests.post(f"{API_BASE}/assignments", json=payload, headers=admin_headers, timeout=30)
    assert response.status_code == 200, response.text
    assignment = response.json()
    state["assignment_id"] = assignment["id"]
    assert assignment["is_practicum"] is True
    assert len(assignment["rubric"]) == 3
    assert assignment["publish_status"] == "published"


def test_scheduled_assignment_hidden_until_publish_time(admin_headers, state):
    student_headers = login_student(state["student_email"], state["student_password"])
    before_progress = requests.get(f"{API_BASE}/progress", headers=student_headers, timeout=30)
    assert before_progress.status_code == 200, before_progress.text
    before_missing = before_progress.json()["progress"]["missing"]

    future_publish_at = (datetime.now(timezone.utc) + timedelta(days=1)).isoformat()
    payload = {
        "class_id": state["class_id"],
        "title": "TEST Scheduled Assignment",
        "description": "Should appear after publish time",
        "deadline": (datetime.now(timezone.utc) + timedelta(days=3)).isoformat(),
        "published_at": future_publish_at,
        "tolerance_hours": 0,
        "allowed_formats": ["pdf"],
        "rubric": [{"criterion": "Nilai", "weight": 100}],
        "assignment_type": "individu",
        "allow_revision": True,
        "is_active": True,
        "close_after_deadline": False,
    }
    response = requests.post(f"{API_BASE}/assignments", json=payload, headers=admin_headers, timeout=30)
    assert response.status_code == 200, response.text
    assignment = response.json()
    assert assignment["publish_status"] == "scheduled"

    admin_list = requests.get(f"{API_BASE}/assignments", headers=admin_headers, timeout=30)
    assert admin_list.status_code == 200, admin_list.text
    assert any(item["id"] == assignment["id"] for item in admin_list.json())

    student_list = requests.get(f"{API_BASE}/assignments", headers=student_headers, timeout=30)
    assert student_list.status_code == 200, student_list.text
    assert all(item["id"] != assignment["id"] for item in student_list.json())

    after_progress = requests.get(f"{API_BASE}/progress", headers=student_headers, timeout=30)
    assert after_progress.status_code == 200, after_progress.text
    assert after_progress.json()["progress"]["missing"] == before_missing

    submit_resp = requests.post(
        f"{API_BASE}/assignments/{assignment['id']}/submit",
        headers=student_headers,
        files={"file": ("laporan.pdf", b"%PDF-1.4 scheduled", "application/pdf")},
        data={"note": "Should be rejected"},
        timeout=30,
    )
    assert submit_resp.status_code == 403
    assert "belum tayang" in submit_resp.text.lower()


# --- Submission and grading workflow ---
def test_student_submit_assignment_stores_file_locally(state):
    student_headers = login_student(state["student_email"], state["student_password"])
    files = {"file": ("laporan.pdf", b"%PDF-1.4 test content", "application/pdf")}
    data = {"note": "Pengumpulan otomatis"}
    response = requests.post(
        f"{API_BASE}/assignments/{state['assignment_id']}/submit",
        headers=student_headers,
        files=files,
        data=data,
        timeout=30,
    )
    assert response.status_code == 200, response.text
    submission = response.json()
    state["submission_id"] = submission["id"]
    assert submission["file"]["upload_status"] == "stored_on_server"
    assert submission["file"]["file_url"].startswith("/api/files/")

    list_resp = requests.get(f"{API_BASE}/submissions", headers=student_headers, timeout=30)
    assert list_resp.status_code == 200
    items = list_resp.json()
    current = next((item for item in items if item["id"] == state["submission_id"]), None)
    assert current is not None
    assert current["file"]["upload_status"] == "stored_on_server"


def test_admin_grade_submission_weighted_rubric(admin_headers, state):
    payload = {
        "rubric_scores": [
            {"criterion": "Ketepatan", "weight": 40, "score": 85},
            {"criterion": "Kerapian", "weight": 30, "score": 90},
            {"criterion": "Kreativitas", "weight": 30, "score": 88},
        ],
        "feedback": "Bagus, lanjutkan.",
        "revision_note": "",
        "status": "Dinilai",
    }
    response = requests.post(
        f"{API_BASE}/submissions/{state['submission_id']}/grade",
        headers=admin_headers,
        json=payload,
        timeout=30,
    )
    assert response.status_code == 200, response.text
    graded = response.json()
    assert graded["status"] == "Dinilai"
    assert graded["feedback"] == "Bagus, lanjutkan."
    assert isinstance(graded["grade"], (int, float))


def test_student_can_see_grade_and_feedback(state):
    student_headers = login_student(state["student_email"], state["student_password"])
    response = requests.get(f"{API_BASE}/assignments", headers=student_headers, timeout=30)
    assert response.status_code == 200
    assignments = response.json()
    target = next((item for item in assignments if item["id"] == state["assignment_id"]), None)
    assert target is not None
    assert isinstance(target["my_submission"]["grade"], (int, float))
    assert target["my_submission"]["feedback"] == "Bagus, lanjutkan."


# --- Reminders, calendar, and report export ---
def test_reminder_send_and_logs_listing(admin_headers, state):
    payload = {
        "assignment_id": state["assignment_id"],
        "student_id": "",
        "reminder_type": "manual",
        "message": "Reminder test",
    }
    send_resp = requests.post(f"{API_BASE}/reminders/send", json=payload, headers=admin_headers, timeout=30)
    assert send_resp.status_code == 200, send_resp.text
    reminder = send_resp.json()
    assert reminder["status"] == "in_app"

    list_resp = requests.get(f"{API_BASE}/reminders", headers=admin_headers, timeout=30)
    assert list_resp.status_code == 200
    logs = list_resp.json()
    assert any(item["id"] == reminder["id"] for item in logs)


def test_calendar_contains_deadline_events_for_admin(admin_headers, state):
    response = requests.get(f"{API_BASE}/calendar", headers=admin_headers, timeout=30)
    assert response.status_code == 200
    events = response.json()
    assert any(event["id"] == state["assignment_id"] and event["type"] == "deadline" for event in events)


def test_export_grades_excel_is_valid_xlsx(admin_headers):
    response = requests.get(f"{API_BASE}/reports/grades.xlsx", headers=admin_headers, timeout=30)
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    assert sheet.title == "Rekap Nilai"
    assert sheet["A1"].value == "NIM"
    assert sheet["B1"].value == "Nama"
